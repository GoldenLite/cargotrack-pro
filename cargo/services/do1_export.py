"""Генерация ДО1 (опись товаров партии) в формате Excel.

Формат соответствует шаблону ДО1: шапка с реквизитами получателя и общим
итогом, затем построчная опись HAWBGood'ов с группировкой по HAWB.
"""
from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


# Стили (как в образце ДО1: Calibri 11/10pt, тонкие рамки, выравн. left/center)
_FONT_HEADER = Font(name='Calibri', size=11)
_FONT_TABLE_HEAD = Font(name='Calibri', size=10)
_FONT_TABLE_BODY = Font(name='Calibri', size=11)
_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
_ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
_ALIGN_LEFT_TABLE_HEAD = Alignment(horizontal='left', vertical='center', wrap_text=True)
_THIN = Side(style='thin')
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Колонки таблицы и их соответствие данным
COLUMN_HEADERS = [
    '№',                                                     # A
    'HAWB, идентификационный номер',                         # B
    'Описание (содержание) груза',                           # C
    'Количество мест, шт',                                   # D
    'Вес товара, кг',                                        # E
    'Код по ТНВЭД',                                          # F
    'Стоимость товара в валюте, указанной в коммерческом '
    'или сопроводительном документе для перемещения '
    'международной отправки',                                # G
    'Стоимость как валюта',                                  # H
    'Получатель по HAWB, идентификационный номер',           # I
]

COLUMN_WIDTHS = {
    'A': 6,
    'B': 20.18,
    'C': 30,
    'D': 8,
    'E': 9,
    'F': 14,
    'G': 14,
    'H': 8,
    'I': 45.18,
}


def _decimal_to_excel(value):
    """Decimal → float для openpyxl, иначе None."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def build_do1_workbook(*, cargo, goods, organization):
    """Сгенерировать workbook ДО1 для партии cargo с товарами goods.

    Args:
        cargo: Cargo (MAWB)
        goods: list[HAWBGood] — товары всех HAWB партии, отсортированы.
        organization: OrganizationSettings — реквизиты получателя.

    Returns:
        openpyxl.Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'ДО1'

    # ── Ширины колонок ──
    for letter, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    # ── Шапка ──
    org_name = organization.name if organization and organization.name else ''
    org_inn = f'ИНН:{organization.inn}' if organization and organization.inn else ''

    # Суммы по партии: количество мест и вес
    pieces_total = 0
    weight_total = Decimal('0')
    seen_hawbs: set[int] = set()
    for g in goods:
        # места считаем по HAWB (только при первом появлении HAWB),
        # вес — по каждому товару (берём weight_net, fallback weight_gross)
        if g.hawb_id not in seen_hawbs:
            seen_hawbs.add(g.hawb_id)
            pieces_total += int(g.hawb.pieces_declared or 0)
        w = g.weight_net if g.weight_net is not None else g.weight_gross
        if w is not None:
            weight_total += Decimal(str(w))

    # Если в партии нет товаров с весом — берём вес самой партии
    if weight_total == 0 and cargo.weight is not None:
        weight_total = Decimal(str(cargo.weight))
    if pieces_total == 0 and cargo.pieces_declared is not None:
        pieces_total = int(cargo.pieces_declared)

    header_rows = [
        ('Получатель:', org_name, org_inn),
        ('Идентификатор:', cargo.awb_number or '', cargo.flight_date),
        ('Общий мест по AWB:', pieces_total, ''),
        ('Общий вес по AWB:', float(weight_total) if weight_total else '', ''),
    ]
    for i, (a, b, c) in enumerate(header_rows, start=1):
        ws.cell(row=i, column=1, value=a).font = _FONT_HEADER
        ws.cell(row=i, column=1).alignment = _ALIGN_LEFT
        ws.cell(row=i, column=2, value=b).font = _FONT_HEADER
        ws.cell(row=i, column=2).alignment = _ALIGN_LEFT
        if c not in ('', None):
            ws.cell(row=i, column=3, value=c).font = _FONT_HEADER
            ws.cell(row=i, column=3).alignment = _ALIGN_LEFT

    # ── Заголовки таблицы (строка 5) ──
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = _FONT_TABLE_HEAD
        cell.alignment = _ALIGN_LEFT_TABLE_HEAD
        cell.border = _BORDER_ALL

    # ── Нумерация колонок (строка 6: 1..9) ──
    for col_idx in range(1, len(COLUMN_HEADERS) + 1):
        cell = ws.cell(row=6, column=col_idx, value=col_idx)
        cell.font = _FONT_TABLE_BODY
        cell.alignment = _ALIGN_LEFT
        cell.border = _BORDER_ALL

    # ── Данные ──
    row = 7
    seq = 0
    seen_hawbs_in_table: set[int] = set()
    for g in goods:
        seq += 1
        is_first_in_hawb = g.hawb_id not in seen_hawbs_in_table
        seen_hawbs_in_table.add(g.hawb_id)

        weight = g.weight_net if g.weight_net is not None else g.weight_gross
        pieces = int(g.hawb.pieces_declared or 0) if is_first_in_hawb else None
        consignee = g.hawb.consignee_name or ''

        cells = [
            seq,                                # A: №
            g.hawb.hawb_number,                 # B: HAWB
            g.name or '',                       # C: описание
            pieces,                             # D: места (только в первой строке HAWB)
            _decimal_to_excel(weight),          # E: вес
            g.tnved_code or '',                 # F: ТНВЭД
            _decimal_to_excel(g.total_value),   # G: стоимость
            g.currency or '',                   # H: валюта
            consignee,                          # I: получатель HAWB
        ]
        for col_idx, value in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _FONT_TABLE_BODY
            # Описание и Получатель — с переносом текста (длинный текст)
            cell.alignment = _ALIGN_LEFT_WRAP if col_idx in (3, 9) else _ALIGN_LEFT
            cell.border = _BORDER_ALL
        row += 1

    return wb
