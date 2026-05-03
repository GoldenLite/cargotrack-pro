"""Генерация Грузового манифеста (Таиланд) — опись товаров партии в формате Excel.

Формат соответствует шаблону: шапка с реквизитами компании-перевозчика,
заголовок «Грузовой манифест», подзаголовок «Товары, доставляемые перевозчиком»,
таблица из 10 колонок (по строке на каждый товар), подвал с итогами.
"""
from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


# ── Стили (как в образце) ──
_FONT_HEADER     = Font(name='Calibri', size=10)              # реквизиты в шапке
_FONT_TITLE      = Font(name='Calibri', size=22, bold=True)   # «Грузовой манифест»
_FONT_SUBTITLE   = Font(name='Calibri', size=10, bold=True)   # «Товары, доставляемые перевозчиком»
_FONT_TABLE_HEAD = Font(name='Calibri', size=10, bold=True)
_FONT_TABLE_BODY = Font(name='Calibri', size=10)
_FONT_FOOTER     = Font(name='Calibri', size=10, bold=True)

_ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')

_THIN = Side(style='thin')
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Заголовки колонок
COLUMN_HEADERS = [
    '№',                    # A
    'Грузополучатель ФИО',  # B
    'Грузоотправитель',     # C
    'Номер накладной',      # D
    'Наименование',         # E
    'Код ТН ВЭД',           # F
    'Кол-во во вложении',   # G
    'Вес посылки',          # H
    'Стоимость',            # I
    'Валюта',               # J
]

# Ширины как в образце
COLUMN_WIDTHS = {
    'A': 4.7,  'B': 28.0, 'C': 20.7, 'D': 15.5, 'E': 28.3,
    'F': 10.3, 'G': 6.3,  'H': 10.1, 'I': 14.8, 'J': 9.5,
}


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    return v


def build_manifest_workbook(*, cargo, goods, organization):
    """Сгенерировать workbook «Грузовой манифест» для партии.

    Args:
        cargo: Cargo (MAWB).
        goods: list[HAWBGood] — товары партии, отсортированы (HAWB.hawb_number, id).
        organization: OrganizationSettings — реквизиты компании.

    Returns:
        openpyxl.Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'ГМ'

    # ── Ширины колонок ──
    for letter, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    # ── Шапка: реквизиты в столбце E (E1:J7) ──
    org = organization
    header_lines = [
        org.name or '',
        f'ОГРН {org.ogrn}' if org.ogrn else '',
        f'ИНН {org.inn}' if org.inn else '',
        f'Р/с {org.bank_account}' if org.bank_account else '',
        org.bank_name or '',
        f'К/с {org.bank_corr_account}' if org.bank_corr_account else '',
        f'БИК {org.bank_bik}' if org.bank_bik else '',
    ]
    for i, text in enumerate(header_lines, start=1):
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=10)
        cell = ws.cell(row=i, column=5, value=text)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_LEFT

    # ── Заголовок документа (строка 9, A9:J9) ──
    ws.merge_cells('A9:J9')
    title = ws.cell(row=9, column=1, value='Грузовой манифест')
    title.font = _FONT_TITLE
    title.alignment = _ALIGN_CENTER
    ws.row_dimensions[9].height = 28.5

    # ── Подзаголовок (строка 10, E10:J10) ──
    ws.merge_cells('E10:J10')
    subtitle = ws.cell(row=10, column=5, value='Товары, доставляемые перевозчиком')
    subtitle.font = _FONT_SUBTITLE
    subtitle.alignment = _ALIGN_LEFT

    # ── Заголовки таблицы (строка 11) ──
    ws.row_dimensions[11].height = 52.0
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = _FONT_TABLE_HEAD
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_ALL

    # ── Данные ──
    # Грузоотправитель: берём с HAWB (поле shipper_name) — это источник правды для каждой накладной.
    row = 12
    seq = 0
    for g in goods:
        seq += 1
        weight = g.weight_net if g.weight_net is not None else g.weight_gross
        shipper = g.hawb.shipper_name or ''
        cells = [
            seq,                                # A
            g.hawb.consignee_name or '',        # B Грузополучатель ФИО
            shipper,                            # C Грузоотправитель
            g.hawb.hawb_number,                 # D Номер накладной
            g.name or '',                       # E Наименование
            g.tnved_code or '',                 # F Код ТНВЭД
            _to_float(g.quantity),              # G Кол-во во вложении
            _to_float(weight),                  # H Вес
            _to_float(g.total_value),           # I Стоимость
            g.currency or '',                   # J Валюта
        ]
        for col_idx, value in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _FONT_TABLE_BODY
            # B/C/E — текст, остальное — центр
            cell.alignment = _ALIGN_LEFT if col_idx in (2, 3, 5) else _ALIGN_CENTER
            cell.border = _BORDER_ALL
        row += 1

    last_data_row = row - 1
    if last_data_row < 12:
        last_data_row = 12  # на случай пустой партии — иначе формула SUM сломается

    # ── Подвал: итоги ──
    # Строка last_data_row+1: «Итого количество мест» | сумма pieces | sum(H) | sum(I) | валюта
    total_pieces = sum(int(h.pieces_declared or 0) for h in {g.hawb_id: g.hawb for g in goods}.values())
    unique_hawbs = len({g.hawb_id for g in goods})
    total_goods = len(goods)
    total_currency = ''
    if goods:
        # Берём наиболее частую валюту
        from collections import Counter
        counts = Counter(g.currency for g in goods if g.currency)
        if counts:
            total_currency = counts.most_common(1)[0][0]

    footer_row1 = last_data_row + 1
    ws.merge_cells(start_row=footer_row1, start_column=2, end_row=footer_row1, end_column=3)
    ws.cell(row=footer_row1, column=2, value='Итого количество мест').font = _FONT_FOOTER
    ws.cell(row=footer_row1, column=2).alignment = _ALIGN_LEFT
    ws.cell(row=footer_row1, column=4, value=total_pieces).font = _FONT_FOOTER
    ws.cell(row=footer_row1, column=4).alignment = _ALIGN_CENTER
    # Формулы — пользователю удобнее: можно отредактировать ячейки и сумма пересчитается
    if last_data_row >= 12:
        ws.cell(row=footer_row1, column=8, value=f'=SUM(H12:H{last_data_row})').font = _FONT_FOOTER
        ws.cell(row=footer_row1, column=8).alignment = _ALIGN_CENTER
        ws.cell(row=footer_row1, column=9, value=f'=SUM(I12:I{last_data_row})').font = _FONT_FOOTER
        ws.cell(row=footer_row1, column=9).alignment = _ALIGN_CENTER
    if total_currency:
        ws.cell(row=footer_row1, column=10, value=total_currency).font = _FONT_FOOTER
        ws.cell(row=footer_row1, column=10).alignment = _ALIGN_CENTER

    # Строка footer_row1+1: «Итого количество экспресс-грузов (посылок)» | unique HAWB | МП
    footer_row2 = footer_row1 + 1
    ws.merge_cells(start_row=footer_row2, start_column=2, end_row=footer_row2, end_column=3)
    ws.cell(row=footer_row2, column=2, value='Итого количество экспресс-грузов (посылок)').font = _FONT_FOOTER
    ws.cell(row=footer_row2, column=2).alignment = _ALIGN_LEFT
    ws.cell(row=footer_row2, column=4, value=unique_hawbs).font = _FONT_FOOTER
    ws.cell(row=footer_row2, column=4).alignment = _ALIGN_CENTER
    ws.cell(row=footer_row2, column=5, value='МП').font = _FONT_FOOTER
    ws.cell(row=footer_row2, column=5).alignment = _ALIGN_LEFT

    # Строка footer_row2+1: «Количество товаров» | total_goods
    footer_row3 = footer_row2 + 1
    ws.merge_cells(start_row=footer_row3, start_column=2, end_row=footer_row3, end_column=3)
    ws.cell(row=footer_row3, column=2, value='Количество товаров').font = _FONT_FOOTER
    ws.cell(row=footer_row3, column=2).alignment = _ALIGN_LEFT
    ws.cell(row=footer_row3, column=4, value=total_goods).font = _FONT_FOOTER
    ws.cell(row=footer_row3, column=4).alignment = _ALIGN_CENTER

    return wb
