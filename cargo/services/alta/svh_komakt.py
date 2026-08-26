"""Авто-комакт (коммерческий акт) для СВХ Внуково.

Собирает per-HAWB Excel-разрез товаров поданной ДТЭГ/ДТ, АГРЕГИРОВАННЫЙ по коду
ТН ВЭД (наименования через запятую, вес/стоимость суммируются), в формате,
который принимает СВХ. Заменяет ручную работу декларанта.

Данные берём из нашей БД (outbox подачи), парсер ДТЭГ переиспользуем из
dteg_reestr (parse_express_cargo_declaration/select_house_shipment) — там уже
есть per-позиция tnved/description/gross_kg/value/currency.

Только для нашего СВХ Внуково (лицензия начинается с 10001), ИМПОРТ.
См. [[svh-komakt-automation]].
"""
from __future__ import annotations

import io
import xml.etree.ElementTree as ET
from collections import OrderedDict
from decimal import Decimal, InvalidOperation

VNUKOVO_LICENSE_PREFIX = '10001'

# Тег ESADout_CUPresentedDocument с ИНД. НАКЛАДНОЙ — так HAWB привязан к товару
# в регулярной ДТ (в ДТЭГ он в HouseShipment).
_HAWB_DOC_NAME = 'ИНДИВИДУАЛЬНАЯ НАКЛАДНАЯ'


def _ln(tag: str) -> str:
    return tag.rsplit('}', 1)[-1]


def _direct(el, name):
    """Прямые дети с локальным именем name (без глубоких потомков)."""
    return [c for c in list(el) if _ln(c.tag) == name]


def _direct_text(el, name: str) -> str:
    for c in list(el):
        if _ln(c.tag) == name and c.text:
            return c.text.strip()
    return ''


def _deep_text(el, name: str) -> str:
    for e in el.iter():
        if e is not el and _ln(e.tag) == name and e.text:
            return e.text.strip()
    return ''


def _deep_first(el, name):
    for e in el.iter():
        if e is not el and _ln(e.tag) == name:
            return e
    return None

# Формат комакта = формат ДО1 (столбцы совпадают с ручным файлом).
HEADERS = [
    '№№',
    'HAWB, индивидуальная накладная',
    'Природа (характер) груза',
    'Количество мест, шт',
    'Вес брутто, кг',
    'Код ТН ВЭД',
    'Стоимость товара в валюте, указанной в транспортных или коммерческих документах',
    'Буквенный код валюты',
    'Получатель по HAWB, индивидуальной накладной',
]
SUBNUM = ['1', '2', '3', '5', '6', '7', '8', '9', '10']


def _dec(s) -> Decimal:
    if s is None:
        return Decimal('0')
    try:
        return Decimal(str(s).replace(',', '.').strip() or '0')
    except (InvalidOperation, ValueError):
        return Decimal('0')


def is_vnukovo_import(hawb) -> bool:
    """HAWB оформляется на нашем СВХ Внуково и это импорт."""
    c = hawb.mawb if getattr(hawb, 'mawb_id', None) else None
    if c is None:
        return False
    lic = (c.warehouse_license or '').strip()
    if not lic.startswith(VNUKOVO_LICENSE_PREFIX):
        return False
    return (hawb.shipment_type or 'IMPORT').upper() != 'EXPORT'


def _find_dt_submission(hawb_number: str) -> str:
    """raw_xml подачи РЕГУЛЯРНОЙ ДТ (ESADout_CU) по накладной, или ''."""
    from cargo.models import AltaOutboxObservation, AltaOutboxWaybill
    obs_ids = list(AltaOutboxWaybill.objects
                   .filter(hawb_number=hawb_number)
                   .values_list('observation_id', flat=True))
    qs = (AltaOutboxObservation.objects
          .filter(id__in=obs_ids, msg_type__in=['CMN.11023', 'CMN.11335'])
          .order_by('-prepared_at'))
    for o in qs:
        rx = (o.parsed_meta or {}).get('raw_xml') or ''
        if 'ESADout_CU' in rx and 'ExpressCargoDeclaration' not in rx:
            return rx
    return ''


def _parse_regular_dt(rx: str, hawb_number: str):
    """(consignee_name, positions|None, reason) для регулярной ДТ (ESADout_CU).

    Товар — ESADout_CUGoods; вес — GrossWeightQuantity (прямой текст), стоимость —
    InvoicedCost (валюта ContractCurrencyCode из шапки), tnved — GoodsTNVEDCode,
    описание — прямые GoodsDescription (склеиваем). HAWB товара —
    ESADout_CUPresentedDocument с PrDocumentName='ИНДИВИДУАЛЬНАЯ НАКЛАДНАЯ'."""
    try:
        root = ET.fromstring(rx.encode('utf-8') if isinstance(rx, str) else rx)
    except ET.ParseError:
        return '', None, 'raw_xml ДТ не распарсился'
    # Получатель: ESADout_CUConsignee.OrganizationName; если пусто (часто
    # EqualIndicator=true — получатель==декларант) → ESADout_CUDeclarant.
    consignee = ''
    cons = _deep_first(root, 'ESADout_CUConsignee')
    if cons is not None:
        consignee = _deep_text(cons, 'OrganizationName')
    if not consignee:
        decl = _deep_first(root, 'ESADout_CUDeclarant')
        if decl is not None:
            consignee = _deep_text(decl, 'OrganizationName')
    currency = _deep_text(root, 'ContractCurrencyCode')

    positions = []
    for g in root.iter():
        if _ln(g.tag) != 'ESADout_CUGoods':
            continue
        gh = ''
        for pd in _direct(g, 'ESADout_CUPresentedDocument'):
            if _direct_text(pd, 'PrDocumentName').strip().upper() == _HAWB_DOC_NAME:
                gh = _direct_text(pd, 'PrDocumentNumber')
                break
        if gh and gh != hawb_number:
            continue
        descr = ' '.join(c.text.strip() for c in _direct(g, 'GoodsDescription')
                         if c.text and c.text.strip())
        positions.append({
            'tnved': _direct_text(g, 'GoodsTNVEDCode'),
            'description': descr,
            'gross_kg': _dec(_direct_text(g, 'GrossWeightQuantity')),
            'value': _dec(_direct_text(g, 'InvoicedCost')),
            'currency': currency,
        })
    if not positions:
        return consignee, None, 'в подаче ДТ нет товарных позиций по накладной'
    return consignee, positions, ''


def extract_dt_positions(hawb_number: str):
    """(consignee_name, positions|None, reason). positions: list of
    {tnved, description, gross_kg, value, currency}.

    Сначала ДТЭГ (ExpressCargoDeclaration, штатный парсер), затем регулярная ДТ
    (ESADout_CU)."""
    from cargo.services.alta.dteg_reestr import (
        find_submission_for_hawb, parse_express_cargo_declaration,
        select_house_shipment)

    rx, _mt, _obs = find_submission_for_hawb(hawb_number)
    if rx:
        parsed = parse_express_cargo_declaration(rx)
        if parsed:
            hs = select_house_shipment(parsed, hawb_number)
            if hs is None:
                return '', None, 'HouseShipment по накладной не найден в подаче ДТЭГ'
            consignee = (hs.get('consignee') or {}).get('name') or ''
            positions = [{
                'tnved': (g.get('tnved') or '').strip(),
                'description': (g.get('description') or '').strip(),
                'gross_kg': _dec(g.get('gross_kg')),
                'value': _dec(g.get('value')),
                'currency': (g.get('currency') or '').strip(),
            } for g in (hs.get('goods') or [])]
            if not positions:
                return '', None, 'в подаче ДТЭГ нет товарных позиций'
            return consignee, positions, ''

    dt_rx = _find_dt_submission(hawb_number)
    if dt_rx:
        return _parse_regular_dt(dt_rx, hawb_number)
    return '', None, 'подача (ДТЭГ/ДТ) не найдена в БД'


def aggregate_by_tnved(positions):
    """GROUP BY код ТН ВЭД: наименования через запятую, Σ вес, Σ стоимость.
    Порядок групп = первого появления кода."""
    groups: 'OrderedDict[str, dict]' = OrderedDict()
    for p in positions:
        k = p['tnved']
        g = groups.get(k)
        if g is None:
            g = groups[k] = {'tnved': k, 'descs': [], 'gross_kg': Decimal('0'),
                             'value': Decimal('0'), 'currency': p['currency'], 'n': 0}
        if p['description']:
            g['descs'].append(p['description'])
        g['gross_kg'] += p['gross_kg']
        g['value'] += p['value']
        g['n'] += 1
        if p['currency']:
            g['currency'] = p['currency']
    return list(groups.values())


def _fnum(d: Decimal) -> float:
    return float(d)


def build_komakt_xlsx(hawb_number: str, place_count=None):
    """→ (xlsx_bytes, meta) | (None, {'reason': ...}).

    meta: {consignee, num_positions, num_codes, total_gross_kg, total_value,
    currency, place_count}."""
    from cargo.models import HouseWaybill
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    h = HouseWaybill.objects.filter(hawb_number=hawb_number).order_by('-id').first()
    if not h:
        return None, {'reason': 'HAWB не найдена в БД'}
    if place_count is None:
        place_count = h.svh_do1_place_count
    mawb = (h.mawb.awb_number if getattr(h, 'mawb_id', None) and h.mawb else '')

    consignee, positions, reason = extract_dt_positions(hawb_number)
    if positions is None:
        return None, {'reason': reason}

    groups = aggregate_by_tnved(positions)

    wb = Workbook()
    ws = wb.active
    ws.title = 'комакт'
    thin = Side(style='thin')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    ws.append(SUBNUM)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = bord
    for c in ws[2]:
        c.border = bord

    tot_w = Decimal('0')
    tot_v = Decimal('0')
    curr = ''
    for i, g in enumerate(groups, start=1):
        row = [
            i,
            hawb_number,
            ', '.join(g['descs']),
            (place_count if (i == 1 and place_count is not None) else ''),
            _fnum(g['gross_kg']),
            g['tnved'],
            _fnum(g['value']),
            g['currency'],
            consignee,
        ]
        ws.append(row)
        for c in ws[ws.max_row]:
            c.border = bord
            c.alignment = Alignment(wrap_text=True, vertical='top')
        tot_w += g['gross_kg']
        tot_v += g['value']
        curr = g['currency'] or curr

    ws.append(['', '', 'итого',
               (place_count if place_count is not None else ''),
               _fnum(tot_w), '', _fnum(tot_v), curr, ''])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.border = bord

    for i, w in enumerate([5, 16, 60, 12, 11, 14, 16, 8, 32], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    meta = {
        'mawb': mawb,
        'consignee': consignee,
        'num_positions': len(positions),
        'num_codes': len(groups),
        'total_gross_kg': tot_w,
        'total_value': tot_v,
        'currency': curr,
        'place_count': place_count,
    }
    return buf.getvalue(), meta
