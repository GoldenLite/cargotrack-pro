"""
Представления CargoTrack Pro
"""
import io
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from django.contrib import messages
from django.views.decorators.http import require_POST

from .models import (
    Cargo, StatusHistory, Warehouse, CargoAssignment, Label,
    ROLE_CHOICES, STAGE_CHOICES, TRANSPORT_MODE_CHOICES,
    DocumentType, CargoCategoryDocRule, HAWBChecklistItem,
)

logger = logging.getLogger('cargo')

STAGE_COLORS = {
    'DRAFT':      'secondary',
    'FORMED':     'info',
    'DISPATCHED': 'primary',
    'ARRIVED':    'warning',
    'CUSTOMS':    'warning',
    'RELEASED':   'success',
}


@login_required
def dashboard(request):
    qs = Cargo.objects.select_related('warehouse').all()
    q = request.GET.get('q', '').strip()
    filter_stage = request.GET.get('stage', request.GET.get('queue', ''))
    filter_date_from = request.GET.get('date_from', '')
    filter_date_to = request.GET.get('date_to', '')
    filter_warehouse = request.GET.get('warehouse', '')

    if q:
        qs = qs.filter(
            Q(awb_number__icontains=q) | Q(customs_declaration_number__icontains=q) |
            Q(description__icontains=q) | Q(description_ru__icontains=q) |
            Q(assignments__user__username__icontains=q) |
            Q(assignments__user__last_name__icontains=q) |
            Q(assignments__user__first_name__icontains=q)
        ).distinct()
    if filter_stage:
        qs = qs.filter(stage=filter_stage)
    if filter_date_from:
        qs = qs.filter(flight_date__gte=filter_date_from)
    if filter_date_to:
        qs = qs.filter(flight_date__lte=filter_date_to)
    if filter_warehouse:
        qs = qs.filter(warehouse_id=filter_warehouse)

    total = qs.count()
    total_weight = qs.aggregate(s=Sum('weight'))['s'] or 0
    total_pieces = qs.aggregate(s=Sum('pieces_declared'))['s'] or 0
    problematic = [c for c in qs if c.is_problematic]

    # Канбан по этапам — один запрос вместо N штук на N этапов
    kanban_stages = [s[0] for s in STAGE_CHOICES]
    kanban_qs = qs.filter(stage__in=kanban_stages)
    kanban = {s: [] for s in kanban_stages}
    for c in kanban_qs:
        if c.stage in kanban and len(kanban[c.stage]) < 50:
            kanban[c.stage].append(c)

    stage_stats = qs.values('stage').annotate(cnt=Count('id')).order_by('stage')
    wh_stats = (
        qs.exclude(warehouse_license='')
          .values('warehouse_name', 'warehouse_license')
          .annotate(cnt=Count('id'), total_w=Sum('weight'))
          .order_by('-cnt')
    )

    from .widget_columns import (
        get_column_catalog, get_default_columns, serialize_column,
    )
    table_columns_cargo = [serialize_column(c) for c in get_column_catalog('cargo')]
    table_columns_hawb  = [serialize_column(c) for c in get_column_catalog('hawb')]

    context = {
        'cargos': qs, 'total': total, 'total_weight': total_weight,
        'total_pieces': total_pieces, 'problematic': problematic[:20],
        'kanban': kanban, 'kanban_stages': kanban_stages,
        'stage_choices': STAGE_CHOICES, 'stage_labels': dict(STAGE_CHOICES),
        'stage_stats': stage_stats,
        'wh_stats': wh_stats,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'stage_colors': STAGE_COLORS,
        'q': q, 'filter_stage': filter_stage,
        'filter_date_from': filter_date_from, 'filter_date_to': filter_date_to,
        'filter_warehouse': filter_warehouse,
        'table_columns_cargo': table_columns_cargo,
        'table_columns_hawb':  table_columns_hawb,
        'table_default_cols_cargo': get_default_columns('cargo'),
        'table_default_cols_hawb':  get_default_columns('hawb'),
    }
    return render(request, 'cargo/dashboard_pro.html', context)


@login_required
def cargo_detail(request, awb_number: str):
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    history = cargo.status_history.select_related('changed_by').all()
    history_list = list(history)
    history_with_duration = []
    for i, record in enumerate(history_list):
        hours = None
        if i < len(history_list) - 1:
            delta = history_list[i].changed_at - history_list[i + 1].changed_at
            hours = int(delta.total_seconds() // 3600)
        history_with_duration.append({'record': record, 'hours': hours})

    hawbs = cargo.hawbs.select_related('assigned_to').order_by('hawb_number')
    hawbs_list = list(hawbs)

    inbox_messages = (
        cargo.inbox_messages
        .order_by('-received_at')
        .only('id', 'received_at', 'msg_type', 'msg_kind', 'declaration_number',
              'status_applied', 'envelope_id')
    )

    context = {
        'cargo': cargo,
        'history': history_with_duration,
        'assignments': cargo.assignments.filter(is_active=True).select_related('user', 'assigned_by'),
        'all_users': User.objects.filter(is_active=True).order_by('last_name', 'first_name'),
        'role_choices': ROLE_CHOICES,
        'stage_choices': STAGE_CHOICES,
        'stage_colors': STAGE_COLORS,
        'hawbs': hawbs_list,
        'total_hawbs': len(hawbs_list),
        'docs_complete': sum(1 for h in hawbs_list if h.docs_ready),
        'released': sum(1 for h in hawbs_list if h.customs_status == 'RELEASED'),
        'inbox_messages': inbox_messages,
    }
    return render(request, 'cargo/detail_pro.html', context)


@login_required
def update_status(request, awb_number: str):
    """Совместимый endpoint: переводит партию на указанный этап.
    Принимает POST с параметром `stage` (старые параметры status/queue игнорируются).
    """
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    if request.method == 'POST':
        new_stage = request.POST.get('stage', '').strip()
        valid_stages = [s[0] for s in STAGE_CHOICES]
        if new_stage and new_stage in valid_stages:
            cargo.set_stage(new_stage, user=request.user)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'stage': cargo.stage})
        messages.success(request, f'Этап партии {awb_number} обновлён.')
        return redirect('cargo_detail', awb_number=awb_number)
    return redirect('cargo_detail', awb_number=awb_number)


@login_required
def assign_user(request, awb_number: str):
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'assign':
            user_id = request.POST.get('user_id')
            role = request.POST.get('role', 'declarant')
            note = request.POST.get('note', '')
            try:
                user = User.objects.get(pk=user_id)
                CargoAssignment.objects.filter(cargo=cargo, role=role, is_active=True).update(is_active=False)
                CargoAssignment.objects.create(
                    cargo=cargo, user=user, role=role,
                    assigned_by=request.user, note=note, is_active=True,
                )
                messages.success(request, f'{user.get_full_name() or user.username} назначен как {dict(ROLE_CHOICES).get(role, role)}')
            except User.DoesNotExist:
                messages.error(request, 'Пользователь не найден')
        elif action == 'remove':
            assignment_id = request.POST.get('assignment_id')
            CargoAssignment.objects.filter(pk=assignment_id, cargo=cargo).update(is_active=False)
            messages.success(request, 'Назначение снято')
    return redirect('cargo_detail', awb_number=awb_number)


@login_required
def cargo_list(request):
    qs = Cargo.objects.select_related('warehouse').prefetch_related('hawbs').all()
    q = request.GET.get('q', '').strip()
    filter_stage = request.GET.get('stage', request.GET.get('queue', ''))
    filter_unreleased_hawb = request.GET.get('unreleased_hawb', '')
    filter_transport = request.GET.get('transport', '')
    cql_raw = (request.GET.get('cql', '') or '').strip()
    cql_error = ''

    if cql_raw:
        try:
            from .cql_parser import parse_cql as _parse_cql, CQLError as _CQLErr
            qs = qs.filter(_parse_cql(cql_raw, {'me': request.user.username}, entity_type='cargo'))
        except Exception as _e:
            cql_error = str(_e)

    if q:
        qs = qs.filter(
            Q(awb_number__icontains=q) | Q(customs_declaration_number__icontains=q) |
            Q(description_ru__icontains=q)
        )
    if filter_stage:
        qs = qs.filter(stage=filter_stage)
    if filter_unreleased_hawb:
        qs = qs.exclude(hawbs__customs_status='RELEASED').distinct()
    if filter_transport:
        qs = qs.filter(transportation_mode=filter_transport)

    # Сортировка
    sort = request.GET.get('sort', '-created_at')
    direction = request.GET.get('dir', 'desc')
    allowed_sorts = ['awb_number', 'stage', 'departure_date',
                     'flight_date', 'weight', 'pieces_declared',
                     'warehouse_license', 'scan_into_bond', 'created_at']
    if sort in allowed_sorts:
        qs = qs.order_by(f'{"-" if direction == "desc" else ""}{sort}')

    # Аннотируем количество HAWB и веса
    from django.db.models import Count as DCount, Sum as DSum
    qs = qs.annotate(
        hawb_total=DCount('hawbs', distinct=True),
        hawb_released=DCount('hawbs', filter=Q(hawbs__customs_status='RELEASED'), distinct=True),
        hawb_weight_released=DSum('hawbs__weight', filter=Q(hawbs__customs_status='RELEASED')),
        hawb_weight_svh=DSum('hawbs__weight'),
    )

    # Добавляем типы грузов из HAWB (используем prefetch_related)
    cargos = list(qs[:300])
    for cargo in cargos:
        types = sorted({h.cargo_type for h in cargo.hawbs.all() if h.cargo_type})
        cargo.cargo_types_display = ', '.join(types)

    context = {
        'cargos': cargos,
        'stage_choices': STAGE_CHOICES,
        'stage_colors': STAGE_COLORS,
        'transport_choices': TRANSPORT_MODE_CHOICES,
        'q': q,
        'filter_stage': filter_stage,
        'filter_unreleased_hawb': filter_unreleased_hawb,
        'filter_transport': filter_transport,
        'sort': sort,
        'direction': direction,
        'cql': cql_raw,
        'cql_error': cql_error,
    }
    return render(request, 'cargo/cargo_list.html', context)


@login_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl не установлен', status=500)

    qs = Cargo.objects.select_related('warehouse').all()
    q = request.GET.get('q', '').strip()
    filter_stage = request.GET.get('stage', '')
    if q:
        qs = qs.filter(Q(awb_number__icontains=q) | Q(customs_declaration_number__icontains=q))
    if filter_stage:
        qs = qs.filter(stage=filter_stage)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Грузы'
    headers = [
        'AWB номер', 'Этап', 'Рейс', 'Дата рейса',
        'Откуда (IATA)', 'Куда (IATA)', 'Вес (кг)', 'Мест',
        'Склад', '№ Лицензии', 'Ячейка', 'Въезд на склад', 'Декларация ТД',
        'Дата подачи', 'Дата выпуска', 'Стоимость', 'Валюта', 'Пошлина (RUB)',
        'Описание', 'Назначено на', 'Создан',
    ]
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for cargo in qs:
        assigned = ', '.join(
            (a.user.get_full_name() or a.user.username)
            for a in cargo.assignments.filter(is_active=True).select_related('user')
        )
        ws.append([
            cargo.awb_number, cargo.get_stage_display(),
            cargo.flight_number,
            cargo.flight_date.strftime('%d.%m.%Y') if cargo.flight_date else '',
            cargo.departure_iata, cargo.arrival_iata,
            float(cargo.weight) if cargo.weight else '',
            cargo.pieces_declared,
            cargo.warehouse_name, cargo.warehouse_license, cargo.bond_location,
            cargo.scan_into_bond.strftime('%d.%m.%Y %H:%M') if cargo.scan_into_bond else '',
            cargo.customs_declaration_number,
            cargo.entry_date.strftime('%d.%m.%Y %H:%M') if cargo.entry_date else '',
            cargo.release_date.strftime('%d.%m.%Y %H:%M') if cargo.release_date else '',
            float(cargo.invoice_value) if cargo.invoice_value else '',
            cargo.invoice_currency,
            float(cargo.duty_amount) if cargo.duty_amount else '',
            cargo.description_ru or cargo.description, assigned,
            cargo.created_at.strftime('%d.%m.%Y %H:%M') if cargo.created_at else '',
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'cargotrack_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────── HAWB VIEWS ───────────────────────────

@login_required
def hawb_list(request, awb_number: str):
    """Список HAWB внутри MAWB — с сортировкой"""
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    hawbs = cargo.hawbs.select_related('assigned_to').all()

    # Сортировка
    sort = request.GET.get('sort', 'hawb_number')
    direction = request.GET.get('dir', 'asc')
    allowed_sorts = ['hawb_number', 'status', 'customs_status', 'weight', 'pieces_declared',
                     'customs_declaration_number',
                     'consignee_name', 'invoice_value']
    if sort in allowed_sorts:
        hawbs = hawbs.order_by(f'{"-" if direction == "desc" else ""}{sort}')

    # Статистика по документам
    total_hawbs = hawbs.count()
    docs_complete = sum(1 for h in hawbs if h.docs_ready)
    released = hawbs.filter(customs_status='RELEASED').count()

    # SLA
    from .sla import compute_sla_state, get_active_workflow_step
    hawbs = list(hawbs)
    for h in hawbs:
        step = get_active_workflow_step('hawb', h.pk)
        h.sla_logistics = compute_sla_state(
            'hawb', 'logistics_status', h.logistics_status, h.logistics_status_date,
            workflow_step=step,
        )
        h.sla_customs = compute_sla_state(
            'hawb', 'customs_status', h.customs_status, h.customs_status_date,
            workflow_step=step,
        )

    all_users = User.objects.filter(is_active=True).order_by('last_name', 'first_name')

    context = {
        'cargo': cargo,
        'hawbs': hawbs,
        'all_users': all_users,
        'total_hawbs': total_hawbs,
        'docs_complete': docs_complete,
        'released': released,
        'sort': sort,
        'direction': direction,
    }
    return render(request, 'cargo/hawb_list.html', context)


@login_required
def hawb_create(request, awb_number: str):
    """Создать HAWB внутри конкретного MAWB"""
    from .models import HouseWaybill
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    if request.method == 'POST':
        hawb_number = request.POST.get('hawb_number', '').strip()
        next_url = request.POST.get('next', '')
        def _back():
            return redirect(next_url) if next_url else redirect('cargo_detail', awb_number=awb_number)
        if not hawb_number:
            messages.error(request, 'Номер HAWB обязателен')
            return _back()
        if HouseWaybill.objects.filter(hawb_number=hawb_number).exists():
            messages.error(request, f'HAWB {hawb_number} уже существует')
            return _back()
        assigned_user = None
        assigned_id = request.POST.get('assigned_to')
        if assigned_id:
            try:
                assigned_user = User.objects.get(pk=assigned_id)
            except User.DoesNotExist:
                pass
        hawb = HouseWaybill.objects.create(
            mawb=cargo,
            hawb_number=hawb_number,
            description=request.POST.get('description', ''),
            cargo_type=request.POST.get('cargo_type', 'B2C'),
            shipment_type=request.POST.get('shipment_type', 'IMPORT'),
            consignee_name=request.POST.get('consignee_name', ''),
            consignee_city=request.POST.get('consignee_city', ''),
            consignee_phone=request.POST.get('consignee_phone', ''),
            consignee_inn=request.POST.get('consignee_inn', ''),
            shipper_name=request.POST.get('shipper_name', ''),
            shipper_inn=request.POST.get('shipper_inn', ''),
            shipper_city=request.POST.get('shipper_city', ''),
            shipper_address=request.POST.get('shipper_address', ''),
            shipper_phone=request.POST.get('shipper_phone', ''),
            weight=request.POST.get('weight') or None,
            pieces_declared=request.POST.get('pieces_declared') or 1,
            invoice_value=request.POST.get('invoice_value') or None,
            invoice_currency=request.POST.get('invoice_currency', 'USD'),
            logistics_status='AT_ORIGIN_WH',  # Авто-статус при создании внутри партии
            logistics_status_date=timezone.now(),
            last_status_change=timezone.now(),
            assigned_to=assigned_user,
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, f'HAWB {hawb_number} создан')
        return _back()
    return redirect('cargo_detail', awb_number=awb_number)


@login_required
def hawb_create_standalone(request):
    """Создать HAWB без привязки к MAWB (или с выбором MAWB)"""
    from .models import HouseWaybill
    if request.method == 'POST':
        hawb_number = request.POST.get('hawb_number', '').strip()
        if not hawb_number:
            messages.error(request, 'Номер HAWB обязателен')
            return redirect('hawb_create_standalone')
        if HouseWaybill.objects.filter(hawb_number=hawb_number).exists():
            messages.error(request, f'HAWB {hawb_number} уже существует')
            return redirect('hawb_create_standalone')

        # MAWB опциональный
        mawb_obj = None
        mawb_awb = request.POST.get('mawb_awb', '').strip()
        if mawb_awb:
            mawb_obj = Cargo.objects.filter(awb_number=mawb_awb).first()
            if not mawb_obj:
                messages.error(request, f'Партия MAWB {mawb_awb} не найдена')
                return redirect('hawb_create_standalone')
            # При привязке к партии — накладная автоматически получит AT_ORIGIN_WH
            # (см. create ниже)

        assigned_user = None
        assigned_id = request.POST.get('assigned_to')
        if assigned_id:
            try:
                assigned_user = User.objects.get(pk=assigned_id)
            except User.DoesNotExist:
                pass

        # Если привязываем к партии — ставим AT_ORIGIN_WH автоматически
        initial_logistics = 'AT_ORIGIN_WH' if mawb_obj else 'CREATED'

        hawb = HouseWaybill.objects.create(
            mawb=mawb_obj,  # None если не привязан
            hawb_number=hawb_number,
            description=request.POST.get('description', ''),
            cargo_type=request.POST.get('cargo_type', 'B2C'),
            shipment_type=request.POST.get('shipment_type', 'IMPORT'),
            consignee_name=request.POST.get('consignee_name', ''),
            consignee_address=request.POST.get('consignee_address', ''),
            consignee_city=request.POST.get('consignee_city', ''),
            consignee_phone=request.POST.get('consignee_phone', ''),
            consignee_inn=request.POST.get('consignee_inn', ''),
            shipper_name=request.POST.get('shipper_name', ''),
            shipper_inn=request.POST.get('shipper_inn', ''),
            shipper_city=request.POST.get('shipper_city', ''),
            shipper_address=request.POST.get('shipper_address', ''),
            shipper_phone=request.POST.get('shipper_phone', ''),
            weight=request.POST.get('weight') or None,
            pieces_declared=request.POST.get('pieces_declared') or 1,
            invoice_value=request.POST.get('invoice_value') or None,
            invoice_currency=request.POST.get('invoice_currency', 'USD'),
            logistics_status=initial_logistics,
            logistics_status_date=timezone.now() if mawb_obj else None,
            last_status_change=timezone.now(),
            assigned_to=assigned_user,
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, f'HAWB {hawb_number} создан')
        return redirect('hawb_detail', hawb_id=hawb.pk)

    # GET — форма создания
    context = {
        'all_users': User.objects.filter(is_active=True).order_by('last_name', 'first_name'),
        'mawb_list': Cargo.objects.values('awb_number', 'flight_number', 'flight_date').order_by('-created_at')[:100],
    }
    return render(request, 'cargo/hawb_create.html', context)


@login_required
def hawb_update(request, hawb_id: int):
    """Обновить статус / документы / ответственного HAWB"""
    from .models import HouseWaybill
    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)
    if request.method == 'POST':
        # Ответственный
        assigned_id = request.POST.get('assigned_to')
        if assigned_id:
            try:
                hawb.assigned_to = User.objects.get(pk=assigned_id)
            except User.DoesNotExist:
                pass

        # Таможенная декларация
        td = request.POST.get('customs_declaration_number', '').strip()
        if td:
            hawb.customs_declaration_number = td

        # Документы — чекбоксы
        hawb.doc_invoice      = 'doc_invoice'      in request.POST
        hawb.doc_packing_list = 'doc_packing_list' in request.POST
        hawb.doc_permit       = 'doc_permit'       in request.POST
        hawb.doc_tech_desc    = 'doc_tech_desc'    in request.POST

        # Прочие поля
        for field in ('pieces_declared', 'weight', 'notes', 'consignee_name'):
            val = request.POST.get(field, '').strip()
            if val:
                setattr(hawb, field, val)

        # Поля отправителя — сохраняем как есть, в т.ч. пустые (чтобы можно было очистить)
        for field in ('shipper_name', 'shipper_inn', 'shipper_city',
                      'shipper_address', 'shipper_phone'):
            if field in request.POST:
                setattr(hawb, field, request.POST.get(field, '').strip())

        hawb.save()
        messages.success(request, f'HAWB {hawb.hawb_number} обновлён')
    # Редирект: если есть MAWB — в список HAWB партии, иначе — в карточку HAWB
    if hawb.mawb:
        return redirect('hawb_list', awb_number=hawb.mawb.awb_number)
    return redirect('hawb_detail', hawb_id=hawb.pk)


@login_required
def hawb_delete(request, hawb_id: int):
    """Удалить HAWB"""
    from .models import HouseWaybill
    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)
    mawb_awb = hawb.mawb.awb_number if hawb.mawb else None
    if request.method == 'POST':
        hawb.delete()
        messages.success(request, f'HAWB {hawb.hawb_number} удалён')
    if mawb_awb:
        return redirect('hawb_list', awb_number=mawb_awb)
    return redirect('all_hawbs')


@login_required
def all_hawbs(request):
    """Глобальный список всех HAWB"""
    from django.db.models import F
    from .models import HouseWaybill
    qs = HouseWaybill.objects.select_related('mawb', 'mawb__warehouse', 'assigned_to').annotate(
        total_goods_count=Count('goods', distinct=True),
        approved_goods_count=Count('goods', filter=Q(goods__approval_status='approved'), distinct=True),
    )

    q = request.GET.get('q', '').strip()
    filter_logistics = request.GET.get('logistics_status', '')
    filter_type = request.GET.get('cargo_type', '')
    filter_ship = request.GET.get('shipment_type', '')
    filter_unreleased = request.GET.get('unreleased', '')
    cql_raw = (request.GET.get('cql', '') or '').strip()
    cql_error = ''

    if cql_raw:
        try:
            from .cql_parser import parse_cql as _parse_cql
            qs = qs.filter(_parse_cql(cql_raw, {'me': request.user.username}, entity_type='hawb'))
        except Exception as _e:
            cql_error = str(_e)

    if q:
        qs = qs.filter(
            Q(hawb_number__icontains=q) |
            Q(consignee_name__icontains=q) |
            Q(consignee_inn__icontains=q) |
            Q(customs_declaration_number__icontains=q) |
            Q(mawb__awb_number__icontains=q)
        )
    if filter_logistics:
        qs = qs.filter(logistics_status=filter_logistics)
    if filter_type:
        qs = qs.filter(cargo_type=filter_type)
    if filter_ship:
        qs = qs.filter(shipment_type=filter_ship)
    if filter_unreleased:
        qs = qs.exclude(customs_status='RELEASED')

    sort = request.GET.get('sort', 'hawb_number')
    direction = request.GET.get('dir', 'asc')
    allowed_sorts = [
        'hawb_number', 'cargo_type', 'shipment_type',
        'mawb__awb_number', 'mawb__flight_number', 'mawb__flight_date',
        'mawb__departure_iata', 'scan_into_bond', 'mawb__warehouse_license',
        'customs_declaration_number', 'weight', 'pieces_declared',
        'logistics_status', 'customs_status',
    ]
    if sort in allowed_sorts:
        prefix = '-' if direction == 'desc' else ''
        qs = qs.order_by(f'{prefix}{sort}')

    context = {
        'hawbs': qs[:500],
        'total': qs.count(),
        'logistics_choices': HouseWaybill.LOGISTICS_STATUS_CHOICES,
        'q': q,
        'filter_logistics': filter_logistics,
        'filter_type': filter_type,
        'filter_ship': filter_ship,
        'filter_unreleased': filter_unreleased,
        'sort': sort,
        'direction': direction,
        'cql': cql_raw,
        'cql_error': cql_error,
    }
    return render(request, 'cargo/all_hawbs.html', context)


@login_required
def faq(request):
    """Справочник этапов"""
    return render(request, 'cargo/faq.html', {
        'stage_choices': STAGE_CHOICES,
    })


# ─────────────────────────── HAWB DETAIL VIEWS ───────────────────────────

@login_required
def hawb_detail(request, hawb_id: int):
    """Карточка индивидуальной накладной — товары, документы, чеклист"""
    from .models import HouseWaybill, HAWBGood, HAWBDocument
    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)
    goods = hawb.goods.all()
    documents = hawb.documents.all()

    if request.method == 'POST':
        action = request.POST.get('action')

        # Обновление чеклиста документов и кол-ва требуемых
        if action == 'update_docs':
            hawb.doc_invoice      = 'doc_invoice'      in request.POST
            hawb.doc_packing_list = 'doc_packing_list' in request.POST
            hawb.doc_permit       = 'doc_permit'       in request.POST
            hawb.doc_tech_desc    = 'doc_tech_desc'    in request.POST
            req = request.POST.get('docs_required', '4')
            try:
                hawb.docs_required = max(1, min(10, int(req)))
            except ValueError:
                pass
            hawb.save()
            messages.success(request, 'Чеклист документов обновлён')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Добавить товарную позицию
        if action == 'add_good':
            name = request.POST.get('name', '').strip()
            if name:
                # Для B2C обязательна ссылка на товар — проверяем перед созданием
                product_url = request.POST.get('product_url', '').strip()
                good_cargo_type = request.POST.get('cargo_type', '').strip()
                effective_type = good_cargo_type or hawb.cargo_type
                if effective_type == 'B2C' and not product_url:
                    messages.error(request, 'Для B2C товаров обязательна ссылка на товар (URL)')
                    return redirect('hawb_detail', hawb_id=hawb_id)
                HAWBGood.objects.create(
                    hawb=hawb,
                    name=name,
                    tnved_code=request.POST.get('tnved_code', ''),
                    brand=request.POST.get('brand', ''),
                    manufacturer=request.POST.get('manufacturer', ''),
                    model=request.POST.get('model', ''),
                    article=request.POST.get('article', ''),
                    product_url=product_url,
                    quantity=request.POST.get('quantity') or 1,
                    unit=request.POST.get('unit', 'шт'),
                    quantity_additional=request.POST.get('quantity_additional') or None,
                    unit_additional=request.POST.get('unit_additional', ''),
                    weight_net=request.POST.get('weight_net') or None,
                    weight_gross=request.POST.get('weight_gross') or None,
                    unit_price=request.POST.get('unit_price') or None,
                    total_value=request.POST.get('total_value') or None,
                    currency=request.POST.get('currency', 'USD'),
                    cargo_type=good_cargo_type,
                )
                messages.success(request, 'Товарная позиция добавлена')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Удалить товарную позицию
        if action == 'delete_good':
            good_id = request.POST.get('good_id')
            HAWBGood.objects.filter(pk=good_id, hawb=hawb).delete()
            messages.success(request, 'Позиция удалена')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Обновить поле товарной позиции (inline-редактирование)
        if action == 'update_good':
            good_id = request.POST.get('good_id')
            field = request.POST.get('field', '').strip()
            value = request.POST.get('value', '').strip()
            allowed_fields = {
                'name', 'tnved_code', 'brand', 'manufacturer', 'model',
                'article', 'product_url',
                'quantity', 'unit', 'quantity_additional', 'unit_additional',
                'weight_net', 'weight_gross',
                'unit_price', 'total_value', 'currency', 'cargo_type',
            }
            numeric_fields = {
                'quantity', 'quantity_additional',
                'weight_net', 'weight_gross', 'unit_price', 'total_value',
            }
            good = HAWBGood.objects.filter(pk=good_id, hawb=hawb).first()
            if good and field in allowed_fields:
                if field in numeric_fields:
                    try:
                        parsed = float(value) if value else None
                    except ValueError:
                        return JsonResponse({'ok': False, 'error': 'Неверный формат числа'})
                    setattr(good, field, parsed)
                else:
                    setattr(good, field, value)
                good.save(update_fields=[field])
            return JsonResponse({'ok': True})

        # Добавить документ
        if action == 'add_document':
            doc_name = request.POST.get('doc_name', '').strip()
            if doc_name:
                HAWBDocument.objects.create(
                    hawb=hawb,
                    doc_type=request.POST.get('doc_type', 'other'),
                    name=doc_name,
                    number=request.POST.get('doc_number', ''),
                    issue_date=request.POST.get('issue_date') or None,
                    is_received='is_received' in request.POST,
                    notes=request.POST.get('doc_notes', ''),
                )
                messages.success(request, 'Документ добавлен')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Обновить документ (получен/не получен)
        if action == 'toggle_document':
            doc_id = request.POST.get('doc_id')
            from .models import HAWBDocument as D
            doc = D.objects.filter(pk=doc_id, hawb=hawb).first()
            if doc:
                doc.is_received = not doc.is_received
                doc.save()
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Удалить документ
        if action == 'delete_document':
            doc_id = request.POST.get('doc_id')
            HAWBDocument.objects.filter(pk=doc_id, hawb=hawb).delete()
            messages.success(request, 'Документ удалён')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Смена логистического статуса
        if action == 'change_logistics_status':
            from .models import HouseWaybill as HWB
            new_ls = request.POST.get('logistics_status', '').strip()
            valid_ls = [s[0] for s in HWB.LOGISTICS_STATUS_CHOICES]
            if new_ls in valid_ls:
                error = hawb.change_logistics_status(new_ls, user=request.user,
                                                     comment=request.POST.get('comment', ''))
                if error:
                    messages.error(request, error)
                else:
                    messages.success(request, f'Логистический статус обновлён: {hawb.logistics_status_display}')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # Смена таможенного статуса
        if action == 'change_customs_status':
            from .models import HouseWaybill as HWB
            new_cs = request.POST.get('customs_status', '').strip()
            valid_cs = [s[0] for s in HWB.CUSTOMS_STATUS_CHOICES]
            if new_cs in valid_cs:
                error = hawb.change_customs_status(new_cs, user=request.user)
                if error:
                    messages.error(request, error)
                else:
                    messages.success(request, f'Таможенный статус обновлён: {hawb.customs_status_label}')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # ── Чеклист: добавить документ ──────────────────────────────
        if action == 'checklist_add':
            doc_type_id = request.POST.get('document_type_id', '').strip()
            if doc_type_id and doc_type_id.isdigit():
                doc_type = DocumentType.objects.filter(pk=int(doc_type_id), is_active=True).first()
                if doc_type:
                    is_req = request.POST.get('is_required', 'true') != 'false'
                    HAWBChecklistItem.objects.get_or_create(
                        hawb=hawb,
                        document_type=doc_type,
                        defaults={
                            'is_required': is_req,
                            'added_by': request.user,
                        }
                    )
                    messages.success(request, f'Документ «{doc_type.name}» добавлен в чеклист')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # ── Чеклист: отметить получен / не получен ───────────────────
        if action == 'checklist_toggle':
            item_id = request.POST.get('item_id', '').strip()
            if item_id and item_id.isdigit():
                item = HAWBChecklistItem.objects.filter(pk=int(item_id), hawb=hawb).first()
                if item:
                    item.is_received = not item.is_received
                    item.save(update_fields=['is_received'])
            return redirect('hawb_detail', hawb_id=hawb_id)

        # ── Чеклист: изменить обязательность ─────────────────────────
        if action == 'checklist_toggle_required':
            item_id = request.POST.get('item_id', '').strip()
            if item_id and item_id.isdigit():
                item = HAWBChecklistItem.objects.filter(pk=int(item_id), hawb=hawb).first()
                if item:
                    item.is_required = not item.is_required
                    item.save(update_fields=['is_required'])
            return redirect('hawb_detail', hawb_id=hawb_id)

        # ── Чеклист: удалить элемент ──────────────────────────────────
        if action == 'checklist_remove':
            item_id = request.POST.get('item_id', '').strip()
            if item_id and item_id.isdigit():
                HAWBChecklistItem.objects.filter(pk=int(item_id), hawb=hawb).delete()
                messages.success(request, 'Документ удалён из чеклиста')
            return redirect('hawb_detail', hawb_id=hawb_id)

        # ── Чеклист: заполнить по шаблону категории ───────────────────
        if action == 'checklist_populate':
            rules = CargoCategoryDocRule.objects.filter(
                cargo_type=hawb.cargo_type
            ).select_related('document_type')
            added = 0
            for rule in rules:
                _, created = HAWBChecklistItem.objects.get_or_create(
                    hawb=hawb,
                    document_type=rule.document_type,
                    defaults={'is_required': True, 'added_by': request.user}
                )
                if created:
                    added += 1
            if added:
                messages.success(request, f'Добавлено {added} документов по шаблону категории {hawb.cargo_type}')
            else:
                messages.info(request, 'Все документы шаблона уже в чеклисте')
            return redirect('hawb_detail', hawb_id=hawb_id)

    # ── Авто-заполнение чеклиста при первом просмотре ─────────────────
    checklist = list(hawb.checklist_items.select_related('document_type').order_by('-is_required', 'document_type__name'))
    if not checklist:
        rules = CargoCategoryDocRule.objects.filter(
            cargo_type=hawb.cargo_type
        ).select_related('document_type')
        for rule in rules:
            HAWBChecklistItem.objects.get_or_create(
                hawb=hawb,
                document_type=rule.document_type,
                defaults={'is_required': True, 'added_by': None}
            )
        checklist = list(hawb.checklist_items.select_related('document_type').order_by('-is_required', 'document_type__name'))

    checklist_total    = len(checklist)
    checklist_received = sum(1 for i in checklist if i.is_received)
    checklist_required = sum(1 for i in checklist if i.is_required)
    checklist_ready    = (
        checklist_required == 0 or
        all(i.is_received for i in checklist if i.is_required)
    ) if checklist else False

    from .models import HouseWaybill as HWB, HawbWorkflowEvent, ImportedSheetRow, AltaInboxMessage
    workflow_events = (
        HawbWorkflowEvent.objects
        .filter(hawb=hawb)
        .select_related('set_by', 'source_row__source')
        .order_by('-occurred_at', '-id')
    )
    sheet_rows = list(
        ImportedSheetRow.objects
        .filter(Q(promoted_hawb=hawb) | Q(matched_hawb=hawb))
        .select_related('source')
        .order_by('source__kind', 'source_row_index')
        .distinct()
    )
    inbox_messages = list(
        AltaInboxMessage.objects
        .filter(hawb=hawb)
        .order_by('-received_at')[:50]
    )
    from django.conf import settings as _settings
    cdek_enabled = getattr(_settings, 'CDEK_ENABLED', False)
    cdek_events = list(hawb.cdek_events.all()[:50]) if cdek_enabled else []
    context = {
        'hawb': hawb,
        'goods': goods,
        'documents': documents,
        'cdek_enabled': cdek_enabled,
        'cdek_events': cdek_events,
        'checklist': checklist,
        'checklist_total': checklist_total,
        'checklist_received': checklist_received,
        'checklist_required': checklist_required,
        'checklist_ready': checklist_ready,
        'workflow_events': workflow_events,
        'sheet_rows': sheet_rows,
        'inbox_messages': inbox_messages,
        'logistics_status_choices': HWB.LOGISTICS_STATUS_CHOICES,
        'customs_status_choices': HWB.CUSTOMS_STATUS_CHOICES,
        'all_users': User.objects.filter(is_active=True).order_by('last_name', 'first_name'),
        'doc_type_choices': [
            ('invoice', 'Инвойс'), ('packing_list', 'Упаковочный лист'),
            ('permit', 'Разрешительный документ'), ('tech_desc', 'Техническое описание'),
            ('certificate', 'Сертификат'), ('declaration', 'Декларация'), ('other', 'Иное'),
        ],
        'currency_choices': [('USD','USD'),('EUR','EUR'),('CNY','CNY'),('RUB','RUB'),('GBP','GBP')],
    }
    return render(request, 'cargo/hawb_detail.html', context)


# ─────────────────────────── СОЗДАНИЕ ПАРТИИ ИЗ HAWB ───────────────────────

@login_required
def cargo_create_from_hawbs(request):
    """Шаг 1 — выбор HAWB для партии. Шаг 2 — данные партии."""
    from django.db.models import F as DjF
    from .models import HouseWaybill

    step = request.GET.get('step', '1')

    # ── ШАГ 1: выбор HAWB по фильтрам ──
    if step == '1' or request.method == 'GET':
        # Только накладные без партии, в статусе «Принят на склад отправки»
        # и у которых все товарные позиции согласованы (либо товаров нет)
        qs = HouseWaybill.objects.select_related('assigned_to').filter(
            mawb__isnull=True,
            logistics_status='AT_ORIGIN_WH',
        ).annotate(
            total_goods=Count('goods', distinct=True),
            approved_goods=Count('goods', filter=Q(goods__approval_status='approved'), distinct=True),
        ).filter(total_goods=DjF('approved_goods'))

        q = request.GET.get('q', '').strip()
        filter_type = request.GET.get('cargo_type', '')
        filter_ship = request.GET.get('shipment_type', '')

        if q:
            qs = qs.filter(
                Q(hawb_number__icontains=q) |
                Q(consignee_name__icontains=q) |
                Q(consignee_inn__icontains=q)
            )
        if filter_type:
            qs = qs.filter(cargo_type=filter_type)
        if filter_ship:
            qs = qs.filter(shipment_type=filter_ship)

        # Если POST на шаге 1 — переход к шагу 2
        if request.method == 'POST' and request.POST.get('step') == '1':
            selected = request.POST.getlist('hawb_ids')
            if not selected:
                messages.error(request, 'Выберите хотя бы одну накладную')
                return redirect(f'{request.path}?step=1&q={q}&cargo_type={filter_type}&shipment_type={filter_ship}')
            # Передаём выбранные ID на шаг 2 через GET
            ids_param = ','.join(selected)
            return redirect(f'{request.path}?step=2&ids={ids_param}')

        context = {
            'step': '1',
            'hawbs': qs[:300],
            'total': qs.count(),
            'q': q,
            'filter_type': filter_type,
            'filter_ship': filter_ship,
        }
        return render(request, 'cargo/cargo_create_wizard.html', context)

    # ── ШАГ 2: данные партии и финальное создание ──
    if step == '2':
        ids_param = request.GET.get('ids', '')
        hawb_ids = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        selected_hawbs = HouseWaybill.objects.filter(pk__in=hawb_ids)

        if request.method == 'POST' and request.POST.get('step') == '2':
            awb_number = request.POST.get('awb_number', '').strip()
            if not awb_number:
                messages.error(request, 'Номер партии (AWB) обязателен')
                return redirect(f'{request.path}?step=2&ids={ids_param}')
            if Cargo.objects.filter(awb_number=awb_number).exists():
                messages.error(request, f'Партия {awb_number} уже существует')
                return redirect(f'{request.path}?step=2&ids={ids_param}')

            # Проверяем что все выбранные HAWB в статусе AT_ORIGIN_WH
            invalid_hawbs = selected_hawbs.exclude(logistics_status='AT_ORIGIN_WH')
            if invalid_hawbs.exists():
                messages.error(request, 'Некоторые накладные не в статусе «Принят на склад отправки» и не могут быть добавлены в партию')
                return redirect(f'{request.path}?step=2&ids={ids_param}')

            # Проверяем что у всех выбранных HAWB все товары согласованы
            not_approved_hawbs = selected_hawbs.annotate(
                _total=Count('goods', distinct=True),
                _approved=Count('goods', filter=Q(goods__approval_status='approved'), distinct=True),
            ).exclude(_total=DjF('_approved'))
            if not_approved_hawbs.exists():
                nums = ', '.join(h.hawb_number for h in not_approved_hawbs[:5])
                messages.error(request, f'Накладные с несогласованными товарами нельзя добавить в партию: {nums}')
                return redirect(f'{request.path}?step=2&ids={ids_param}')

            cargo = Cargo.objects.create(
                awb_number=awb_number,
                description=request.POST.get('description', ''),
                description_ru=request.POST.get('description_ru', ''),
                shp_type=request.POST.get('shp_type', 'IMPEX'),
                stage='DRAFT',
                flight_number=request.POST.get('flight_number', ''),
                flight_date=request.POST.get('flight_date') or None,
                departure_iata=request.POST.get('departure_iata', '').upper(),
                arrival_iata=request.POST.get('arrival_iata', '').upper(),
                transportation_mode=4,
                weight=request.POST.get('weight') or None,
                pieces_declared=request.POST.get('pieces_declared') or 0,
                invoice_currency=request.POST.get('invoice_currency', 'USD'),
                invoice_value=request.POST.get('invoice_value') or None,
                # Склад не назначается на этапе DRAFT
                last_status_change=timezone.now(),
                created_by=request.user,
            )

            # Привязываем HAWB к созданной партии и переводим в статус CONSOLIDATED
            count = selected_hawbs.update(
                mawb=cargo,
                logistics_status='CONSOLIDATED',
                logistics_status_date=timezone.now(),
            )
            messages.success(request, f'Партия {awb_number} создана, привязано {count} накладных')
            return redirect('cargo_detail', awb_number=awb_number)

        # Суммарный вес выбранных HAWB
        total_weight = sum(h.weight or 0 for h in selected_hawbs)
        total_pieces = sum(h.pieces_declared for h in selected_hawbs)

        context = {
            'step': '2',
            'selected_hawbs': selected_hawbs,
            'ids_param': ids_param,
            'total_weight': total_weight,
            'total_pieces': total_pieces,
            'warehouses': __import__('cargo.models', fromlist=['Warehouse']).Warehouse.objects.filter(is_active=True),
        }
        return render(request, 'cargo/cargo_create_wizard.html', context)


@login_required
def export_hawbs_excel(request):
    """Экспорт HAWB в Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl не установлен', status=500)

    from .models import HouseWaybill
    qs = HouseWaybill.objects.select_related('mawb', 'assigned_to').all()

    q = request.GET.get('q', '').strip()
    filter_logistics = request.GET.get('logistics_status', '')
    filter_type = request.GET.get('cargo_type', '')
    filter_ship = request.GET.get('shipment_type', '')
    filter_unreleased = request.GET.get('unreleased', '')

    if q:
        qs = qs.filter(Q(hawb_number__icontains=q) | Q(consignee_name__icontains=q) | Q(mawb__awb_number__icontains=q))
    if filter_logistics:
        qs = qs.filter(logistics_status=filter_logistics)
    if filter_type:
        qs = qs.filter(cargo_type=filter_type)
    if filter_ship:
        qs = qs.filter(shipment_type=filter_ship)
    if filter_unreleased:
        qs = qs.exclude(customs_status='RELEASED')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HAWB'

    headers = [
        '№ HAWB', 'Тип груза', 'Тип отправки', 'MAWB партия',
        'Маршрут', 'Дата рейса', 'Номер рейса',
        'Размещение на СВХ', '№ Лицензии СВХ',
        'Получатель', 'ИНН', 'Город',
        'Вес (кг)', 'Мест',
        'Стоимость', 'Валюта', 'Рег. номер ДТ',
        'Инвойс', 'Упак. лист', 'Разреш. доки', 'Тех. описание',
        'Документов', 'Требуется', 'Лог. статус', 'Там. статус', 'Ответственный',
    ]

    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for hawb in qs:
        ws.append([
            hawb.hawb_number,
            hawb.cargo_type,
            hawb.get_shipment_type_display(),
            hawb.mawb.awb_number if hawb.mawb else '',
            f"{hawb.mawb.departure_iata or ''}→{hawb.mawb.arrival_iata or ''}" if hawb.mawb else '',
            hawb.mawb.flight_date.strftime('%d.%m.%Y') if hawb.mawb and hawb.mawb.flight_date else '',
            hawb.mawb.flight_number if hawb.mawb else '',
            hawb.scan_into_bond.strftime('%d.%m.%Y %H:%M') if hawb.scan_into_bond else '',
            hawb.mawb.warehouse_license if hawb.mawb else '',
            hawb.consignee_name,
            hawb.consignee_inn,
            hawb.consignee_city,
            float(hawb.weight) if hawb.weight else '',
            hawb.pieces_declared,
            float(hawb.invoice_value) if hawb.invoice_value else '',
            hawb.invoice_currency,
            hawb.customs_declaration_number,
            'Да' if hawb.doc_invoice else 'Нет',
            'Да' if hawb.doc_packing_list else 'Нет',
            'Да' if hawb.doc_permit else 'Нет',
            'Да' if hawb.doc_tech_desc else 'Нет',
            hawb.docs_count,
            hawb.docs_required,
            hawb.get_logistics_status_display(),
            hawb.get_customs_status_display() if hawb.customs_status else '',
            hawb.assigned_to.get_full_name() if hawb.assigned_to else '',
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    output = __import__('io').BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'hawb_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def document_types_search(request):
    """JSON-автокомплит для поиска типов документов (для чеклиста HAWB)"""
    q = request.GET.get('q', '').strip()
    exclude_hawb = request.GET.get('hawb_id', '')

    qs = DocumentType.objects.filter(is_active=True)
    if q:
        qs = qs.filter(name__icontains=q)

    # Исключаем уже добавленные в чеклист данного HAWB
    if exclude_hawb and exclude_hawb.isdigit():
        already = HAWBChecklistItem.objects.filter(
            hawb_id=int(exclude_hawb)
        ).values_list('document_type_id', flat=True)
        qs = qs.exclude(pk__in=already)

    results = [
        {
            'id': dt.id,
            'name': dt.name,
            'category': dt.get_category_display(),
        }
        for dt in qs.order_by('category', 'name')[:50]
    ]
    return JsonResponse({'results': results})


@login_required
def cargo_set_stage(request, awb_number: str):
    """Изменить этап партии"""
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'close_draft':
            try:
                cargo.close_draft(user=request.user)
                messages.success(request, f'Партия {awb_number} сформирована и закрыта для редактирования')
            except ValueError as e:
                messages.error(request, str(e))

        elif action == 'set_stage':
            new_stage = request.POST.get('stage', '').strip()
            valid = [s[0] for s in __import__('cargo.models', fromlist=['STAGE_CHOICES']).STAGE_CHOICES]
            if new_stage in valid:
                cargo.set_stage(new_stage, user=request.user)
                messages.success(request, f'Этап изменён: {cargo.stage_display}')

        elif action == 'advance':
            next_s = cargo.advance_stage(user=request.user)
            messages.success(request, f'Этап изменён: {cargo.stage_display}')

        elif action == 'update_dates':
            dep_str = request.POST.get('departure_date', '').strip()
            arr_str = request.POST.get('flight_date', '').strip()
            from datetime import date as date_type
            try:
                dep = date_type.fromisoformat(dep_str) if dep_str else None
                arr = date_type.fromisoformat(arr_str) if arr_str else None
                cargo.departure_date = dep
                cargo.flight_date = arr
                cargo.save()
                messages.success(request, 'Даты рейса обновлены')
            except ValueError as e:
                messages.error(request, str(e))

        elif action == 'toggle_self_clearance':
            cargo.is_self_clearance = not cargo.is_self_clearance
            cargo.save(update_fields=['is_self_clearance'])
            messages.success(
                request,
                'Партия отмечена как «ТО клиентом»' if cargo.is_self_clearance
                else 'Партия отмечена как «Наше ТО»'
            )

        elif action == 'correct_weight':
            # Корректировка веса с пропорциональным пересчётом HAWB
            new_weight_str = request.POST.get('new_weight', '').strip()
            if not new_weight_str:
                messages.error(request, 'Укажите новый вес партии')
            elif not cargo.can_correct_weight:
                messages.error(request, 'Корректировка веса доступна только на этапах «Отправлена» и «Прибыла»')
            else:
                try:
                    new_weight = float(new_weight_str)
                    if new_weight <= 0:
                        raise ValueError
                    cargo.redistribute_hawb_weights(new_weight, user=request.user)
                    messages.success(
                        request,
                        f'Вес партии изменён на {new_weight} кг. '
                        f'Веса накладных пропорционально пересчитаны.'
                    )
                except ValueError:
                    messages.error(request, 'Некорректное значение веса')

    return redirect('cargo_detail', awb_number=awb_number)


# ─────────────────────────── GOODS (ТОВАРЫ) ───────────────────────────

@login_required
def goods_list(request):
    """Раздел «Товары» — все товарные позиции из всех накладных"""
    from .models import HAWBGood
    qs = HAWBGood.objects.select_related(
        'hawb', 'hawb__mawb', 'approved_by'
    ).all()

    # ── Поиск / фильтрация ──
    q              = request.GET.get('q', '').strip()
    filter_hawb    = request.GET.get('hawb', '').strip()
    filter_tnved   = request.GET.get('tnved', '').strip()
    filter_approve = request.GET.get('approval', '').strip()

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(manufacturer__icontains=q) |
            Q(tnved_code__icontains=q) |
            Q(article__icontains=q) |
            Q(hawb__hawb_number__icontains=q)
        )
    if filter_hawb:
        qs = qs.filter(hawb__hawb_number__icontains=filter_hawb)
    if filter_tnved:
        qs = qs.filter(tnved_code__icontains=filter_tnved)
    if filter_approve:
        qs = qs.filter(approval_status=filter_approve)

    # ── Сортировка ──
    sort      = request.GET.get('sort', 'id')
    direction = request.GET.get('dir', 'asc')
    allowed_sorts = [
        'id', 'name', 'tnved_code', 'brand', 'manufacturer',
        'model', 'article', 'quantity', 'unit',
        'weight_net', 'weight_gross', 'unit_price', 'total_value', 'currency',
        'hawb__hawb_number', 'approval_status', 'approved_at',
    ]
    if sort in allowed_sorts:
        prefix = '-' if direction == 'desc' else ''
        qs = qs.order_by(f'{prefix}{sort}')

    from .models import HAWBGood as G
    approval_choices = G.APPROVAL_STATUS_CHOICES

    goods_cols = [
        ('name',                'Наименование'),
        ('tnved_code',          'ТН ВЭД'),
        ('brand',               'Бренд'),
        ('manufacturer',        'Изготовитель'),
        ('model',               'Модель'),
        ('article',             'Артикул'),
        ('product_url',         'Ссылка'),
        ('quantity',            'Кол-во'),
        ('unit',                'Ед.'),
        ('quantity_additional', 'ДЕИ'),
        ('unit_additional',     'Ед. ДЕИ'),
        ('weight_net',          'Нетто кг'),
        ('weight_gross',        'Брутто кг'),
        ('unit_price',          'Цена ед.'),
        ('total_value',         'Стоим. итого'),
        ('currency',            'Вал.'),
        ('cargo_type',          'Тип груза'),
        ('hawb__hawb_number',   'HAWB'),
        ('approval_status',     'Согласование'),
    ]

    context = {
        'goods': qs[:2000],
        'total': qs.count(),
        'q': q,
        'filter_hawb': filter_hawb,
        'filter_tnved': filter_tnved,
        'filter_approve': filter_approve,
        'sort': sort,
        'direction': direction,
        'approval_choices': approval_choices,
        'goods_cols': goods_cols,
    }
    return render(request, 'cargo/goods_list.html', context)


@login_required
def goods_export_excel(request):
    """Экспорт товарных позиций в Excel (с учётом текущих фильтров)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl не установлен', status=500)

    from .models import HAWBGood
    qs = HAWBGood.objects.select_related('hawb', 'hawb__mawb', 'approved_by').all()

    q              = request.GET.get('q', '').strip()
    filter_hawb    = request.GET.get('hawb', '').strip()
    filter_tnved   = request.GET.get('tnved', '').strip()
    filter_approve = request.GET.get('approval', '').strip()

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(manufacturer__icontains=q) |
            Q(tnved_code__icontains=q) |
            Q(article__icontains=q) |
            Q(hawb__hawb_number__icontains=q)
        )
    if filter_hawb:
        qs = qs.filter(hawb__hawb_number__icontains=filter_hawb)
    if filter_tnved:
        qs = qs.filter(tnved_code__icontains=filter_tnved)
    if filter_approve:
        qs = qs.filter(approval_status=filter_approve)

    sort      = request.GET.get('sort', 'id')
    direction = request.GET.get('dir', 'asc')
    allowed_sorts = [
        'id', 'name', 'tnved_code', 'brand', 'manufacturer',
        'model', 'article', 'quantity', 'unit',
        'weight_net', 'weight_gross', 'unit_price', 'total_value', 'currency',
        'hawb__hawb_number', 'approval_status', 'approved_at',
    ]
    if sort in allowed_sorts:
        prefix = '-' if direction == 'desc' else ''
        qs = qs.order_by(f'{prefix}{sort}')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Товары'

    headers = [
        'Наименование', 'ТН ВЭД', 'Бренд', 'Изготовитель', 'Модель', 'Артикул',
        'Кол-во', 'Ед.', 'Нетто кг', 'Брутто кг', 'Цена ед.', 'Стоим. итого',
        'Валюта', 'Тип груза', 'HAWB', 'MAWB', 'Согласование',
    ]
    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for good in qs:
        ws.append([
            good.name,
            good.tnved_code or '',
            good.brand or '',
            good.manufacturer or '',
            good.model or '',
            good.article or '',
            float(good.quantity) if good.quantity is not None else '',
            good.unit or '',
            float(good.weight_net) if good.weight_net is not None else '',
            float(good.weight_gross) if good.weight_gross is not None else '',
            float(good.unit_price) if good.unit_price is not None else '',
            float(good.total_value) if good.total_value is not None else '',
            good.currency or '',
            good.cargo_type or '',
            good.hawb.hawb_number,
            good.hawb.mawb.awb_number if good.hawb.mawb else '',
            good.approval_label,
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = __import__('io').BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'goods_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def goods_approve(request):
    """Массовое согласование/отклонение товарных позиций"""
    if request.method != 'POST':
        return redirect('goods_list')

    from .models import HAWBGood
    ids_raw  = request.POST.getlist('good_ids')
    action   = request.POST.get('action', '').strip()
    comment  = request.POST.get('comment', '').strip()

    valid_actions = {
        'approve':       'approved',
        'clarification': 'clarification',
        'reject':        'rejected',
    }
    if action not in valid_actions or not ids_raw:
        messages.error(request, 'Не выбраны товары или действие')
        return redirect('goods_list')

    ids = [int(i) for i in ids_raw if i.isdigit()]
    qs  = HAWBGood.objects.filter(pk__in=ids)

    new_status = valid_actions[action]
    now        = timezone.now()

    for good in qs:
        good.approval_status  = new_status
        good.approval_comment = comment
        good.approved_by      = request.user
        good.approved_at      = now
    HAWBGood.objects.bulk_update(qs, ['approval_status', 'approval_comment', 'approved_by', 'approved_at'])

    label_map = {
        'approved':      'Согласовано',
        'clarification': 'Отправлено на уточнение',
        'rejected':      'Отклонено',
    }
    messages.success(request, f'{label_map[new_status]}: {len(ids)} поз.')

    return_params = request.POST.get('return_params', '')
    url = '/goods/'
    if return_params:
        url += '?' + return_params
    from django.http import HttpResponseRedirect
    return HttpResponseRedirect(url)


# ─────────────────────────── REST API ───────────────────────────

from rest_framework import generics, filters
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from django.views.decorators.csrf import csrf_exempt
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .serializers import CargoListSerializer, CargoDetailSerializer, HouseWaybillSerializer


class CargoListAPIView(generics.ListAPIView):
    """GET /api/v1/cargo/ — список партий с фильтрацией"""
    serializer_class = CargoListSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['awb_number', 'customs_declaration_number', 'description',
                     'assignments__user__username', 'assignments__user__last_name']
    ordering_fields = ['created_at', 'flight_date', 'weight', 'stage']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = Cargo.objects.select_related('warehouse').all()
        stage = self.request.query_params.get('stage')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        warehouse = self.request.query_params.get('warehouse')

        if stage:
            qs = qs.filter(stage=stage)
        if date_from:
            qs = qs.filter(flight_date__gte=date_from)
        if date_to:
            qs = qs.filter(flight_date__lte=date_to)
        if warehouse:
            qs = qs.filter(warehouse_id=warehouse)
        return qs


class CargoDetailAPIView(generics.RetrieveAPIView):
    """GET /api/v1/cargo/<awb_number>/ — детали партии с накладными"""
    serializer_class = CargoDetailSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'awb_number'
    queryset = Cargo.objects.select_related('warehouse').prefetch_related('hawbs')


class HawbListAPIView(generics.ListAPIView):
    """GET /api/v1/hawbs/ — список накладных с фильтрацией"""
    serializer_class = HouseWaybillSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['hawb_number', 'consignee_name', 'customs_declaration_number']
    ordering_fields = ['created_at', 'weight', 'logistics_status']
    ordering = ['-created_at']

    def get_queryset(self):
        from .models import HouseWaybill
        qs = HouseWaybill.objects.select_related('mawb').all()
        mawb = self.request.query_params.get('mawb')
        status = self.request.query_params.get('status')

        if mawb:
            qs = qs.filter(mawb__awb_number=mawb)
        if status:
            qs = qs.filter(logistics_status=status)
        return qs


@api_view(['GET'])
@permission_classes([])
def api_health(request):
    """GET /api/v1/health/ — проверка доступности API (без авторизации)"""
    return Response({
        'status': 'ok',
        'service': 'CargoTrack Pro',
        'version': '1.0',
    })


@csrf_exempt
@api_view(['GET'])
@permission_classes([])
@authentication_classes([])
def api_status_text(request):
    """GET /api/v1/status/text/?key=... — collect_status() как plain text.

    Используется telegram_poll.py (standalone bot-клиент на ноуте юзера),
    т.к. VPS-провайдер блокирует outbound к api.telegram.org. Бот сам
    бегает к Telegram (с ноута оно работает) и за статусом ходит сюда.

    Авторизация: ?key=<secret> из settings.STATUS_API_KEY (или TELEGRAM_ALERT.chat_id
    как fallback shared secret).
    """
    from django.conf import settings
    from django.http import HttpResponse
    expected = (getattr(settings, 'STATUS_API_KEY', None)
                or (getattr(settings, 'TELEGRAM_ALERT', None) or {}).get('chat_id', ''))
    given = request.GET.get('key', '')
    if not expected or str(given) != str(expected):
        return HttpResponse('forbidden', status=403, content_type='text/plain')
    try:
        from cargo.services.notify.status import collect_status
        return HttpResponse(collect_status(), content_type='text/plain; charset=utf-8')
    except Exception as e:
        return HttpResponse(f'collect_status error: {e}', status=500,
                            content_type='text/plain; charset=utf-8')


@csrf_exempt
@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def api_telegram_webhook(request):
    """POST /api/v1/telegram/webhook/ — приём update от Telegram bot API.

    Поддерживаемые команды:
      /status, /health — health-status всех модулей через collect_status()
      /start, /help    — справка

    Защита: chat_id whitelist из settings.TELEGRAM_ALERT['chat_id']. Любой
    update от других chat'ов отбрасывается.
    """
    from django.conf import settings
    import logging, requests
    logger = logging.getLogger('cargo.telegram.webhook')

    cfg = getattr(settings, 'TELEGRAM_ALERT', None) or getattr(
        settings, 'DEKLARANT_ALERT_TELEGRAM', None)
    if not cfg:
        return Response({'ok': True, 'note': 'no TELEGRAM_ALERT config'})

    allowed_chat = str(cfg.get('chat_id', '')).strip()
    bot_token = cfg.get('bot_token', '')
    if not bot_token:
        return Response({'ok': True, 'note': 'no bot_token'})

    data = request.data or {}
    msg = data.get('message') or data.get('edited_message') or {}
    chat = msg.get('chat') or {}
    chat_id = str(chat.get('id', ''))
    text = (msg.get('text') or '').strip()

    if not chat_id or not text:
        return Response({'ok': True})
    if allowed_chat and chat_id != allowed_chat:
        logger.warning('telegram webhook: chat_id %s not in whitelist', chat_id)
        return Response({'ok': True})

    cmd = text.split()[0].lower().split('@')[0]
    reply = ''
    parse_mode = 'Markdown'

    if cmd in ('/status', '/health'):
        try:
            from cargo.services.notify.status import collect_status
            reply = collect_status()
        except Exception as e:
            logger.exception('collect_status failed')
            reply = f'❌ collect_status error: {e}'
            parse_mode = ''
    elif cmd in ('/start', '/help'):
        reply = ('🤖 *CargoTrack Alerts Bot*\n\n'
                 'Команды:\n'
                 '`/status` — статус всех модулей\n'
                 '`/health` — то же что /status')
    else:
        return Response({'ok': True})  # неизвестная команда — молчим

    try:
        requests.post(
            f'https://api.telegram.org/bot{bot_token}/sendMessage',
            data={'chat_id': chat_id, 'text': reply,
                  'parse_mode': parse_mode} if parse_mode else
                 {'chat_id': chat_id, 'text': reply},
            timeout=10,
        )
    except Exception:
        logger.exception('telegram reply failed')
    return Response({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD WIDGETS API
# ═══════════════════════════════════════════════════════════════════════════════

import json
import datetime as _dt
from django.views.decorators.http import require_http_methods
from .models import DashboardWidget
from .cql_parser import (
    parse_cql, parse_to_ast, CQLError, FIELD_REFERENCE, get_field_reference,
    GROUPABLE_FIELDS, AGGREGATABLE_FIELDS, ALLOWED_AGGS,
    get_groupable_fields, get_aggregatable_fields,
)
from .widget_registry import (
    ENTITY_REGISTRY, get_entities, get_field_catalog_union,
    resolve_entity_for_field,
)


def _parse_json_body(request):
    """Парсит JSON-тело запроса. Возвращает (data, error_response).
    Если error_response не None — её нужно вернуть из view как 400."""
    try:
        return json.loads(request.body or b'{}'), None
    except json.JSONDecodeError as e:
        return None, JsonResponse(
            {'error': f'Некорректный JSON: {e}'}, status=400
        )


def _paginate(request, qs, default_page_size: int = 200, max_page_size: int = 500):
    """Постраничная выборка с метаданными. Возвращает (items_slice, meta_dict).
    Параметры из query: ?page=N&page_size=M. Без них берётся весь qs (до max).
    UI, не использующий пагинацию, продолжит работать — он читает массив, а
    pagination игнорирует."""
    try:
        page = max(1, int(request.GET.get('page', 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        size = int(request.GET.get('page_size', default_page_size))
    except (TypeError, ValueError):
        size = default_page_size
    size = max(1, min(max_page_size, size))
    total = qs.count()
    pages = (total + size - 1) // size if size else 1
    items = qs[(page - 1) * size: page * size]
    return items, {
        'page': page,
        'page_size': size,
        'total': total,
        'pages': pages,
    }


def _widget_to_dict(w: DashboardWidget) -> dict:
    return {
        'id':           w.id,
        'widget_type':  w.widget_type,
        'entity_type':  w.entity_type,
        'title':        w.title,
        'filter_query': w.filter_query,
        'config':       w.config,
        'pos_x':        w.pos_x,
        'pos_y':        w.pos_y,
        'width':        w.width,
        'height':       w.height,
    }


@login_required
@require_http_methods(['GET', 'POST'])
def api_dashboard_widgets(request):
    """GET  /api/v1/dashboard/widgets/  — список виджетов пользователя
       POST /api/v1/dashboard/widgets/  — создать виджет"""
    if request.method == 'GET':
        widgets, page = _paginate(
            request,
            DashboardWidget.objects.filter(user=request.user).order_by('id'),
        )
        return JsonResponse({
            'widgets': [_widget_to_dict(w) for w in widgets],
            'pagination': page,
        })

    data, err = _parse_json_body(request)
    if err: return err
    cql = data.get('filter_query', '').strip()
    entity_type = data.get('entity_type', 'cargo')
    if entity_type not in ('cargo', 'hawb'):
        entity_type = 'cargo'
    if cql:
        try:
            parse_cql(cql, {'me': request.user.username}, entity_type=entity_type)
        except CQLError as e:
            return JsonResponse({'error': str(e)}, status=400)

    widget = DashboardWidget.objects.create(
        user=request.user,
        widget_type=data.get('widget_type', 'stat'),
        entity_type=entity_type,
        title=data.get('title', 'Виджет'),
        filter_query=cql,
        config=data.get('config', {}),
        pos_x=data.get('pos_x', 0),
        pos_y=data.get('pos_y', 9999),
        width=data.get('width', 3),
        height=data.get('height', 2),
    )
    return JsonResponse(_widget_to_dict(widget), status=201)


@login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def api_dashboard_widget(request, widget_id):
    """GET/PUT/DELETE /api/v1/dashboard/widgets/<id>/"""
    try:
        widget = DashboardWidget.objects.get(id=widget_id, user=request.user)
    except DashboardWidget.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

    if request.method == 'GET':
        return JsonResponse(_widget_to_dict(widget))

    if request.method == 'DELETE':
        widget.delete()
        return JsonResponse({'ok': True})

    # PUT
    data, err = _parse_json_body(request)
    if err: return err
    cql = data.get('filter_query', widget.filter_query).strip()
    entity_type = data.get('entity_type', widget.entity_type)
    if entity_type not in ('cargo', 'hawb'):
        entity_type = widget.entity_type
    if cql:
        try:
            parse_cql(cql, {'me': request.user.username}, entity_type=entity_type)
        except CQLError as e:
            return JsonResponse({'error': str(e)}, status=400)

    widget.title        = data.get('title',       widget.title)
    widget.widget_type  = data.get('widget_type', widget.widget_type)
    widget.entity_type  = entity_type
    widget.filter_query = cql
    widget.config       = data.get('config',      widget.config)
    widget.width        = data.get('width',        widget.width)
    widget.height       = data.get('height',       widget.height)
    widget.save()
    return JsonResponse(_widget_to_dict(widget))


@login_required
@require_http_methods(['GET'])
def api_dashboard_widget_data(request, widget_id):
    """GET /api/v1/dashboard/widgets/<id>/data/ — данные виджета.
    Для pivot-виджета поддерживается параметр ?format=xlsx — возвращает
    Excel-файл с теми же данными (включая итоги)."""
    try:
        widget = DashboardWidget.objects.get(id=widget_id, user=request.user)
    except DashboardWidget.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

    if request.GET.get('format', '').lower() == 'xlsx':
        if widget.widget_type != 'pivot':
            return JsonResponse({'error': 'Excel-экспорт доступен только для pivot-виджета'}, status=400)
        return _export_pivot_xlsx(widget, request)

    # Workload-виджеты не привязаны к одной сущности и не используют CQL —
    # обрабатываем до основного диспетчера.
    if widget.widget_type.startswith('workload_'):
        return _widget_workload_dispatch(widget, request)

    try:
        q = parse_cql(widget.filter_query, {'me': request.user.username},
                      entity_type=widget.entity_type)
    except CQLError as e:
        return JsonResponse({'error': f'CQL: {e}'}, status=400)

    if widget.entity_type == 'hawb':
        from .models import HouseWaybill
        qs = HouseWaybill.objects.filter(q).select_related('mawb', 'assigned_to').distinct()

        if widget.widget_type == 'stat':
            return _hawb_widget_stat(qs, widget.config)
        if widget.widget_type == 'kanban':
            return _hawb_widget_kanban(qs, widget.config)
        if widget.widget_type == 'table':
            return _hawb_widget_table(qs, widget.config)
        if widget.widget_type == 'chart_stage':
            return _hawb_widget_chart_logistics(qs)
        if widget.widget_type == 'chart_warehouse':
            return _hawb_widget_chart_warehouse(qs)
        if widget.widget_type == 'chart_pie':
            return _widget_chart_pie(qs, widget.config, 'hawb')
        if widget.widget_type == 'forecast_arrivals':
            return _widget_forecast_arrivals(widget.config)
        if widget.widget_type == 'pivot':
            return _compute_pivot(qs, widget.config, 'hawb', request.user)
        return JsonResponse({'error': 'Неизвестный тип виджета'}, status=400)

    qs = Cargo.objects.filter(q).select_related('warehouse').distinct()

    if widget.widget_type == 'stat':
        return _widget_stat(qs, widget.config)
    if widget.widget_type == 'kanban':
        return _widget_kanban(qs, widget.config)
    if widget.widget_type == 'table':
        return _widget_table(qs, widget.config)
    if widget.widget_type == 'chart_stage':
        return _widget_chart_stage(qs)
    if widget.widget_type == 'chart_warehouse':
        return _widget_chart_warehouse(qs)
    if widget.widget_type == 'chart_pie':
        return _widget_chart_pie(qs, widget.config, 'cargo')
    if widget.widget_type == 'forecast_arrivals':
        return _widget_forecast_arrivals(widget.config)
    if widget.widget_type == 'pivot':
        return _compute_pivot(qs, widget.config, 'cargo', request.user)
    return JsonResponse({'error': 'Неизвестный тип виджета'}, status=400)


@login_required
@require_http_methods(['POST'])
def api_dashboard_layout(request):
    """POST /api/v1/dashboard/layout/ — сохранить расположение виджетов"""
    items, err = _parse_json_body(request)
    if err: return err
    for item in items:
        DashboardWidget.objects.filter(
            id=item['id'], user=request.user
        ).update(
            pos_x=item.get('x', 0),
            pos_y=item.get('y', 0),
            width=item.get('w', 3),
            height=item.get('h', 2),
        )
    return JsonResponse({'ok': True})


@login_required
@require_http_methods(['POST'])
def api_dashboard_cql_validate(request):
    """POST /api/v1/dashboard/cql/validate/ — проверка CQL-запроса"""
    data, err = _parse_json_body(request)
    if err: return err
    cql = data.get('query', '').strip()
    entity_type = data.get('entity_type', 'cargo')
    if entity_type not in ('cargo', 'hawb'):
        entity_type = 'cargo'
    if not cql:
        return JsonResponse({'valid': True})
    try:
        parse_cql(cql, {'me': request.user.username}, entity_type=entity_type)
        return JsonResponse({'valid': True})
    except CQLError as e:
        return JsonResponse({'valid': False, 'error': str(e)})


@login_required
@require_http_methods(['GET'])
def api_dashboard_cql_fields(request):
    """GET /api/v1/dashboard/cql/fields/ — справка по полям CQL"""
    entity_type = request.GET.get('entity_type', 'cargo')
    if entity_type not in ('cargo', 'hawb'):
        entity_type = 'cargo'
    reference = get_field_reference(entity_type)
    fields = [
        {'field': f, 'label': l, 'example': ex, 'hint': h}
        for f, l, ex, h in reference
    ]
    return JsonResponse({'fields': fields, 'entity_type': entity_type})


@login_required
@require_http_methods(['GET'])
def api_dashboard_cql_values(request):
    """GET /api/v1/dashboard/cql/values/?entity=cargo|hawb&field=<key>&q=<prefix>
    Возвращает топ-N уникальных значений поля для типа-ввода в фильтре."""
    from .cql_parser import CARGO_FIELDS, HAWB_FIELDS, _normalize_fields
    from .models import HouseWaybill

    entity = request.GET.get('entity', 'cargo')
    if entity not in ('cargo', 'hawb'):
        return JsonResponse({'error': 'entity должен быть cargo или hawb'}, status=400)
    field_key = (request.GET.get('field') or '').strip()
    if not field_key:
        return JsonResponse({'error': 'не указано поле'}, status=400)

    fields = _normalize_fields(CARGO_FIELDS if entity == 'cargo' else HAWB_FIELDS)
    spec = fields.get(field_key)
    if not spec or spec['db_field'] is None or spec['type'] != 'str':
        return JsonResponse({'values': []})

    db_field = spec['db_field']
    q = (request.GET.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.GET.get('limit') or 20), 50))
    except (TypeError, ValueError):
        limit = 20

    Model = Cargo if entity == 'cargo' else HouseWaybill
    qs = Model.objects.all()
    if q:
        qs = qs.filter(**{f'{db_field}__icontains': q})
    qs = qs.exclude(**{db_field: ''}).exclude(**{f'{db_field}__isnull': True})
    values = list(
        qs.order_by().values_list(db_field, flat=True).distinct()[:limit]
    )
    return JsonResponse({'values': values})


@login_required
@require_http_methods(['POST'])
def api_dashboard_cql_parse_tree(request):
    """POST /api/v1/dashboard/cql/parse-tree/ — разобрать CQL-строку в JSON AST.
    Используется визуальным builder'ом для восстановления дерева из сохранённого фильтра."""
    data, err = _parse_json_body(request)
    if err: return err
    cql = (data.get('query') or '').strip()
    entity_type = data.get('entity_type', 'cargo')
    if entity_type not in ('cargo', 'hawb'):
        entity_type = 'cargo'
    try:
        tree = parse_to_ast(cql, entity_type=entity_type)
        return JsonResponse({'ok': True, 'tree': tree})
    except CQLError as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ── Метки (Labels) ────────────────────────────────────────────────────────────

def _label_to_dict(label: Label, with_usage: bool = False) -> dict:
    d = {
        'id':          label.id,
        'name':        label.name,
        'color':       label.color,
        'description': label.description,
    }
    if with_usage:
        d['usage_count'] = label.cargos.count() + label.hawbs.count()
    return d


@login_required
@require_http_methods(['GET', 'POST'])
def api_labels(request):
    """GET  /api/v1/labels/  — список всех меток.
       POST /api/v1/labels/  — создать метку {name, color?, description?}."""
    if request.method == 'GET':
        with_usage = request.GET.get('with_usage') == '1'
        qs = Label.objects.all().order_by('name')
        items, page = _paginate(request, qs)
        return JsonResponse({
            'labels': [_label_to_dict(l, with_usage=with_usage) for l in items],
            'pagination': page,
        })
    # POST
    data, err = _parse_json_body(request)
    if err: return err
    name = (data.get('name') or '').strip()
    if not name:
        return JsonResponse({'error': 'Название обязательно'}, status=400)
    if Label.objects.filter(name__iexact=name).exists():
        return JsonResponse({'error': f'Метка {name!r} уже существует'}, status=400)
    label = Label.objects.create(
        name=name,
        color=(data.get('color') or '#6c757d').strip(),
        description=(data.get('description') or '').strip(),
        created_by=request.user,
    )
    return JsonResponse(_label_to_dict(label), status=201)


@login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def api_label(request, label_id: int):
    """GET/PUT/DELETE /api/v1/labels/<id>/"""
    try:
        label = Label.objects.get(pk=label_id)
    except Label.DoesNotExist:
        return JsonResponse({'error': 'Метка не найдена'}, status=404)

    if request.method == 'GET':
        return JsonResponse(_label_to_dict(label, with_usage=True))

    if request.method == 'PUT':
        data, err = _parse_json_body(request)
        if err: return err
        new_name = (data.get('name') or label.name).strip()
        if new_name != label.name and Label.objects.filter(name__iexact=new_name).exists():
            return JsonResponse({'error': f'Метка {new_name!r} уже существует'}, status=400)
        label.name = new_name
        if 'color' in data:
            label.color = (data.get('color') or '#6c757d').strip()
        if 'description' in data:
            label.description = (data.get('description') or '').strip()
        label.save()
        return JsonResponse(_label_to_dict(label))

    # DELETE — требует ?force=1, если метка используется
    usage = label.cargos.count() + label.hawbs.count()
    if usage > 0 and request.GET.get('force') != '1':
        return JsonResponse({
            'error': f'Метка используется в {usage} сущностях. '
                     f'Добавьте ?force=1 для удаления.',
            'usage_count': usage,
        }, status=409)
    label.delete()
    return JsonResponse({'ok': True})


@login_required
@require_http_methods(['GET', 'PUT'])
def api_cargo_labels(request, awb_number: str):
    """GET/PUT /api/v1/cargo/<awb>/labels/ — метки конкретной партии.
    PUT body: {label_ids: [1, 2, 3]} — заменить набор."""
    try:
        cargo = Cargo.objects.prefetch_related('labels').get(awb_number=awb_number)
    except Cargo.DoesNotExist:
        return JsonResponse({'error': 'Партия не найдена'}, status=404)
    if request.method == 'GET':
        return JsonResponse({
            'labels': [_label_to_dict(l) for l in cargo.labels.all()]
        })
    # PUT
    data, err = _parse_json_body(request)
    if err: return err
    ids = data.get('label_ids') or []
    if not isinstance(ids, list):
        return JsonResponse({'error': 'label_ids должен быть массивом'}, status=400)
    labels = list(Label.objects.filter(pk__in=ids))
    cargo.labels.set(labels)
    return JsonResponse({
        'labels': [_label_to_dict(l) for l in labels]
    })


@login_required
@require_http_methods(['GET', 'PUT'])
def api_hawb_labels(request, hawb_id: int):
    """GET/PUT /api/v1/hawbs/<id>/labels/ — метки конкретной накладной."""
    from .models import HouseWaybill
    try:
        hawb = HouseWaybill.objects.prefetch_related('labels').get(pk=hawb_id)
    except HouseWaybill.DoesNotExist:
        return JsonResponse({'error': 'Накладная не найдена'}, status=404)
    if request.method == 'GET':
        return JsonResponse({
            'labels': [_label_to_dict(l) for l in hawb.labels.all()]
        })
    data, err = _parse_json_body(request)
    if err: return err
    ids = data.get('label_ids') or []
    if not isinstance(ids, list):
        return JsonResponse({'error': 'label_ids должен быть массивом'}, status=400)
    labels = list(Label.objects.filter(pk__in=ids))
    hawb.labels.set(labels)
    return JsonResponse({
        'labels': [_label_to_dict(l) for l in labels]
    })


@login_required
def labels_page(request):
    """Страница управления метками /labels/"""
    return render(request, 'cargo/labels.html')


# ── Типы виджетов ─────────────────────────────────────────────────────────────

_STAGE_LABELS = dict(STAGE_CHOICES)
_STAGE_CSS = {
    'DRAFT':      'stage-draft',
    'FORMED':     'stage-formed',
    'DISPATCHED': 'stage-dispatched',
    'ARRIVED':    'stage-arrived',
    'CUSTOMS':    'stage-customs',
    'RELEASED':   'stage-released',
}


def _widget_stat(qs, config: dict):
    metric = config.get('metric', 'count')
    if metric == 'weight':
        val = float(qs.aggregate(s=Sum('weight'))['s'] or 0)
        return JsonResponse({'value': round(val, 1), 'suffix': ' кг', 'color': '#10b981'})
    if metric == 'pieces':
        val = qs.aggregate(s=Sum('pieces_declared'))['s'] or 0
        return JsonResponse({'value': val, 'suffix': ' мест', 'color': '#f59e0b'})
    if metric == 'problematic':
        from .cql_parser import parse_cql
        val = qs.filter(parse_cql('is_problematic = true', entity_type='cargo')).count()
        return JsonResponse({'value': val, 'suffix': '', 'color': '#ef4444'})
    # default: count
    val = qs.count()
    return JsonResponse({'value': val, 'suffix': ' партий', 'color': '#2563eb'})


def _widget_kanban(qs, config: dict):
    stages_filter = config.get('stages') or [s[0] for s in STAGE_CHOICES]
    limit = min(int(config.get('limit', 20)), 50)
    columns = []
    for code, label in STAGE_CHOICES:
        if code not in stages_filter:
            continue
        col_qs = qs.filter(stage=code)
        total = col_qs.count()
        items = list(
            col_qs[:limit].values(
                'awb_number', 'departure_iata', 'arrival_iata',
                'weight', 'flight_number', 'flight_date',
                'warehouse_license', 'is_draft',
            )
        )
        for item in items:
            if item['flight_date']:
                item['flight_date'] = item['flight_date'].strftime('%d.%m.%y')
        columns.append({
            'stage': code,
            'label': label,
            'css':   _STAGE_CSS.get(code, ''),
            'total': total,
            'items': items,
        })
    return JsonResponse({'columns': columns})


def _widget_table(qs, config: dict):
    return _render_table_widget(qs, config, 'cargo')


def _render_table_widget(qs, config: dict, entity_type: str):
    from .widget_columns import (
        get_column_catalog, get_default_columns, sortable_fields, serialize_column,
    )

    limit      = min(int(config.get('limit', 20)), 100)
    catalog    = get_column_catalog(entity_type)
    by_key     = {c['key']: c for c in catalog}
    sortable   = sortable_fields(entity_type)

    selected_keys = config.get('columns') or get_default_columns(entity_type)
    selected = [by_key[k] for k in selected_keys if k in by_key]
    if not selected:
        selected = [by_key[k] for k in get_default_columns(entity_type) if k in by_key]

    order_by = config.get('order_by') or '-created_at'
    desc = order_by.startswith('-')
    order_field = order_by.lstrip('-')
    # Принимаем как ORM-поле, так и `key` колонки из каталога —
    # фронт чаще присылает key (например, 'pieces'), а ORM-имя — 'pieces_declared'.
    if order_field in by_key and by_key[order_field].get('sortable'):
        order_by = ('-' if desc else '') + by_key[order_field]['db_fields'][0]
    elif order_field not in sortable:
        order_by = '-created_at'

    # Fields needed for row-link (independent of selected columns)
    if entity_type == 'cargo':
        mandatory = {'id', 'awb_number'}
        row_link  = '/cargo/{awb_number}/'
    else:
        mandatory = {'id', 'hawb_number'}
        row_link  = '/hawbs/{id}/'

    db_fields = set(mandatory)
    for col in selected:
        db_fields.update(col['db_fields'])

    items = list(qs.order_by(order_by)[:limit].values(*db_fields))

    sla_cols = [c for c in selected if c.get('type') == 'sla_progress']
    if sla_cols:
        from .sla import compute_sla_state, get_active_workflow_step
        step_cache = {}
        for it in items:
            eid = it.get('id')
            for col in sla_cols:
                meta  = col.get('sla', {})
                sval  = it.get(col['db_fields'][0])
                sdt   = it.get(col['db_fields'][1])
                step  = step_cache.get(eid)
                if step is None and eid is not None:
                    step = get_active_workflow_step(meta['entity_type'], eid)
                    step_cache[eid] = step
                it[f"__sla_{col['key']}__"] = compute_sla_state(
                    meta['entity_type'], meta['status_field'],
                    sval, sdt, workflow_step=step or None,
                )

    for it in items:
        for k, v in list(it.items()):
            if isinstance(v, _dt.datetime):
                it[k] = v.strftime('%d.%m.%Y %H:%M')
            elif isinstance(v, _dt.date):
                it[k] = v.strftime('%d.%m.%Y')

    cols_out = [serialize_column(c) for c in selected]

    return JsonResponse({
        'columns':  cols_out,
        'items':    items,
        'total':    qs.count(),
        'row_link': row_link,
    })


def _widget_chart_stage(qs):
    stats = qs.values('stage').annotate(cnt=Count('id')).order_by('stage')
    labels = [_STAGE_LABELS.get(s['stage'], s['stage']) for s in stats]
    values = [s['cnt'] for s in stats]
    group_keys = [s['stage'] for s in stats]
    return JsonResponse({
        'labels': labels, 'values': values,
        'group_keys': group_keys, 'group_by': 'stage',
    })


def _widget_chart_warehouse(qs):
    stats = (
        qs.exclude(warehouse_license='')
        .values('warehouse_license', 'warehouse_name')
        .annotate(cnt=Count('id'), total_w=Sum('weight'))
        .order_by('-cnt')[:12]
    )
    labels  = [s['warehouse_license'] or s['warehouse_name'] or '?' for s in stats]
    values  = [s['cnt'] for s in stats]
    weights = [float(s['total_w'] or 0) for s in stats]
    group_keys = [s['warehouse_license'] for s in stats]
    return JsonResponse({
        'labels': labels, 'values': values, 'weights': weights,
        'group_keys': group_keys, 'group_by': 'warehouse',
    })


def _widget_chart_pie(qs, config: dict, entity: str):
    """Круговая диаграмма: одна группировка + одна агрегируемая метрика.

    config: {
      group_by: 'field_key',
      metric:   {agg: 'count'|'sum'|'avg', field: 'field_key' (для sum/avg)},
      limit:    int (top-N слайсов, по умолчанию 12)
    }

    Ответ: {labels, values, group_by, group_keys, metric_label}.
    group_keys передаётся отдельным массивом «сырых» значений для drill-down
    (чтобы фронт мог построить CQL: `<field> = <raw_key>`).
    """
    if not isinstance(config, dict):
        return JsonResponse({'error': 'Некорректная конфигурация'}, status=400)
    group_by = (config.get('group_by') or '').strip()
    if not group_by:
        return JsonResponse({'error': 'Не задано поле группировки (group_by)'}, status=400)
    gdef, is_standalone = _resolve_group(group_by, entity)
    if not gdef:
        return JsonResponse({'error': f'Недопустимое поле: {group_by!r}'}, status=400)

    metric = config.get('metric') or {'agg': 'count'}
    agg = (metric.get('agg') or 'count').lower()
    if agg not in ('count', 'sum', 'avg', 'count_distinct'):
        return JsonResponse({'error': f'Недопустимая агрегация: {agg!r}'}, status=400)

    aggregatable = get_aggregatable_fields(entity)
    metric_field_key = metric.get('field')
    metric_field_orm = None
    metric_label = {
        'count': 'Количество',
        'count_distinct': 'Уникальных',
        'sum': 'Сумма',
        'avg': 'Среднее',
    }[agg]
    if agg in ('sum', 'avg', 'count_distinct'):
        if not metric_field_key or metric_field_key not in aggregatable:
            return JsonResponse(
                {'error': f'Метрика {agg!r} требует поле из списка: {sorted(aggregatable)}'},
                status=400,
            )
        metric_field_orm = aggregatable[metric_field_key]['orm']
        metric_label = f'{metric_label}: {aggregatable[metric_field_key]["label"]}'

    orm = gdef['orm']
    label_orm = gdef.get('label_orm')
    if is_standalone:
        qs = qs.annotate(_pie_standalone=Q(mawb__isnull=True))
        orm = '_pie_standalone'

    values_fields = [orm]
    if label_orm and label_orm != orm:
        values_fields.append(label_orm)

    if agg == 'count':
        annotated = qs.values(*values_fields).annotate(_val=Count('id', distinct=True))
    elif agg == 'count_distinct':
        annotated = qs.values(*values_fields).annotate(_val=Count(metric_field_orm, distinct=True))
    elif agg == 'sum':
        annotated = qs.values(*values_fields).annotate(_val=Sum(metric_field_orm))
    else:  # avg
        annotated = qs.values(*values_fields).annotate(_val=Avg(metric_field_orm))

    try:
        limit = max(1, min(int(config.get('limit') or 12), 50))
    except (TypeError, ValueError):
        limit = 12

    rows = list(annotated.order_by('-_val')[:limit])

    labels = []
    values = []
    group_keys = []
    for r in rows:
        code = r.get(orm)
        name_val = r.get(label_orm) if label_orm else None
        k, l = _group_key_label(gdef, code, name_val)
        labels.append(l)
        v = r.get('_val')
        try:
            values.append(float(v) if v is not None else 0)
        except (TypeError, ValueError):
            values.append(0)
        group_keys.append(k)

    return JsonResponse({
        'labels':       labels,
        'values':       values,
        'group_keys':   group_keys,
        'group_by':     group_by,
        'metric_label': metric_label,
        'agg':          agg,
    })


_DAY_RU = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
_SHP_TYPE_LABELS = {
    'IMPEX': 'IMPEX', 'B2C': 'B2C', 'B2B': 'B2B',
    'EXPRESS': 'EXPRESS', 'RTO': 'RTO',
}


def _widget_forecast_arrivals(config: dict):
    """Прогноз прилётов: агрегация Cargo по flight_date на ближайшие N дней."""
    days = min(int(config.get('days', 7)), 30)
    today = timezone.localdate()
    end   = today + _dt.timedelta(days=days - 1)

    # Базовый QS: партии с датой прилёта в диапазоне
    base_qs = Cargo.objects.filter(flight_date__gte=today, flight_date__lte=end)

    # Агрегат по дате
    by_date = (
        base_qs
        .values('flight_date')
        .annotate(
            cargo_count=Count('id', distinct=True),
            hawb_count=Count('hawbs__id'),
            weight_total=Sum('weight'),
            pieces_total=Sum('pieces_declared'),
        )
        .order_by('flight_date')
    )

    # Разбивка по shp_type по дате
    by_type = (
        base_qs.exclude(shp_type='')
        .values('flight_date', 'shp_type')
        .annotate(cnt=Count('id'))
    )
    type_map = {}  # date_str → {shp_type: count}
    for row in by_type:
        ds = row['flight_date'].isoformat()
        type_map.setdefault(ds, {})[row['shp_type']] = row['cnt']

    # Собираем все даты диапазона (включая пустые)
    result = []
    for i in range(days):
        d = today + _dt.timedelta(days=i)
        ds = d.isoformat()
        agg = next((r for r in by_date if r['flight_date'] == d), None)
        result.append({
            'date':        ds,
            'day_ru':      _DAY_RU[d.weekday()],
            'display':     d.strftime('%d.%m'),
            'cargo_count': agg['cargo_count'] if agg else 0,
            'hawb_count':  agg['hawb_count']  if agg else 0,
            'weight':      round(float(agg['weight_total'] or 0), 1) if agg else 0,
            'pieces':      agg['pieces_total'] if agg else 0,
            'by_type':     type_map.get(ds, {}),
            'is_today':    d == today,
            'is_weekend':  d.weekday() >= 5,
        })

    # Все встреченные типы для заголовков колонок
    all_types = sorted({t for tm in type_map.values() for t in tm})

    return JsonResponse({'days': result, 'types': all_types})


# ── Workload-виджеты (планирование нагрузки сотрудников) ────────────────────

_DAY_RU_SHORT = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']


def _widget_workload_dispatch(widget, request):
    """Маршрутизация workload-виджетов по типу."""
    from .services.workload import (
        compute_load_grid, compute_my_day, compute_workload_forecast,
        find_overloaded, build_rebalance_plan, cells_to_dicts, actions_to_dicts,
    )

    today = timezone.localdate()
    cfg = widget.config or {}
    days = max(1, min(int(cfg.get('days', 7)), 30))

    if widget.widget_type == 'workload_heatmap':
        grid = compute_load_grid(today, today + _dt.timedelta(days=days - 1),
                                 include_inactive=bool(cfg.get('include_inactive', False)))
        # Группируем по пользователям — для удобства фронта
        users_order = []
        seen = set()
        for c in grid:
            if c.user_id not in seen:
                seen.add(c.user_id)
                users_order.append({'user_id': c.user_id, 'user_name': c.user_name})
        dates = []
        for i in range(days):
            d = today + _dt.timedelta(days=i)
            dates.append({
                'date': d.isoformat(),
                'display': d.strftime('%d.%m'),
                'weekday': d.weekday(),
                'day_ru': _DAY_RU_SHORT[d.weekday()],
                'is_today': d == today,
                'is_weekend': d.weekday() >= 5,
            })
        return JsonResponse({
            'users': users_order,
            'dates': dates,
            'cells': cells_to_dicts(grid),
        })

    if widget.widget_type == 'workload_my_day':
        horizon = cfg.get('horizon', 'today')
        if horizon == 'tomorrow':
            target = today + _dt.timedelta(days=1)
        else:
            target = today
        return JsonResponse(compute_my_day(request.user, target))

    if widget.widget_type == 'workload_overloaded':
        offset = int(cfg.get('date_offset', 0))
        target = today + _dt.timedelta(days=offset)
        original_target = target
        # Если target — выходной (никто не работает), ищем ближайший рабочий день
        # на 7 дней вперёд. Виджет подсвечивает, что это смещение от выбранной даты.
        from .services.workload import compute_load_grid
        probe = compute_load_grid(target, target)
        all_off = probe and all(c.capacity_minutes == 0 for c in probe)
        if all_off or not probe:
            for delta in range(1, 8):
                candidate = target + _dt.timedelta(days=delta)
                probe = compute_load_grid(candidate, candidate)
                if probe and any(c.capacity_minutes > 0 for c in probe):
                    target = candidate
                    break
        data = find_overloaded(target)
        plan = build_rebalance_plan(target)
        data['plan'] = actions_to_dicts(plan)
        data['target_date'] = target.isoformat()
        data['target_display'] = target.strftime('%d.%m.%Y')
        if target != original_target:
            data['shifted_from'] = original_target.isoformat()
            data['shift_reason'] = 'выходной — показан ближайший рабочий день'
        return JsonResponse(data)

    if widget.widget_type == 'workload_forecast':
        days = max(1, min(int(cfg.get('days', 7)), 30))
        forecast = compute_workload_forecast(today, days=days)
        # Декорируем для фронта
        for d in forecast:
            d['day_ru'] = _DAY_RU_SHORT[d['weekday']]
            dt = _dt.date.fromisoformat(d['date'])
            d['display'] = dt.strftime('%d.%m')
            d['is_today'] = (dt == today)
            d['is_weekend'] = d['weekday'] >= 5
        return JsonResponse({'days': forecast})

    return JsonResponse({'error': 'Неизвестный workload-виджет'}, status=400)


@login_required
@require_http_methods(['POST'])
def api_workload_rebalance_preview(request):
    """POST /api/v1/workload/rebalance/preview/ — построить план без записи.
    Тело: {date: 'YYYY-MM-DD'}. Возвращает {plan: [...], target_date}."""
    from .services.workload import build_rebalance_plan, actions_to_dicts
    data, err = _parse_json_body(request)
    if err:
        return err
    try:
        target = _dt.date.fromisoformat(data.get('date') or timezone.localdate().isoformat())
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Некорректная дата'}, status=400)
    plan = build_rebalance_plan(target)
    return JsonResponse({
        'target_date': target.isoformat(),
        'plan': actions_to_dicts(plan),
    })


@login_required
@require_http_methods(['POST'])
def api_workload_rebalance_apply(request):
    """POST /api/v1/workload/rebalance/apply/ — применить план.
    Тело: {date: 'YYYY-MM-DD', actions: [{hawb_id, from_user_id, to_user_id}]}.
    Возвращает {applied: N, errors: [...]}."""
    from .services.workload import apply_rebalance_plan
    data, err = _parse_json_body(request)
    if err:
        return err
    try:
        target = _dt.date.fromisoformat(data.get('date') or timezone.localdate().isoformat())
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Некорректная дата'}, status=400)
    actions = data.get('actions') or []
    if not isinstance(actions, list) or not actions:
        return JsonResponse({'error': 'actions: пустой список'}, status=400)
    result = apply_rebalance_plan(actions, by_user=request.user, target_date=target)
    return JsonResponse(result)


# ── Команда: профили и исключения ──────────────────────────────────────────

def _profile_to_dict(profile) -> dict:
    user = profile.user
    return {
        'user_id':                user.id,
        'username':               user.username,
        'full_name':              user.get_full_name() or user.username,
        'is_active':              user.is_active,
        'timezone':               profile.timezone,
        'is_active_op':           profile.is_active_op,
        'primary_role':           profile.primary_role,
        'work_schedule':          profile.work_schedule or {},
        'daily_capacity_minutes': profile.daily_capacity_minutes,
        'notes':                  profile.notes,
    }


def _exception_to_dict(exc) -> dict:
    return {
        'id':         exc.id,
        'user_id':    exc.user_id,
        'username':   exc.user.username,
        'date_from':  exc.date_from.isoformat(),
        'date_to':    exc.date_to.isoformat(),
        'kind':       exc.kind,
        'kind_label': exc.get_kind_display(),
        'note':       exc.note,
    }


@login_required
def api_team_profiles(request):
    """GET /api/v1/team/profiles/ — список профилей сотрудников (для редактора команды)."""
    from .models import UserProfile
    if not request.user.is_staff:
        return JsonResponse({'error': 'Только для is_staff'}, status=403)
    qs = UserProfile.objects.select_related('user').order_by(
        'user__last_name', 'user__username',
    )
    return JsonResponse({
        'roles':    [{'value': v, 'label': l} for v, l in ROLE_CHOICES],
        'profiles': [_profile_to_dict(p) for p in qs],
    })


@login_required
@require_http_methods(['PUT'])
def api_team_profile(request, user_id: int):
    """PUT /api/v1/team/profiles/<user_id>/ — обновить профиль сотрудника."""
    from .models import UserProfile
    if not request.user.is_staff:
        return JsonResponse({'error': 'Только для is_staff'}, status=403)
    data, err = _parse_json_body(request)
    if err:
        return err
    try:
        profile = UserProfile.objects.select_related('user').get(user_id=user_id)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Профиль не найден'}, status=404)

    # Валидация work_schedule (если передан)
    schedule = data.get('work_schedule')
    if schedule is not None:
        if not isinstance(schedule, dict):
            return JsonResponse({'error': 'work_schedule: ожидается объект'}, status=400)
        valid_keys = {'mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun'}
        for k, intervals in schedule.items():
            if k not in valid_keys:
                return JsonResponse({'error': f'Неизвестный ключ дня: {k}'}, status=400)
            if not isinstance(intervals, list):
                return JsonResponse({'error': f'{k}: ожидается список интервалов'}, status=400)
            for pair in intervals:
                if not (isinstance(pair, (list, tuple)) and len(pair) == 2):
                    return JsonResponse({'error': f'{k}: интервал должен быть [start, end]'}, status=400)
                # Минимальная синтаксическая проверка HH:MM
                for t in pair:
                    if not isinstance(t, str) or len(t) != 5 or t[2] != ':':
                        return JsonResponse({'error': f'{k}: время в формате HH:MM'}, status=400)
        profile.work_schedule = schedule

    if 'timezone' in data:
        tz = (data['timezone'] or '').strip() or 'Europe/Moscow'
        profile.timezone = tz
    if 'is_active_op' in data:
        profile.is_active_op = bool(data['is_active_op'])
    if 'primary_role' in data:
        role = data['primary_role'] or ''
        if role and role not in dict(ROLE_CHOICES):
            return JsonResponse({'error': 'Неизвестная роль'}, status=400)
        profile.primary_role = role
    if 'daily_capacity_minutes' in data:
        try:
            profile.daily_capacity_minutes = max(0, int(data['daily_capacity_minutes']))
        except (TypeError, ValueError):
            return JsonResponse({'error': 'daily_capacity_minutes: число'}, status=400)
    if 'notes' in data:
        profile.notes = (data['notes'] or '')[:200]

    profile.save()
    return JsonResponse(_profile_to_dict(profile))


@login_required
@require_http_methods(['GET', 'POST'])
def api_team_exceptions(request):
    """GET — список исключений (с фильтрами ?user=N&from=X&to=Y).
    POST — создать новое исключение."""
    from .models import WorkScheduleException
    if not request.user.is_staff:
        return JsonResponse({'error': 'Только для is_staff'}, status=403)

    if request.method == 'GET':
        qs = WorkScheduleException.objects.select_related('user')
        if request.GET.get('user'):
            try:
                qs = qs.filter(user_id=int(request.GET['user']))
            except ValueError:
                pass
        if request.GET.get('from'):
            try:
                qs = qs.filter(date_to__gte=_dt.date.fromisoformat(request.GET['from']))
            except ValueError:
                pass
        if request.GET.get('to'):
            try:
                qs = qs.filter(date_from__lte=_dt.date.fromisoformat(request.GET['to']))
            except ValueError:
                pass
        return JsonResponse({'items': [_exception_to_dict(e) for e in qs[:500]]})

    # POST
    data, err = _parse_json_body(request)
    if err:
        return err
    try:
        df = _dt.date.fromisoformat(data['date_from'])
        dt = _dt.date.fromisoformat(data['date_to'])
    except (KeyError, TypeError, ValueError):
        return JsonResponse({'error': 'date_from / date_to: YYYY-MM-DD'}, status=400)
    if dt < df:
        return JsonResponse({'error': 'date_to раньше date_from'}, status=400)
    try:
        user_id = int(data['user_id'])
        target_user = User.objects.get(id=user_id)
    except (KeyError, TypeError, ValueError, User.DoesNotExist):
        return JsonResponse({'error': 'user_id некорректен'}, status=400)
    kind = data.get('kind', 'vacation')
    if kind not in dict(WorkScheduleException.KIND_CHOICES):
        return JsonResponse({'error': 'Неизвестный тип'}, status=400)
    exc = WorkScheduleException.objects.create(
        user=target_user,
        date_from=df, date_to=dt,
        kind=kind,
        note=(data.get('note') or '')[:200],
    )
    return JsonResponse(_exception_to_dict(exc))


@login_required
@require_http_methods(['DELETE'])
def api_team_exception(request, exception_id: int):
    """DELETE /api/v1/team/exceptions/<id>/ — удалить исключение."""
    from .models import WorkScheduleException
    if not request.user.is_staff:
        return JsonResponse({'error': 'Только для is_staff'}, status=403)
    deleted, _ = WorkScheduleException.objects.filter(id=exception_id).delete()
    if not deleted:
        return JsonResponse({'error': 'Не найдено'}, status=404)
    return JsonResponse({'ok': True})


@login_required
def team_page(request):
    """Страница управления командой: графики, таймзоны, отпуска."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Только для staff-пользователей')
    return render(request, 'cargo/team.html')


# ── Настройки организации (для печатных форм) ───────────────────────────────

@login_required
@require_http_methods(['GET', 'PUT'])
def api_organization_settings(request):
    """GET / PUT /api/v1/team/organization/ — реквизиты организации (singleton)."""
    from .models import OrganizationSettings
    if not request.user.is_staff:
        return JsonResponse({'error': 'Только для is_staff'}, status=403)
    org = OrganizationSettings.get_solo()
    # Поля и их максимальные длины
    fields = {
        'name': 200, 'inn': 20, 'ogrn': 20,
        'bank_account': 30, 'bank_name': 200,
        'bank_corr_account': 30, 'bank_bik': 20,
    }
    if request.method == 'PUT':
        data, err = _parse_json_body(request)
        if err:
            return err
        for f, max_len in fields.items():
            if f in data:
                setattr(org, f, (data[f] or '')[:max_len])
        org.save()
    return JsonResponse({
        **{f: getattr(org, f) for f in fields},
        'updated_at': org.updated_at.isoformat() if org.updated_at else None,
    })


# ── Экспорт ДО1 (опись товаров партии) ──────────────────────────────────────

@login_required
def cargo_export_do1(request, awb_number: str):
    """GET /cargo/<awb>/export/do1/ — выгрузка ДО1 (опись товаров) в Excel."""
    from .models import OrganizationSettings, HAWBGood
    from .services.do1_export import build_do1_workbook

    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    org = OrganizationSettings.get_solo()

    # HAWBGood'ы партии в порядке: HAWB.hawb_number, потом id
    goods_qs = (
        HAWBGood.objects
        .filter(hawb__mawb=cargo)
        .select_related('hawb')
        .order_by('hawb__hawb_number', 'id')
    )

    wb = build_do1_workbook(cargo=cargo, goods=list(goods_qs), organization=org)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_awb = (cargo.awb_number or 'do1').replace('/', '_')
    response['Content-Disposition'] = f'attachment; filename="DO1 {safe_awb}.xlsx"'
    wb.save(response)
    return response


@login_required
def cargo_export_manifest(request, awb_number: str):
    """GET /cargo/<awb>/export/manifest/ — Грузовой манифест (Таиланд) в Excel."""
    from .models import OrganizationSettings, HAWBGood
    from .services.manifest_export import build_manifest_workbook

    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    org = OrganizationSettings.get_solo()

    goods_qs = (
        HAWBGood.objects
        .filter(hawb__mawb=cargo)
        .select_related('hawb')
        .order_by('hawb__hawb_number', 'id')
    )

    wb = build_manifest_workbook(cargo=cargo, goods=list(goods_qs), organization=org)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_awb = (cargo.awb_number or 'manifest').replace('/', '_')
    response['Content-Disposition'] = f'attachment; filename="Manifest {safe_awb}.xlsx"'
    wb.save(response)
    return response


# ── Экспорт XML для Альта-ГТД ───────────────────────────────────────────────

@login_required
def hawb_export_alta_xml(request, hawb_id: int):
    """GET /hawb/<id>/export/alta/ — WayBillExpressIndividual XML для Альта-ГТД.

    Берёт реквизиты участника ВЭД и таможенного представителя из переменных
    окружения (см. .env.example, блок ALTA_*). Возвращает готовый Envelope,
    подписан он будет уже Альта-Подписью после импорта в Альта-ГТД.
    """
    import os
    from .models import HouseWaybill
    from .services.alta import envelope
    from .services.alta.generators import waybill_individual

    hawb = get_object_or_404(
        HouseWaybill.objects.select_related('mawb'),
        pk=hawb_id,
    )

    body = waybill_individual.build(
        hawb,
        carrier_name=os.environ.get('ALTA_CARRIER_NAME', 'ТЕСТ-ПЕРЕВОЗЧИК'),
        carrier_cert_number=os.environ.get('ALTA_CARRIER_CERT', '0000/00'),
        carrier_inn=os.environ.get('ALTA_CARRIER_INN', ''),
        carrier_okpo=os.environ.get('ALTA_CARRIER_OKPO', ''),
        carrier_legal_city=os.environ.get('ALTA_CARRIER_CITY', ''),
        carrier_legal_street=os.environ.get('ALTA_CARRIER_STREET', ''),
        carrier_fact_city=os.environ.get('ALTA_CARRIER_CITY', ''),
        carrier_fact_street=os.environ.get('ALTA_CARRIER_STREET', ''),
    )

    if request.GET.get('raw') == '1':
        from lxml import etree
        xml_bytes = etree.tostring(body, xml_declaration=True, encoding='UTF-8', pretty_print=True, standalone=False)
    else:
        xml_bytes = envelope.wrap(
            body_element=body,
            message_type='ED.1002018',
            participant_id=os.environ.get('ALTA_PARTICIPANT_ID', '0000000000000'),
            receiver_customs_code=os.environ.get('ALTA_CUSTOMS_CODE', '10005030'),
        )

    safe_num = (hawb.hawb_number or 'hawb').replace('/', '_').replace('\\', '_')
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="WayBill_{safe_num}.xml"'
    return response


@login_required
def hawb_export_alta_indpost(request, hawb_id: int):
    """GET /hawb/<id>/export/alta/indpost/ — почтовая накладная (Alta-диалект).

    Внутренний XML-формат Альты (корень <AltaIndPost>, cp1251). Загружается
    в Альту через hot-folder и служит источником для формирования реестра
    экспресс-грузов (ДТЭГ).
    """
    import os
    from .models import HouseWaybill
    from .services.alta.generators import indpost

    hawb = get_object_or_404(
        HouseWaybill.objects.select_related('mawb'),
        pk=hawb_id,
    )

    xml_bytes = indpost.build(
        hawb,
        customs_code=os.environ.get('ALTA_CUSTOMS_CODE', ''),
        origin_country=os.environ.get('ALTA_DEFAULT_ORIGIN_COUNTRY', 'CN'),
    )

    safe_num = (hawb.hawb_number or 'indpost').replace('/', '_').replace('\\', '_')
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=windows-1251')
    response['Content-Disposition'] = f'attachment; filename="IndPost_{safe_num}.xml"'
    return response


@login_required
def hawb_export_alta_invoice(request, hawb_id: int):
    """GET /hawb/<id>/export/alta/invoice/ — Invoice (коммерческий инвойс) XML.

    Собирает инвойс по одной HAWB: данные продавца/покупателя из shipper_*
    и consignee_*, товары из HAWBGood, условия поставки — заглушка FCA.
    """
    import os
    from .models import HouseWaybill
    from .services.alta import envelope
    from .services.alta.generators import invoice

    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)

    body = invoice.build(hawb)

    if request.GET.get('raw') == '1':
        from lxml import etree
        xml_bytes = etree.tostring(body, xml_declaration=True, encoding='UTF-8', pretty_print=True, standalone=False)
    else:
        xml_bytes = envelope.wrap(
            body_element=body,
            message_type='ED.1002007',
            participant_id=os.environ.get('ALTA_PARTICIPANT_ID', '0000000000000'),
            receiver_customs_code=os.environ.get('ALTA_CUSTOMS_CODE', '10005030'),
        )

    safe_num = (hawb.hawb_number or 'invoice').replace('/', '_').replace('\\', '_')
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{safe_num}.xml"'
    return response


@login_required
def cargo_export_alta_dt(request, awb_number: str):
    """GET /cargo/<awb>/export/alta/dt/ — ESADout_CU XML (классическая ДТ).

    Минимально валидная декларация на товары по схеме 5.27.0.
    Для боевой подачи требует дозаполнения реальной ДТ-формой в Альте.
    """
    import os
    from .services.alta import envelope
    from .services.alta.generators import goods_declaration

    cargo = get_object_or_404(Cargo, awb_number=awb_number)

    body = goods_declaration.build(
        cargo,
        declarant_name=os.environ.get('ALTA_CARRIER_NAME', ''),
        declarant_inn=os.environ.get('ALTA_CARRIER_INN', ''),
    )

    if request.GET.get('raw') == '1':
        from lxml import etree
        xml_bytes = etree.tostring(body, xml_declaration=True, encoding='UTF-8', pretty_print=True, standalone=False)
    else:
        xml_bytes = envelope.wrap(
            body_element=body,
            message_type='ED.1006107',
            participant_id=os.environ.get('ALTA_PARTICIPANT_ID', '0000000000000'),
            receiver_customs_code=os.environ.get('ALTA_CUSTOMS_CODE', '10005030'),
        )

    safe_awb = (cargo.awb_number or 'dt').replace('/', '_').replace('\\', '_')
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="DT_{safe_awb}.xml"'
    return response


@login_required
def cargo_export_alta_express(request, awb_number: str):
    """GET /cargo/<awb>/export/alta/ — ExpressCargoDeclaration XML для Альта-ГТД.

    Собирает декларацию на экспресс-грузы (ДТЭГ) из Cargo + всех его HAWB.
    Реквизиты участника ВЭД и таможни — из переменных окружения (.env).
    """
    import os
    from .services.alta import envelope
    from .services.alta.generators import express_declaration

    cargo = get_object_or_404(Cargo, awb_number=awb_number)

    body = express_declaration.build(cargo)

    if request.GET.get('raw') == '1':
        from lxml import etree
        xml_bytes = etree.tostring(body, xml_declaration=True, encoding='UTF-8', pretty_print=True, standalone=False)
    else:
        xml_bytes = envelope.wrap(
            body_element=body,
            message_type='ED.1006275',
            participant_id=os.environ.get('ALTA_PARTICIPANT_ID', '0000000000000'),
            receiver_customs_code=os.environ.get('ALTA_CUSTOMS_CODE', '10005030'),
        )

    safe_awb = (cargo.awb_number or 'declaration').replace('/', '_').replace('\\', '_')
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="ExpressDecl_{safe_awb}.xml"'
    return response


# ── Отправка в hot-folder Альты через очередь ──────────────────────────────

@login_required
@require_POST
def hawb_send_alta_indpost(request, hawb_id: int):
    from .models import HouseWaybill
    from .services.alta import queue as alta_queue
    hawb = get_object_or_404(HouseWaybill.objects.select_related('mawb'), pk=hawb_id)
    item = alta_queue.enqueue_indpost(hawb, user=request.user)
    messages.success(request, f'Накладная поставлена в очередь Альты (#{item.pk}).')
    return redirect('hawb_detail', hawb_id=hawb_id)


@login_required
@require_POST
def hawb_send_alta_waybill(request, hawb_id: int):
    from .models import HouseWaybill
    from .services.alta import queue as alta_queue
    hawb = get_object_or_404(HouseWaybill.objects.select_related('mawb'), pk=hawb_id)
    item = alta_queue.enqueue_waybill(hawb, user=request.user)
    messages.success(request, f'Накладная ЭД-2 поставлена в очередь Альты (#{item.pk}).')
    return redirect('hawb_detail', hawb_id=hawb_id)


@login_required
@require_POST
def hawb_send_alta_invoice(request, hawb_id: int):
    from .models import HouseWaybill
    from .services.alta import queue as alta_queue
    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)
    item = alta_queue.enqueue_invoice(hawb, user=request.user)
    messages.success(request, f'Инвойс поставлен в очередь Альты (#{item.pk}).')
    return redirect('hawb_detail', hawb_id=hawb_id)


@login_required
@require_POST
def hawb_refresh_cdek(request, hawb_id: int):
    """Ручной on-demand fetch статуса СДЭК по im_number (=hawb_number)."""
    from django.conf import settings
    from .models import HouseWaybill
    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)
    if not getattr(settings, 'CDEK_ENABLED', False):
        messages.error(request, 'Интеграция СДЭК выключена (CDEK_ENABLED).')
        return redirect('hawb_detail', hawb_id=hawb_id)
    from .services.cdek import applier as cdek_applier
    from .services.cdek.client import CdekConfigError
    try:
        res = cdek_applier.fetch_and_apply(hawb, source='manual')
    except CdekConfigError as e:
        messages.error(request, f'СДЭК не сконфигурирован: {e}')
        return redirect('hawb_detail', hawb_id=hawb_id)
    except Exception as e:
        logger.exception('cdek manual refresh failed for HAWB %s', hawb_id)
        messages.error(request, f'Ошибка обновления СДЭК: {e}')
        return redirect('hawb_detail', hawb_id=hawb_id)
    if res is None:
        messages.info(request, f'Заказ СДЭК для {hawb.hawb_number} не найден.')
    elif res:
        messages.success(request, f'Статус СДЭК обновлён: {hawb.cdek_status_display}')
    else:
        messages.info(request, 'Статус СДЭК уже актуален.')
    return redirect('hawb_detail', hawb_id=hawb_id)


@login_required
@require_POST
def cargo_send_alta_express(request, awb_number: str):
    from .services.alta import queue as alta_queue
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    item = alta_queue.enqueue_express(cargo, user=request.user)
    messages.success(request, f'Реестр экспресс-грузов поставлен в очередь Альты (#{item.pk}).')
    return redirect('cargo_detail', awb_number=awb_number)


@login_required
@require_POST
def cargo_send_alta_dt(request, awb_number: str):
    from .services.alta import queue as alta_queue
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    item = alta_queue.enqueue_dt(cargo, user=request.user)
    messages.success(request, f'ДТ поставлена в очередь Альты (#{item.pk}).')
    return redirect('cargo_detail', awb_number=awb_number)


@login_required
def alta_queue_page(request):
    """Страница очереди — что отправили, что ушло, что упало."""
    from .models import AltaQueueItem
    qs = AltaQueueItem.objects.select_related('hawb', 'cargo', 'created_by') \
                              .order_by('-created_at')[:300]
    return render(request, 'cargo/alta_queue.html', {
        'items': qs,
        'pending_count': AltaQueueItem.objects.filter(status='pending').count(),
        'failed_count':  AltaQueueItem.objects.filter(status='failed').count(),
    })


@login_required
@require_POST
def alta_queue_retry(request, item_id: int):
    """Пометить failed/sent запись как pending — агент заберёт повторно."""
    from .models import AltaQueueItem
    item = get_object_or_404(AltaQueueItem, pk=item_id)
    item.status = 'pending'
    item.error_message = ''
    item.sent_at = None
    item.save(update_fields=['status', 'error_message', 'sent_at'])
    messages.success(request, f'Документ #{item.pk} возвращён в очередь.')
    return redirect('alta_queue_page')


# ── API для агента Альты ──────────────────────────────────────────────────

def _check_alta_agent_token(request) -> bool:
    """Проверяет заголовок Authorization: Bearer <ALTA_AGENT_TOKEN>."""
    import os
    expected = (os.environ.get('ALTA_AGENT_TOKEN') or '').strip()
    if not expected:
        return False
    header = request.headers.get('Authorization', '')
    if not header.startswith('Bearer '):
        return False
    provided = header[len('Bearer '):].strip()
    # constant-time compare
    import hmac
    return hmac.compare_digest(provided.encode(), expected.encode())


@api_view(['GET'])
@permission_classes([])
@authentication_classes([])
def api_alta_queue_list(request):
    """GET /api/v1/alta/queue/ — список pending-документов для агента."""
    if not _check_alta_agent_token(request):
        return Response({'detail': 'Unauthorized'}, status=401)
    from .models import AltaQueueItem
    qs = AltaQueueItem.objects.filter(status='pending').order_by('created_at')[:50]
    return Response([{
        'id': it.pk,
        'doc_type': it.doc_type,
        'filename': it.filename,
        'encoding': it.content_encoding,
        'created_at': it.created_at.isoformat(),
        'size_bytes': len(it.content) if it.content else 0,
    } for it in qs])


@api_view(['GET'])
@permission_classes([])
@authentication_classes([])
def api_alta_queue_file(request, item_id: int):
    """GET /api/v1/alta/queue/<id>/file/ — XML-байты в исходной кодировке."""
    if not _check_alta_agent_token(request):
        return HttpResponse('Unauthorized', status=401)
    from .models import AltaQueueItem
    item = get_object_or_404(AltaQueueItem, pk=item_id)
    response = HttpResponse(
        bytes(item.content) if item.content else b'',
        content_type=f'application/xml; charset={item.content_encoding}',
    )
    response['Content-Disposition'] = f'attachment; filename="{item.filename}"'
    response['X-Alta-Filename'] = item.filename
    response['X-Alta-Encoding'] = item.content_encoding
    return response


@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def api_alta_queue_ack(request, item_id: int):
    """POST /api/v1/alta/queue/<id>/ack/ — агент успешно положил файл в hot-folder."""
    if not _check_alta_agent_token(request):
        return Response({'detail': 'Unauthorized'}, status=401)
    from .models import AltaQueueItem
    item = get_object_or_404(AltaQueueItem, pk=item_id)
    item.mark_sent()
    return Response({'id': item.pk, 'status': item.status})


@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def api_alta_queue_fail(request, item_id: int):
    """POST /api/v1/alta/queue/<id>/fail/ — агент не смог обработать (FS error и т.д.)."""
    if not _check_alta_agent_token(request):
        return Response({'detail': 'Unauthorized'}, status=401)
    from .models import AltaQueueItem
    item = get_object_or_404(AltaQueueItem, pk=item_id)
    message = (request.data.get('message') or '')[:5000] if hasattr(request, 'data') else ''
    item.mark_failed(message)
    return Response({'id': item.pk, 'status': item.status, 'retry_count': item.retry_count})


def _check_alta_inbox_token(request) -> bool:
    """Проверяет Authorization: Bearer <ALTA_INBOX_TOKEN>. Отдельный токен от queue."""
    import os, hmac
    expected = (os.environ.get('ALTA_INBOX_TOKEN') or '').strip()
    if not expected:
        return False
    header = request.headers.get('Authorization', '')
    if not header.startswith('Bearer '):
        return False
    provided = header[len('Bearer '):].strip()
    return hmac.compare_digest(provided.encode(), expected.encode())


@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def api_alta_inbox_post(request):
    """POST /api/v1/alta/inbox/ — приём входящего ЭД-сообщения от агента.

    Body (JSON):
      envelope_id   str  (обязательно, unique-ключ для идемпотентности)
      msg_type      str  (ED.xxx)
      prepared_at   str  (ISO datetime, опц)
      waybill_number str (опц — из тела XML, по нему матчим HAWB)
      declaration_number str (опц — если регистрация)
      raw_xml       str  (опц)
      parsed_meta   dict (опц — что ещё парсер выкусил)

    Возвращает 200 на любой повтор по envelope_id (idempotent).
    """
    if not _check_alta_inbox_token(request):
        return Response({'detail': 'Unauthorized'}, status=401)

    data = request.data if hasattr(request, 'data') else {}
    envelope_id = (data.get('envelope_id') or '').strip()
    if not envelope_id:
        return Response({'detail': 'envelope_id required'}, status=400)

    # Тех.шум — мы не используем эти типы, не сохраняем (юзер давно просил).
    # CMN.00003 — envelope ACK (29k/сутки), CMN.00006 — receipt ack,
    # ED.11001/11002 — receipt confirmation. None из них не матчится с
    # Cargo/HAWB, только засоряет БД и orphan-счётчик.
    msg_type = (data.get('msg_type') or '').strip()
    NOISE_TYPES = {'CMN.00003', 'CMN.00006', 'ED.11001', 'ED.11002'}
    if msg_type in NOISE_TYPES:
        return Response({'ok': True, 'note': 'noise type, skipped'})

    from .models import AltaInboxMessage
    from .services.alta.inbox import dispatch

    prepared_at = data.get('prepared_at')
    if prepared_at:
        try:
            from django.utils.dateparse import parse_datetime
            prepared_at = parse_datetime(prepared_at)
        except (ValueError, TypeError):
            prepared_at = None

    # Если агент прислал raw_xml — обогащаем parsed_meta сервер-сайдным
    # парсером (parse_raw_xml). Так SVH-поля доедут до dispatch даже если
    # на рабочем сервере крутится старый agent без svh-парсинга.
    raw_xml = data.get('raw_xml') or ''
    parsed_meta = data.get('parsed_meta') or {}
    if raw_xml:
        try:
            from .services.alta.xml_extract import parse_raw_xml
            from_xml = parse_raw_xml(raw_xml)
            # Агент в приоритете: его поля не перетираем
            merged = {**{k: v for k, v in from_xml.items() if v}, **parsed_meta}
            parsed_meta = merged
        except Exception:
            # парсер никогда не должен ронять view
            pass

    msg, created = AltaInboxMessage.objects.update_or_create(
        envelope_id=envelope_id,
        defaults={
            'msg_type': (data.get('msg_type') or '').strip()[:32],
            'waybill_number_raw': (data.get('waybill_number') or '').strip()[:64],
            'declaration_number': (data.get('declaration_number') or '').strip()[:64],
            'prepared_at': prepared_at,
            'raw_xml': raw_xml,
            'parsed_meta': parsed_meta,
        },
    )
    if created:
        dispatch(msg)
    return Response({
        'id': msg.pk,
        'envelope_id': msg.envelope_id,
        'created': created,
        'msg_kind': msg.msg_kind,
        'hawb_id': msg.hawb_id,
        'cargo_id': msg.cargo_id,
        'status_applied': msg.status_applied,
    })


@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def api_alta_inbox_missing(request):
    """POST /api/v1/alta/inbox/missing/

    Body (JSON):
      envelope_ids: list[str]  — массив envelope_id из БД Альты

    Returns:
      missing: list[str] — те envelope_id что отсутствуют в AltaInboxMessage

    Используется db_reconcile-потоком агента для сверки: что у нас нет —
    то агент потом дотянет POST /api/v1/alta/inbox/. Дедуп на envelope_id
    обеспечен уникальным индексом AltaInboxMessage.envelope_id.

    Лимит на один запрос: 5000 envelope_id (чтобы IN-выборка SQLite не
    распухала). Агент пусть чанкует.
    """
    if not _check_alta_inbox_token(request):
        return Response({'detail': 'Unauthorized'}, status=401)

    data = request.data if hasattr(request, 'data') else {}
    envs = data.get('envelope_ids') or []
    if not isinstance(envs, list):
        return Response({'detail': 'envelope_ids must be a list'}, status=400)
    envs = [str(e).strip() for e in envs if e]
    if not envs:
        return Response({'missing': []})
    if len(envs) > 5000:
        return Response(
            {'detail': 'max 5000 envelope_ids per request'}, status=400)

    from .models import AltaInboxMessage
    # Case-insensitive дедуп. Берём из БД через IN — но envelopeid в Альте
    # бывают и в lowercase и в UPPERCASE. У нас хранится как пришло, без
    # нормализации. Делаем сравнение по lower.
    envs_set = {e.lower() for e in envs}
    existing = set(
        AltaInboxMessage.objects.filter(envelope_id__in=envs)
        .values_list('envelope_id', flat=True))
    # Дополнительно проверим lowercase-варианты (на случай если в БД
    # хранится в одном регистре, а агент шлёт в другом).
    existing_lower = {e.lower() for e in existing}
    missing = [e for e in envs if e.lower() not in existing_lower]
    return Response({'missing': missing})


@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def api_alta_outbox_post(request):
    """POST /api/v1/alta/outbox/ — наблюдение исходящей копии Альты.

    Агент читает `538134^*.gz` из C:\\GTDSERV\\ED\\IN и шлёт сюда:
      envelope_id            UUID (обязательно, unique)
      msg_type               'CMN.00202', 'ED.1002018' и т.д.
      prepared_at            ISO datetime, опц
      common_waybill_number  MAWB номер (опц)
      waybill_number         HAWB номер (опц)
      parsed_meta            dict — прочие поля

    Service ищет Cargo по common_waybill_number и/или HAWB по waybill_number.
    Используем тот же ALTA_INBOX_TOKEN — это два потока одного агента.
    """
    if not _check_alta_inbox_token(request):
        return Response({'detail': 'Unauthorized'}, status=401)

    data = request.data if hasattr(request, 'data') else {}
    envelope_id = (data.get('envelope_id') or '').strip()
    if not envelope_id:
        return Response({'detail': 'envelope_id required'}, status=400)

    from .models import AltaOutboxObservation
    from .services.alta.outbox import dispatch as outbox_dispatch

    prepared_at = data.get('prepared_at')
    if prepared_at:
        try:
            from django.utils.dateparse import parse_datetime
            prepared_at = parse_datetime(prepared_at)
        except (ValueError, TypeError):
            prepared_at = None

    msg_type = (data.get('msg_type') or '').strip()[:32]
    common_wb = (data.get('common_waybill_number') or '').strip()[:64]
    parsed_meta = data.get('parsed_meta') or {}

    # ED.DO1: серверный re-парсинг raw_xml через xml_extract.parse_do1_report.
    # Регекс агента не справляется с MAWB-блоком (там доп.теги Avia/FlightNumber
    # ломают плоский regex). На стороне VPS у нас block-based parser устойчив
    # к этому. Перезаписываем common_wb/hawbs из серверного парсинга если оно
    # дало значения.
    if msg_type == 'ED.DO1' and parsed_meta.get('raw_xml'):
        try:
            from .services.alta.xml_extract import parse_do1_report
            parsed = parse_do1_report(parsed_meta['raw_xml'])
            if parsed.get('mawb'):
                common_wb = parsed['mawb'][:64]
            if parsed.get('hawbs'):
                parsed_meta['hawbs'] = parsed['hawbs']
            if parsed.get('report_number'):
                parsed_meta['report_number'] = parsed['report_number']
            if parsed.get('certificate_number'):
                parsed_meta['certificate_number'] = parsed['certificate_number']
            if parsed.get('goods'):
                parsed_meta['goods'] = parsed['goods']
        except Exception:
            import logging
            logging.getLogger('cargo.alta.outbox').exception(
                'parse_do1_report failed for %s', envelope_id)

    obs, created = AltaOutboxObservation.objects.update_or_create(
        envelope_id=envelope_id,
        defaults={
            'msg_type': msg_type,
            'prepared_at': prepared_at,
            'common_waybill_number': common_wb,
            'waybill_number': (data.get('waybill_number') or '').strip()[:64],
            'parsed_meta': parsed_meta,
        },
    )
    # Для типов которые мы обрабатываем post-link (ED.DO1 → svh_do1_sent_at,
    # CMN.11023/CMN.11349 → filed_date) dispatch идемпотентен (только
    # update_fields + writeback diff) — вызываем всегда, чтобы после
    # обновления логики/парсера re-POST уже существующих наблюдений
    # зафиксировал поля в БД.
    if created or msg_type in ('ED.DO1', 'CMN.11023', 'CMN.11349'):
        outbox_dispatch(obs)
    return Response({
        'id': obs.pk,
        'envelope_id': obs.envelope_id,
        'created': created,
        'cargo_id': obs.cargo_id,
        'hawb_id': obs.hawb_id,
    })


@api_view(['GET'])
@permission_classes([])
@authentication_classes([])
def api_alta_agent_download(request):
    """GET /api/v1/alta/agent/download/ — отдаёт текущий alta_agent.py.

    Используется как self-update: на рабочей VPS скачиваем фиксированную версию,
    кладём поверх C:\\ALTA\\IN\\alta_agent\\alta_agent.py и рестартим
    scheduled task. Файл публичный (исходный код агента не секретный).
    """
    from pathlib import Path
    from django.http import FileResponse, HttpResponse
    p = Path(__file__).resolve().parent.parent / 'alta_agent' / 'alta_agent.py'
    if not p.exists():
        return HttpResponse('agent file missing', status=404)
    return FileResponse(
        open(p, 'rb'),
        as_attachment=True,
        filename='alta_agent.py',
        content_type='text/x-python',
    )


# ── СДЭК (CDEK) webhook receiver ─────────────────────────────────────────────

def _check_cdek_webhook_secret(secret: str) -> bool:
    """Сверяет secret из пути с settings.CDEK_WEBHOOK_SECRET (constant-time)."""
    import hmac
    from django.conf import settings
    expected = (getattr(settings, 'CDEK_WEBHOOK_SECRET', '') or '').strip()
    if not expected:
        return False
    return hmac.compare_digest((secret or '').strip().encode(), expected.encode())


def _cdek_ip_allowed(request) -> bool:
    """Опциональный allowlist source-IP. Пусто → не проверяем."""
    from django.conf import settings
    allowed = getattr(settings, 'CDEK_WEBHOOK_ALLOWED_IPS', None) or []
    if not allowed:
        return True
    candidates = set()
    remote = request.META.get('REMOTE_ADDR')
    if remote:
        candidates.add(remote.strip())
    xff = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if xff:
        candidates.add(xff.split(',')[0].strip())
    return bool(candidates & set(allowed))


@api_view(['POST'])
@permission_classes([])
@authentication_classes([])
def cdek_webhook(request, secret: str):
    """POST /api/v1/cdek/webhook/<secret>/ — приёмник вебхуков СДЭК ORDER_STATUS.

    Защита: фича-гейт + неугадываемый secret в пути + опц. IP-allowlist +
    авторитетный до-запрос по uuid в dispatch (у вебхука нет подписи).

    ВСЕГДА возвращает 200 на валидный secret (даже на дубль/несматченное/
    внутреннюю ошибку), чтобы СДЭК не уходил в retry-шторм. Пропущенные
    апдейты подбирает reconcile-команда sync_cdek_statuses.
    """
    from django.conf import settings
    from .services.cdek import webhook as cdek_webhook_svc

    # Неверный secret / выключено — 404 (не раскрываем существование endpoint).
    if not getattr(settings, 'CDEK_ENABLED', False):
        return Response(status=404)
    if not _check_cdek_webhook_secret(secret):
        return Response(status=404)
    if not _cdek_ip_allowed(request):
        logger.warning('cdek webhook: IP %s не в allowlist',
                       request.META.get('REMOTE_ADDR'))
        return Response(status=404)

    data = request.data if hasattr(request, 'data') else {}
    parsed = cdek_webhook_svc.parse_payload(data)
    if not parsed:
        # Не ORDER_STATUS (например PRINT_FORM) или мусор — принимаем молча.
        return Response({'ok': True, 'ignored': True})

    try:
        result = cdek_webhook_svc.dispatch(parsed)
    except Exception:
        logger.exception('cdek webhook dispatch failed for uuid=%s',
                         parsed.get('uuid'))
        return Response({'ok': False})

    return Response({'ok': True, **result})


# ── HAWB-виджеты (для entity_type='hawb') ────────────────────────────────────

_HAWB_LOGISTICS_CSS = {
    'CREATED':         'stage-draft',
    'TO_ORIGIN_WH':    'stage-formed',
    'AT_ORIGIN_WH':    'stage-formed',
    'CONSOLIDATED':    'stage-formed',
    'READY_TO_SHIP':   'stage-dispatched',
    'EXPORT_CUSTOMS':  'stage-customs',
    'IN_TRANSIT_EXP':  'stage-dispatched',
    'ARRIVED_DEST':    'stage-arrived',
    'AT_SVH':          'stage-arrived',
    'IMPORT_CUSTOMS':  'stage-customs',
    'READY_DELIVERY':  'stage-released',
    'TO_SORT_CENTER':  'stage-released',
    'AT_SORT_CENTER':  'stage-released',
    'READY_TO_DEST':   'stage-released',
    'IN_TRANSIT_DEST': 'stage-released',
    'ARRIVED_FINAL':   'stage-released',
    'DELIVERED':       'stage-released',
    'RETURNED':        'stage-draft',
    'LOST':            'stage-draft',
}


def _hawb_widget_stat(qs, config: dict):
    metric = config.get('metric', 'count')
    if metric == 'weight':
        val = float(qs.aggregate(s=Sum('weight'))['s'] or 0)
        return JsonResponse({'value': round(val, 1), 'suffix': ' кг', 'color': '#10b981'})
    if metric == 'pieces':
        val = qs.aggregate(s=Sum('pieces_declared'))['s'] or 0
        return JsonResponse({'value': val, 'suffix': ' мест', 'color': '#f59e0b'})
    if metric == 'invoice_value':
        val = float(qs.aggregate(s=Sum('invoice_value'))['s'] or 0)
        return JsonResponse({'value': round(val, 2), 'suffix': '', 'color': '#6366f1'})
    if metric == 'standalone':
        val = qs.filter(mawb__isnull=True).count()
        return JsonResponse({'value': val, 'suffix': '', 'color': '#ef4444'})
    # default: count
    val = qs.count()
    return JsonResponse({'value': val, 'suffix': ' накл.', 'color': '#2563eb'})


def _hawb_widget_kanban(qs, config: dict):
    from .models import HouseWaybill
    statuses_filter = config.get('statuses') or [s[0] for s in HouseWaybill.LOGISTICS_STATUS_CHOICES]
    limit = min(int(config.get('limit', 20)), 50)
    columns = []
    for code, label in HouseWaybill.LOGISTICS_STATUS_CHOICES:
        if code not in statuses_filter:
            continue
        col_qs = qs.filter(logistics_status=code)
        total = col_qs.count()
        items = list(
            col_qs[:limit].values(
                'id', 'hawb_number', 'cargo_type', 'shipment_type',
                'weight', 'pieces_declared', 'consignee_name',
                'mawb__awb_number',
            )
        )
        columns.append({
            'stage': code,
            'label': label,
            'css':   _HAWB_LOGISTICS_CSS.get(code, ''),
            'total': total,
            'items': items,
        })
    return JsonResponse({'columns': columns})


def _hawb_widget_table(qs, config: dict):
    return _render_table_widget(qs, config, 'hawb')


def _hawb_widget_chart_logistics(qs):
    from .models import HouseWaybill
    labels_map = dict(HouseWaybill.LOGISTICS_STATUS_CHOICES)
    stats = qs.values('logistics_status').annotate(cnt=Count('id')).order_by('logistics_status')
    labels = [labels_map.get(s['logistics_status'], s['logistics_status']) for s in stats]
    values = [s['cnt'] for s in stats]
    group_keys = [s['logistics_status'] for s in stats]
    return JsonResponse({
        'labels': labels, 'values': values,
        'group_keys': group_keys, 'group_by': 'logistics_status',
    })


def _hawb_widget_chart_warehouse(qs):
    stats = (
        qs.exclude(mawb__warehouse_license='')
        .exclude(mawb__warehouse_license__isnull=True)
        .values('mawb__warehouse_license', 'mawb__warehouse_name')
        .annotate(cnt=Count('id'), total_w=Sum('weight'))
        .order_by('-cnt')[:12]
    )
    labels  = [s['mawb__warehouse_license'] or s['mawb__warehouse_name'] or '?' for s in stats]
    values  = [s['cnt'] for s in stats]
    weights = [float(s['total_w'] or 0) for s in stats]
    group_keys = [s['mawb__warehouse_license'] for s in stats]
    return JsonResponse({
        'labels': labels, 'values': values, 'weights': weights,
        'group_keys': group_keys, 'group_by': 'warehouse',
    })


# ═══════════════════════════════════════════════════════════════════════════════
# PIVOT WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

_PIVOT_MISSING = '__pivot_missing__'   # маркер "нет значения" (NULL / '')


def _group_key_label(gdef: dict, code, name_val=None) -> tuple:
    """Возвращает (key, label) для кода значения группировки."""
    if code is None or code == '':
        return (_PIVOT_MISSING, '(не указано)')
    choices = gdef.get('choices')
    if choices is not None:
        # ключ делаем строкой для стабильной сериализации в JSON
        return (str(code), str(choices.get(code, code)))
    if 'choices_fn' in gdef:
        fn_choices = gdef['choices_fn']()
        return (str(code), str(fn_choices.get(code, code)))
    if name_val:
        return (str(code), f'{code} — {name_val}')
    return (str(code), str(code))


def _resolve_group(gdef_key: str, entity: str):
    """Возвращает (gdef, is_standalone_special) либо (None, False) при ошибке."""
    fields = get_groupable_fields(entity)
    if gdef_key not in fields:
        return None, False
    gdef = fields[gdef_key]
    # is_standalone — особый случай: нет прямого orm-поля, группируем по mawb__isnull.
    return gdef, gdef.get('orm') == '__mawb_isnull__'


def _pivot_group_filter(gdef: dict, is_standalone: bool, key: str) -> Q:
    """Обратное преобразование ключа pivot-ячейки в Q-фильтр на QS.

    Нужен для drill-down: ключи в ответе `_compute_pivot()` — это строки,
    но в БД значение может быть int/bool/str. Для choices-полей находим
    «настоящий» ключ, иначе фильтруем по строке как есть.
    """
    if is_standalone:
        return Q(mawb__isnull=(str(key) == 'True'))
    orm = gdef['orm']
    if key == _PIVOT_MISSING:
        q = Q(**{f'{orm}__isnull': True})
        if not gdef.get('choices') and not gdef.get('choices_fn'):
            q |= Q(**{orm: ''})
        return q
    choices = gdef.get('choices')
    if choices is not None:
        for real_k in choices.keys():
            if str(real_k) == str(key):
                return Q(**{orm: real_k})
    return Q(**{orm: key})


def _build_metric_annotation(metric: dict, entity: str, user, index: int):
    """
    Возвращает dict annotations {alias: Expression} (один элемент для count/sum/avg,
    два для ratio) и метаданные {alias, name, format, agg, is_ratio, num_alias, den_alias}.
    Возбуждает CQLError на невалидной конфигурации.
    """
    agg = (metric.get('agg') or 'count').lower()
    if agg not in ALLOWED_AGGS:
        raise CQLError(f'Недопустимая агрегация: {agg!r}')

    name = metric.get('name') or f'Метрика {index + 1}'
    fmt  = metric.get('format') or 'int'
    sub_filter = (metric.get('filter') or '').strip()

    def _sub_q():
        if not sub_filter:
            return None
        return parse_cql(sub_filter, {'me': user.username}, entity_type=entity)

    aggregatable = get_aggregatable_fields(entity)

    def _build_one(a: str, field_cql: str | None, suffix: str = ''):
        """Возвращает (alias, Expression)."""
        alias = f'm{index}{suffix}'
        q = _sub_q()
        field_orm = None
        # Поля со связанными сущностями (hawb_count / mawb_count) принудительно
        # считаются через count_distinct вне зависимости от выбора агрегации.
        if field_cql and field_cql in aggregatable and aggregatable[field_cql].get('count_only'):
            a = 'count_distinct'
        if a in ('sum', 'avg', 'count_distinct'):
            if not field_cql or field_cql not in aggregatable:
                raise CQLError(f'Метрика {a!r} требует поле из списка: {sorted(aggregatable)}')
            field_orm = aggregatable[field_cql]['orm']
        kwargs = {}
        if q is not None:
            kwargs['filter'] = q
        if a == 'count':
            expr = Count('id', distinct=True, **kwargs)
        elif a == 'count_distinct':
            expr = Count(field_orm, distinct=True, **kwargs)
        elif a == 'sum':
            expr = Sum(field_orm, **kwargs)
        elif a == 'avg':
            expr = Avg(field_orm, **kwargs)
        else:
            raise CQLError(f'Недопустимая агрегация: {a!r}')
        return alias, expr

    annotations = {}
    meta = {'name': name, 'format': fmt, 'agg': agg, 'is_ratio': False}

    if agg == 'ratio':
        num = metric.get('numerator') or {}
        den = metric.get('denominator') or {}
        # под-метрики тоже ограничены ALLOWED_AGGS, но не могут сами быть ratio
        for sub in (num, den):
            sub_a = (sub.get('agg') or 'count').lower()
            if sub_a == 'ratio' or sub_a not in ALLOWED_AGGS:
                raise CQLError(f'Недопустимая агрегация в ratio: {sub_a!r}')
        num_meta = {
            'agg': (num.get('agg') or 'count').lower(),
            'field': num.get('field'),
            'filter': (num.get('filter') or '').strip(),
        }
        den_meta = {
            'agg': (den.get('agg') or 'count').lower(),
            'field': den.get('field'),
            'filter': (den.get('filter') or '').strip(),
        }
        # временно подменяем sub_filter для каждой ветви
        for suffix, sub_meta in (('n', num_meta), ('d', den_meta)):
            inner_metric = {
                'agg':    sub_meta['agg'],
                'field':  sub_meta['field'],
                'filter': sub_meta['filter'],
            }
            # переиспользуем _build_one, но с "псевдо-индексом"
            _orig_filter = sub_filter
            sub_filter_cur = sub_meta['filter']

            def _sub_q_local(_f=sub_filter_cur):
                if not _f:
                    return None
                return parse_cql(_f, {'me': user.username}, entity_type=entity)

            q = _sub_q_local()
            kwargs = {}
            if q is not None:
                kwargs['filter'] = q
            a = sub_meta['agg']
            field_orm = None
            if a in ('sum', 'avg', 'count_distinct'):
                f_cql = sub_meta['field']
                if not f_cql or f_cql not in aggregatable:
                    raise CQLError(
                        f'Метрика {a!r} внутри ratio требует поле из списка: {sorted(aggregatable)}'
                    )
                field_orm = aggregatable[f_cql]['orm']
            if a == 'count':
                expr = Count('id', distinct=True, **kwargs)
            elif a == 'count_distinct':
                expr = Count(field_orm, distinct=True, **kwargs)
            elif a == 'sum':
                expr = Sum(field_orm, **kwargs)
            elif a == 'avg':
                expr = Avg(field_orm, **kwargs)
            else:
                raise CQLError(f'Недопустимая агрегация: {a!r}')
            alias = f'm{index}{suffix}'
            annotations[alias] = expr

        meta.update({
            'is_ratio': True,
            'num_alias': f'm{index}n',
            'den_alias': f'm{index}d',
        })
    else:
        alias, expr = _build_one(agg, metric.get('field'))
        annotations[alias] = expr
        meta['alias'] = alias

    return annotations, meta


def _format_metric_value(raw, fmt: str):
    """Возвращает число, подходящее для JSON; форматирование делает фронт."""
    if raw is None:
        return None
    try:
        if fmt in ('int',):
            return int(raw)
        return float(raw)
    except (TypeError, ValueError):
        return None


def _resolve_axis(axis_keys: list, entity: str, side: str) -> list:
    """Резолвит список ключей измерения (rows или cols) в список specs.

    Каждый spec: {gdef, orm, label_orm, is_standalone, field_key}.
    Для is_standalone-полей создаётся уникальное имя аннотации
    `_axis_<side>_<idx>_standalone`, которое должно быть навешено на qs
    вызывающей стороной.

    Raises CQLError при недопустимом ключе.
    """
    specs = []
    for idx, key in enumerate(axis_keys):
        gdef, is_standalone = _resolve_group(key, entity)
        if not gdef:
            raise CQLError(f'Недопустимое поле ({side}): {key!r}')
        orm = gdef['orm']
        standalone_alias = None
        if is_standalone:
            standalone_alias = f'_axis_{side}_{idx}_standalone'
            orm = standalone_alias
        specs.append({
            'field_key':        key,
            'gdef':             gdef,
            'orm':              orm,
            'label_orm':        gdef.get('label_orm'),
            'is_standalone':    is_standalone,
            'standalone_alias': standalone_alias,
        })
    return specs


def _axis_key_label_tuple(specs: list, row: dict) -> tuple:
    """Для одной строки results возвращает (tuple_keys, tuple_labels)
    по списку specs измерения."""
    keys = []
    labels = []
    for spec in specs:
        code = row.get(spec['orm'])
        name_val = row.get(spec['label_orm']) if spec['label_orm'] else None
        k, l = _group_key_label(spec['gdef'], code, name_val)
        keys.append(k)
        labels.append(l)
    return tuple(keys), tuple(labels)


def _compute_pivot(qs, config: dict, entity: str, user):
    """
    Вычисляет pivot-таблицу по настройкам из `config` и возвращает JsonResponse.

    Поддерживает как scalar `row_by`/`col_by` (legacy), так и массивы
    (multi-dim). Формат вывода сохраняется: если вход был массив с >1
    элементом, в row_labels/col_labels entries будут ключи ``keys``+``labels``
    (списки), иначе ``key``+``label`` (скаляры, совместимо со старым фронтом).
    """
    if not isinstance(config, dict):
        return JsonResponse({'error': 'Некорректная конфигурация pivot-виджета'}, status=400)

    row_by_raw = config.get('row_by')
    col_by_raw = config.get('col_by') or None

    # Нормализация к спискам
    def _to_list(v):
        if v is None:
            return []
        if isinstance(v, (list, tuple)):
            return [x for x in v if x]
        return [v] if v else []

    row_by_list = _to_list(row_by_raw)
    col_by_list = _to_list(col_by_raw)
    # Флаги: исходный вход был массивом? Влияет на формат вывода для совместимости.
    row_multi_out = isinstance(row_by_raw, (list, tuple)) and len(row_by_list) > 1
    col_multi_out = isinstance(col_by_raw, (list, tuple)) and len(col_by_list) > 1

    metrics_cfg = config.get('metrics') or []
    if not metrics_cfg:
        # Совместимость с MVP-конфигом "одна метрика на верхнем уровне"
        if config.get('agg'):
            metrics_cfg = [{
                'name':   config.get('metric_name', 'Значение'),
                'agg':    config.get('agg'),
                'field':  config.get('field'),
                'filter': config.get('metric_filter', ''),
                'format': config.get('format', 'int'),
            }]

    if not row_by_list:
        return JsonResponse({'error': 'Не задано поле строк (row_by)'}, status=400)
    if not metrics_cfg:
        return JsonResponse({'error': 'Не задана ни одна метрика'}, status=400)

    # Phase-2: totals / limits / sort / heatmap
    totals_cfg = config.get('show_totals') or {}
    show_row_total   = bool(totals_cfg.get('row'))
    show_col_total   = bool(totals_cfg.get('col'))
    show_grand_total = bool(totals_cfg.get('grand'))
    try:
        row_limit = int(config.get('row_limit') or 0) or None
        col_limit = int(config.get('col_limit') or 0) or None
    except (TypeError, ValueError):
        row_limit = col_limit = None
    sort_cfg = config.get('sort_by') or {}
    try:
        sort_metric_idx = int(sort_cfg.get('metric_idx') or 0)
    except (TypeError, ValueError):
        sort_metric_idx = 0
    sort_dir = 'asc' if (sort_cfg.get('dir') or 'desc').lower() == 'asc' else 'desc'
    heatmap_cfg = config.get('heatmap') or {}
    heatmap_enabled = bool(heatmap_cfg.get('enabled'))
    try:
        heatmap_metric_idx = int(heatmap_cfg.get('metric_idx') or 0)
    except (TypeError, ValueError):
        heatmap_metric_idx = 0
    heatmap_palette = heatmap_cfg.get('palette') or 'green'

    # ── Резолвим оба измерения в списки specs ────────────────────────────────
    try:
        row_specs = _resolve_axis(row_by_list, entity, side='row')
        col_specs = _resolve_axis(col_by_list, entity, side='col') if col_by_list else []
    except CQLError as e:
        return JsonResponse({'error': str(e)}, status=400)

    # ── Группировка: составляем список ORM-полей для .values() ───────────────
    values_fields = []
    for spec in row_specs + col_specs:
        if spec['is_standalone']:
            qs = qs.annotate(**{spec['standalone_alias']: Q(mawb__isnull=True)})
        values_fields.append(spec['orm'])
        if spec['label_orm'] and spec['label_orm'] != spec['orm']:
            values_fields.append(spec['label_orm'])

    # ── Аннотации метрик ──────────────────────────────────────────────────────
    try:
        all_annotations = {}
        metric_meta = []
        for idx, m in enumerate(metrics_cfg):
            ann, meta = _build_metric_annotation(m, entity, user, idx)
            all_annotations.update(ann)
            metric_meta.append(meta)
    except CQLError as e:
        return JsonResponse({'error': str(e)}, status=400)

    try:
        # .order_by() сбрасывает Meta.ordering модели, иначе ORM подмешивает
        # поля сортировки в GROUP BY и группы рассыпаются по каждой строке.
        rows = list(
            qs.order_by().values(*values_fields).annotate(**all_annotations)
        )
    except Exception as e:
        logger.exception('pivot query failed')
        return JsonResponse({'error': f'Ошибка запроса: {e}'}, status=400)

    # ── Сбор уникальных tuple-ключей для row / col + labels ──────────────────
    row_keys_map: dict = {}   # tuple(keys) → tuple(labels)
    col_keys_map: dict = {}   # tuple(keys) → tuple(labels)
    matrix: dict = {}          # (row_tkey, col_tkey) → {alias: value}

    for r in rows:
        rk_t, rl_t = _axis_key_label_tuple(row_specs, r)
        row_keys_map.setdefault(rk_t, rl_t)
        if col_specs:
            ck_t, cl_t = _axis_key_label_tuple(col_specs, r)
            col_keys_map.setdefault(ck_t, cl_t)
        else:
            ck_t = ()
        cell_key = (rk_t, ck_t)
        cell = matrix.setdefault(cell_key, {})
        for alias in all_annotations:
            cell[alias] = r.get(alias)

    def _mk_axis_entry(tkey: tuple, tlabel: tuple, multi_out: bool, empty_label: str = 'Значение') -> dict:
        """Формирует запись для row_labels/col_labels с сохранением служебного _tkey."""
        entry = {'_tkey': tkey}
        if multi_out:
            entry['keys'] = list(tkey)
            entry['labels'] = list(tlabel)
        else:
            entry['key'] = tkey[0] if tkey else '__all__'
            entry['label'] = tlabel[0] if tlabel else empty_label
        return entry

    def _entry_missing(entry: dict) -> bool:
        """Содержит ли запись маркер _PIVOT_MISSING (хотя бы в одном измерении)."""
        return any(k == _PIVOT_MISSING for k in entry['_tkey'])

    def _entry_label_sort(entry: dict) -> tuple:
        return tuple(str(x).lower() for x in (entry.get('labels') or [entry.get('label', '')]))

    # ── Собираем row_labels / col_labels ─────────────────────────────────────
    row_labels = [_mk_axis_entry(k, v, row_multi_out) for k, v in row_keys_map.items()]
    row_labels.sort(key=lambda x: (_entry_missing(x), _entry_label_sort(x)))

    if col_specs:
        col_labels = [_mk_axis_entry(k, v, col_multi_out) for k, v in col_keys_map.items()]
        col_labels.sort(key=lambda x: (_entry_missing(x), _entry_label_sort(x)))
    else:
        col_labels = [_mk_axis_entry((), (), False)]

    metrics_out = [
        {'name': m['name'], 'format': m['format'], 'agg': m['agg']}
        for m in metric_meta
    ]

    # ── Значение ячейки по метрике (None для отсутствия) ─────────────────────
    def _cell_value(data: dict, m: dict):
        if m.get('is_ratio'):
            num = data.get(m['num_alias'])
            den = data.get(m['den_alias'])
            if num is None or not den:
                return None
            try:
                return float(num) / float(den)
            except (TypeError, ValueError, ZeroDivisionError):
                return None
        return _format_metric_value(data.get(m['alias']), m['format'])

    # ── Суммируемые метрики (для итогов и сортировки) ────────────────────────
    SUMMABLE_AGGS = {'count', 'sum', 'count_distinct'}

    def _is_summable(m: dict) -> bool:
        return (not m.get('is_ratio')) and m.get('agg') in SUMMABLE_AGGS

    def _axis_sum(fixed_tkey: tuple, axis: str, m: dict):
        """Сумма значений метрики m по одной оси (row/col) — только для суммируемых."""
        if not _is_summable(m):
            return None
        others = col_labels if axis == 'row' else row_labels
        total = 0.0
        has_any = False
        for o in others:
            cell_key = (fixed_tkey, o['_tkey']) if axis == 'row' else (o['_tkey'], fixed_tkey)
            v = matrix.get(cell_key, {}).get(m['alias'])
            if v is None:
                continue
            try:
                total += float(v)
                has_any = True
            except (TypeError, ValueError):
                pass
        return total if has_any else None

    # ── Сортировка строк/колонок по выбранной метрике (если суммируема) ──────
    if metric_meta and 0 <= sort_metric_idx < len(metric_meta):
        sort_m = metric_meta[sort_metric_idx]
        if _is_summable(sort_m):
            reverse = (sort_dir != 'asc')
            row_labels.sort(
                key=lambda r: (
                    _entry_missing(r),
                    -(_axis_sum(r['_tkey'], 'row', sort_m) or 0) if reverse
                        else (_axis_sum(r['_tkey'], 'row', sort_m) or 0),
                )
            )
            if col_specs:
                col_labels.sort(
                    key=lambda c: (
                        _entry_missing(c),
                        -(_axis_sum(c['_tkey'], 'col', sort_m) or 0) if reverse
                            else (_axis_sum(c['_tkey'], 'col', sort_m) or 0),
                    )
                )

    # ── Лимиты (после сортировки) ────────────────────────────────────────────
    if row_limit and row_limit > 0:
        row_labels = row_labels[:row_limit]
    if col_limit and col_limit > 0 and col_specs:
        col_labels = col_labels[:col_limit]

    # ── Матрица ячеек для видимых row/col ────────────────────────────────────
    cells = []
    for row in row_labels:
        row_line = []
        for col in col_labels:
            data = matrix.get((row['_tkey'], col['_tkey']), {})
            vals = [{'v': _cell_value(data, m)} for m in metric_meta]
            row_line.append(vals)
        cells.append(row_line)

    # ── Итоги (только для суммируемых метрик, на базе видимых ячеек) ─────────
    def _none_list():
        return [None] * len(metric_meta)

    row_totals = None
    col_totals = None
    grand_totals = None

    if show_row_total or show_col_total or show_grand_total:
        # row_totals: [[v_per_metric], ...] — по одному списку на row_labels
        row_totals_vals = [_none_list() for _ in row_labels]
        col_totals_vals = [_none_list() for _ in col_labels]
        grand_vals = _none_list()

        for mi, m in enumerate(metric_meta):
            if not _is_summable(m):
                continue
            grand_acc = 0.0
            grand_has = False
            for ri, row in enumerate(row_labels):
                row_acc = 0.0
                row_has = False
                for ci, col in enumerate(col_labels):
                    v = matrix.get((row['_tkey'], col['_tkey']), {}).get(m['alias'])
                    if v is None:
                        continue
                    try:
                        vf = float(v)
                    except (TypeError, ValueError):
                        continue
                    row_acc += vf
                    row_has = True
                    col_totals_vals[ci][mi] = (col_totals_vals[ci][mi] or 0) + vf
                    grand_acc += vf
                    grand_has = True
                if row_has:
                    row_totals_vals[ri][mi] = _format_metric_value(row_acc, m['format'])
            if grand_has:
                grand_vals[mi] = _format_metric_value(grand_acc, m['format'])
            # format col_totals post-aggregation
            for ci in range(len(col_labels)):
                raw = col_totals_vals[ci][mi]
                if raw is not None:
                    col_totals_vals[ci][mi] = _format_metric_value(raw, m['format'])

        # ── Второй проход: итоги для avg / ratio через маргинальные агрегации ──
        non_sum_idxs = [mi for mi, m in enumerate(metric_meta) if not _is_summable(m)]
        if non_sum_idxs:
            ns_anns = {}
            for mi in non_sum_idxs:
                m = metric_meta[mi]
                if m.get('is_ratio'):
                    ns_anns[m['num_alias']] = all_annotations[m['num_alias']]
                    ns_anns[m['den_alias']] = all_annotations[m['den_alias']]
                else:
                    ns_anns[m['alias']] = all_annotations[m['alias']]

            def _ns_val(data, m):
                if m.get('is_ratio'):
                    n = data.get(m['num_alias'])
                    d = data.get(m['den_alias'])
                    if n is None or not d:
                        return None
                    try:
                        return _format_metric_value(float(n) / float(d), m['format'])
                    except (TypeError, ValueError, ZeroDivisionError):
                        return None
                return _format_metric_value(data.get(m['alias']), m['format'])

            def _axis_margin_fields(specs: list) -> list:
                out = []
                for sp in specs:
                    out.append(sp['orm'])
                    if sp['label_orm'] and sp['label_orm'] != sp['orm']:
                        out.append(sp['label_orm'])
                return out

            if show_row_total:
                rv_fields = _axis_margin_fields(row_specs)
                try:
                    row_margin = {
                        _axis_key_label_tuple(row_specs, r)[0]: r
                        for r in qs.order_by().values(*rv_fields).annotate(**ns_anns)
                    }
                    for ri, row in enumerate(row_labels):
                        data = row_margin.get(row['_tkey'], {})
                        for mi in non_sum_idxs:
                            row_totals_vals[ri][mi] = _ns_val(data, metric_meta[mi])
                except Exception:
                    logger.exception('pivot non-summable row marginals failed')

            if show_col_total and col_specs:
                cv_fields = _axis_margin_fields(col_specs)
                try:
                    col_margin = {
                        _axis_key_label_tuple(col_specs, r)[0]: r
                        for r in qs.order_by().values(*cv_fields).annotate(**ns_anns)
                    }
                    for ci, col in enumerate(col_labels):
                        data = col_margin.get(col['_tkey'], {})
                        for mi in non_sum_idxs:
                            col_totals_vals[ci][mi] = _ns_val(data, metric_meta[mi])
                except Exception:
                    logger.exception('pivot non-summable col marginals failed')

            if show_grand_total:
                try:
                    grand_data = qs.aggregate(**ns_anns)
                    for mi in non_sum_idxs:
                        grand_vals[mi] = _ns_val(grand_data, metric_meta[mi])
                except Exception:
                    logger.exception('pivot non-summable grand total failed')

        if show_row_total:
            row_totals = [
                [{'v': v} for v in vals] for vals in row_totals_vals
            ]
        if show_col_total:
            col_totals = [
                [{'v': v} for v in vals] for vals in col_totals_vals
            ]
        if show_grand_total:
            grand_totals = [{'v': v} for v in grand_vals]

    # Убираем служебный _tkey из сериализуемых записей
    def _clean(entries):
        return [{k: v for k, v in e.items() if k != '_tkey'} for e in entries]

    resp = {
        'row_by':     row_by_raw,
        'col_by':     col_by_raw,
        'row_labels': _clean(row_labels),
        'col_labels': _clean(col_labels),
        'metrics':    metrics_out,
        'cells':      cells,
    }
    if row_totals is not None:
        resp['row_totals'] = row_totals
    if col_totals is not None:
        resp['col_totals'] = col_totals
    if grand_totals is not None:
        resp['grand_totals'] = grand_totals
    if heatmap_enabled and metric_meta and 0 <= heatmap_metric_idx < len(metric_meta):
        resp['heatmap'] = {
            'metric_idx': heatmap_metric_idx,
            'palette': heatmap_palette,
        }
    return JsonResponse(resp)


def _export_pivot_xlsx(widget, request):
    """Экспорт pivot-виджета в .xlsx. Переиспользует `_compute_pivot` и
    разворачивает полученную JSON-структуру в лист с шапкой (одно- или
    двухуровневой), итогами и gran-total."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl не установлен', status=500)
    import json as _json

    try:
        q = parse_cql(widget.filter_query, {'me': request.user.username},
                      entity_type=widget.entity_type)
    except CQLError as e:
        return JsonResponse({'error': f'CQL: {e}'}, status=400)

    if widget.entity_type == 'hawb':
        from .models import HouseWaybill
        qs = HouseWaybill.objects.filter(q).select_related('mawb', 'assigned_to').distinct()
    else:
        qs = Cargo.objects.filter(q).select_related('warehouse').distinct()

    resp = _compute_pivot(qs, widget.config, widget.entity_type, request.user)
    if resp.status_code != 200:
        return resp
    data = _json.loads(resp.content)

    rows    = data.get('row_labels', [])
    cols    = data.get('col_labels', [])
    metrics = data.get('metrics', [])
    cells   = data.get('cells', [])
    row_totals   = data.get('row_totals')
    col_totals   = data.get('col_totals')
    grand_totals = data.get('grand_totals')
    row_by_raw = data.get('row_by') or ''
    col_by_raw = data.get('col_by') or None

    # Multi-dim: ключ оси может быть списком; на XLSX-экспорте склеиваем через " / "
    def _axis_keys_list(v):
        if isinstance(v, (list, tuple)):
            return [k for k in v if k]
        return [v] if v else []

    groupable = get_groupable_fields(widget.entity_type)

    def _axis_by_label(keys):
        labels = [(groupable.get(k) or {}).get('label', k) for k in keys]
        return ' / '.join(labels)

    def _entry_label(entry):
        """Возвращает строковый label записи row_labels/col_labels, в т.ч. multi-dim."""
        if 'label' in entry:
            return entry.get('label') or ''
        labs = entry.get('labels') or []
        return ' / '.join(str(x) for x in labs)

    row_by_list = _axis_keys_list(row_by_raw)
    col_by_list = _axis_keys_list(col_by_raw)
    row_by = row_by_list[0] if len(row_by_list) == 1 else row_by_raw
    col_by = col_by_list[0] if len(col_by_list) == 1 else col_by_raw
    row_by_label = _axis_by_label(row_by_list) if row_by_list else ''
    col_by_label = _axis_by_label(col_by_list) if col_by_list else ''

    col_single_all = (len(cols) == 1
                      and not col_by_list
                      and cols[0].get('key') == '__all__')
    has_col_header = bool(col_by_list) and not col_single_all
    metric_count = max(1, len(metrics))
    col_count = len(cols) if has_col_header else 1
    has_row_tot = row_totals is not None
    has_col_tot = col_totals is not None
    has_grand   = grand_totals is not None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ((widget.title or 'Pivot')[:28] or 'Pivot')

    hfill   = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    subfill = PatternFill(start_color='3b5680', end_color='3b5680', fill_type='solid')
    totfill = PatternFill(start_color='e8eef7', end_color='e8eef7', fill_type='solid')
    hfont   = Font(color='FFFFFF', bold=True)
    totfont = Font(bold=True, color='1e3a5f')
    center  = Alignment(horizontal='center', vertical='center')

    def _fmt_cell(cell, *, fill=None, font=None, align=None):
        if fill:  cell.fill = fill
        if font:  cell.font = font
        if align: cell.alignment = align

    # ── Header ────────────────────────────────────────────────────────────────
    if has_col_header and metric_count > 1:
        # Двухуровневая шапка
        c = ws.cell(row=1, column=1, value=row_by_label)
        _fmt_cell(c, fill=hfill, font=hfont, align=center)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        for ci, col in enumerate(cols):
            start = 2 + ci * metric_count
            end   = start + metric_count - 1
            c = ws.cell(row=1, column=start, value=_entry_label(col))
            _fmt_cell(c, fill=hfill, font=hfont, align=center)
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for mi, m in enumerate(metrics):
                cc = ws.cell(row=2, column=start + mi, value=m.get('name', ''))
                _fmt_cell(cc, fill=subfill, font=hfont, align=center)
        if has_row_tot:
            start = 2 + col_count * metric_count
            end   = start + metric_count - 1
            c = ws.cell(row=1, column=start, value='Итого')
            _fmt_cell(c, fill=totfill, font=totfont, align=center)
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for mi, m in enumerate(metrics):
                cc = ws.cell(row=2, column=start + mi, value=m.get('name', ''))
                _fmt_cell(cc, fill=totfill, font=totfont, align=center)
        data_start_row = 3
    else:
        # Однострочная шапка
        c = ws.cell(row=1, column=1, value=row_by_label)
        _fmt_cell(c, fill=hfill, font=hfont, align=center)
        if has_col_header:
            for ci, col in enumerate(cols):
                cc = ws.cell(row=1, column=2 + ci, value=_entry_label(col))
                _fmt_cell(cc, fill=hfill, font=hfont, align=center)
            if has_row_tot:
                cc = ws.cell(row=1, column=2 + col_count, value='Итого')
                _fmt_cell(cc, fill=totfill, font=totfont, align=center)
        else:
            # Нет col_by — заголовки колонок = имена метрик
            for mi, m in enumerate(metrics):
                cc = ws.cell(row=1, column=2 + mi, value=m.get('name', ''))
                _fmt_cell(cc, fill=hfill, font=hfont, align=center)
        data_start_row = 2

    # ── Data rows ─────────────────────────────────────────────────────────────
    def _put(row, col, v, fill=None, font=None):
        cell = ws.cell(row=row, column=col, value=(v if v is not None else ''))
        if fill: cell.fill = fill
        if font: cell.font = font
        return cell

    for ri, row in enumerate(rows):
        excel_row = data_start_row + ri
        ws.cell(row=excel_row, column=1, value=_entry_label(row))
        row_cells = cells[ri] if ri < len(cells) else []
        if has_col_header:
            for ci in range(col_count):
                cell_vals = row_cells[ci] if ci < len(row_cells) else []
                for mi in range(metric_count):
                    v = (cell_vals[mi] if mi < len(cell_vals) else {}).get('v')
                    _put(excel_row, 2 + ci * metric_count + mi, v)
            if has_row_tot:
                rt = row_totals[ri] if ri < len(row_totals) else []
                for mi in range(metric_count):
                    v = (rt[mi] if mi < len(rt) else {}).get('v')
                    _put(excel_row, 2 + col_count * metric_count + mi, v,
                         fill=totfill, font=totfont)
        else:
            # Без col_by: cells[ri][0][mi]
            cell_vals = row_cells[0] if row_cells else []
            for mi in range(metric_count):
                v = (cell_vals[mi] if mi < len(cell_vals) else {}).get('v')
                _put(excel_row, 2 + mi, v)

    # ── Итоговая строка (col_totals + grand) ──────────────────────────────────
    if has_col_tot or has_grand:
        tot_row = data_start_row + len(rows)
        _put(tot_row, 1, 'Итого', fill=totfill, font=totfont)
        if has_col_header:
            if has_col_tot:
                for ci in range(col_count):
                    ct = col_totals[ci] if ci < len(col_totals) else []
                    for mi in range(metric_count):
                        v = (ct[mi] if mi < len(ct) else {}).get('v')
                        _put(tot_row, 2 + ci * metric_count + mi, v,
                             fill=totfill, font=totfont)
            if has_grand and has_row_tot:
                for mi in range(metric_count):
                    v = (grand_totals[mi] if mi < len(grand_totals) else {}).get('v')
                    _put(tot_row, 2 + col_count * metric_count + mi, v,
                         fill=totfill, font=totfont)
        else:
            # Без col_by: col_totals[0] == grand-по-метрикам
            src = col_totals if has_col_tot else [grand_totals]
            vals = src[0] if src else []
            for mi in range(metric_count):
                v = (vals[mi] if mi < len(vals) else {}).get('v')
                _put(tot_row, 2 + mi, v, fill=totfill, font=totfont)

    # ── Auto-width (учитываем merged cells) ──────────────────────────────────
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 20)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = ''.join(ch if ch.isalnum() or ch in '-_' else '_'
                         for ch in (widget.title or 'pivot'))[:40] or 'pivot'
    filename = f'pivot_{safe_title}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(['GET'])
def api_dashboard_pivot_fields(request):
    """GET /api/v1/dashboard/pivot/fields/?entity_type=cargo|hawb
    Возвращает whitelist допустимых полей группировки / агрегации."""
    entity_type = request.GET.get('entity_type', 'cargo')
    if entity_type not in ('cargo', 'hawb'):
        return JsonResponse(
            {'error': f'entity_type должен быть cargo или hawb (получено: {entity_type!r})'},
            status=400,
        )

    groupable = get_groupable_fields(entity_type)
    aggregatable = get_aggregatable_fields(entity_type)

    return JsonResponse({
        'entity_type': entity_type,
        'groupable': [
            {'key': k, 'label': v['label']}
            for k, v in groupable.items()
        ],
        'aggregatable': [
            {'key': k, 'label': v['label'], 'type': v.get('type', 'num')}
            for k, v in aggregatable.items()
        ],
        'aggs': [
            {'key': 'count',          'label': 'Количество'},
            {'key': 'count_distinct', 'label': 'Количество уникальных'},
            {'key': 'sum',            'label': 'Сумма'},
            {'key': 'avg',            'label': 'Среднее'},
            {'key': 'ratio',          'label': 'Отношение (A / B)'},
        ],
        'formats': [
            {'key': 'int',    'label': 'Целое число'},
            {'key': 'float1', 'label': 'Число (1 знак)'},
            {'key': 'pct',    'label': 'Процент'},
            {'key': 'kg',     'label': 'Килограммы'},
        ],
    })


@login_required
@require_http_methods(['GET'])
def api_widgets_entities(request):
    """GET /api/v1/widgets/entities/

    Список сущностей, по которым можно строить виджеты (для field-first UI).
    """
    return JsonResponse({'entities': get_entities()})


@login_required
@require_http_methods(['GET'])
def api_widgets_fields(request):
    """GET /api/v1/widgets/fields/

    Плоский каталог всех groupable/aggregatable-полей из всех сущностей.
    Результат сгруппирован по (key, role): одинаковые ключи из разных
    сущностей склеиваются с массивом ``entities`` для disambiguation.
    """
    flat = get_field_catalog_union()
    grouped: dict[tuple, dict] = {}
    for f in flat:
        gk = (f['key'], f['role'])
        entry = grouped.get(gk)
        entity_info = {
            'entity': f['entity'],
            'entity_label': ENTITY_REGISTRY[f['entity']]['label'],
            'label': f['label'],
        }
        if entry is None:
            grouped[gk] = {
                'key':      f['key'],
                'role':     f['role'],
                'label':    f['label'],
                'type':     f.get('type'),
                'entities': [entity_info],
            }
        else:
            entry['entities'].append(entity_info)
    return JsonResponse({'fields': list(grouped.values())})


@login_required
@require_http_methods(['GET'])
def api_widgets_drill(request):
    """GET /api/v1/widgets/drill/?entity=<key>&cql=<expr>&page=1&page_size=50

    Универсальный drill-endpoint: возвращает список записей для указанной
    сущности с применённым CQL-фильтром. Используется drill-страницей и
    виджетами, которые делают drill-down после клика (pivot/chart).
    """
    from .widget_columns import serialize_column

    entity = (request.GET.get('entity') or '').strip()
    spec = ENTITY_REGISTRY.get(entity)
    if not spec:
        return JsonResponse({'error': f'Unknown entity: {entity!r}'}, status=400)

    raw_cql = request.GET.get('cql', '') or ''
    ast = None
    try:
        if entity in ('cargo', 'hawb'):
            q = parse_cql(raw_cql, {'me': request.user.username}, entity_type=entity)
            if raw_cql.strip():
                try:
                    ast = parse_to_ast(raw_cql, entity_type=entity)
                except CQLError:
                    ast = None
        else:
            # Для сущностей вне cargo/hawb CQL пока не поддерживаем
            if raw_cql.strip():
                return JsonResponse({'error': f'CQL не поддерживается для сущности {entity!r}'}, status=400)
            q = Q()
    except CQLError as e:
        return JsonResponse({'error': f'CQL: {e}'}, status=400)

    try:
        page = max(1, int(request.GET.get('page', 1)))
        page_size = max(1, min(int(request.GET.get('page_size', 50)), 200))
    except (TypeError, ValueError):
        page, page_size = 1, 50
    offset = (page - 1) * page_size

    model = spec['model']
    qs = model.objects.filter(q)
    if entity == 'cargo':
        qs = qs.select_related('warehouse')
    elif entity == 'hawb':
        qs = qs.select_related('mawb', 'assigned_to')
    qs = qs.distinct()

    catalog = spec.get('columns') or []
    db_fields = set()
    for col in catalog:
        db_fields.update(col.get('db_fields') or [])

    row_link = None
    if entity == 'cargo':
        db_fields.add('awb_number')
        row_link = '/cargo/{awb_number}/'
    elif entity == 'hawb':
        db_fields.update({'id', 'hawb_number'})
        row_link = '/hawb/{id}/'

    total = qs.count()
    # ordering — по created_at ↓ если есть, иначе по pk
    order_field = '-created_at' if 'created_at' in {f.name for f in model._meta.get_fields() if hasattr(f, 'name')} else '-pk'
    try:
        items_qs = qs.order_by(order_field)
    except Exception:
        logger.exception('drill: cannot order by %s', order_field)
        items_qs = qs

    if db_fields:
        items = list(items_qs[offset:offset + page_size].values(*db_fields))
    else:
        items = list(items_qs[offset:offset + page_size].values())
    for it in items:
        for k, v in list(it.items()):
            if isinstance(v, _dt.datetime):
                it[k] = v.strftime('%d.%m.%Y %H:%M')
            elif isinstance(v, _dt.date):
                it[k] = v.strftime('%d.%m.%Y')

    list_url_name = spec.get('list_url')
    list_url = ''
    if list_url_name:
        try:
            from django.urls import reverse
            list_url = reverse(list_url_name)
        except Exception:
            logger.warning('drill: cannot reverse list_url %s', list_url_name, exc_info=True)
            list_url = ''

    return JsonResponse({
        'entity':       entity,
        'entity_label': spec['label'],
        'columns':      [serialize_column(c) for c in catalog],
        'items':        items,
        'total':        total,
        'page':         page,
        'page_size':    page_size,
        'has_more':     offset + len(items) < total,
        'row_link':     row_link,
        'list_url':     list_url,
        'cql':          raw_cql,
        'cql_ast':      ast,
    })


@login_required
def drill_view(request):
    """Рендерит универсальную drill-down страницу. Данные подтягиваются клиентом
    из `/api/v1/widgets/drill/`."""
    entity = (request.GET.get('entity') or '').strip()
    if entity and entity not in ENTITY_REGISTRY:
        return HttpResponse(f'Unknown entity: {entity!r}', status=400)
    ctx = {
        'entity':       entity,
        'entity_label': (ENTITY_REGISTRY.get(entity) or {}).get('label', ''),
        'cql':          request.GET.get('cql', '') or '',
        'widget_id':    request.GET.get('widget_id', '') or '',
        'widget_title': request.GET.get('widget_title', '') or '',
        'entities':     get_entities(),
    }
    return render(request, 'cargo/drill.html', ctx)


@login_required
@require_http_methods(['GET'])
def api_dashboard_pivot_drill(request, widget_id):
    """GET /api/v1/dashboard/widgets/<id>/drill/
        ?row_key=…&col_key=…&metric_idx=…&limit=50&offset=0

    Возвращает список записей, попавших в указанную ячейку pivot-виджета.
    """
    from .widget_columns import get_column_catalog, serialize_column

    try:
        widget = DashboardWidget.objects.get(id=widget_id, user=request.user)
    except DashboardWidget.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    if widget.widget_type != 'pivot':
        return JsonResponse({'error': 'Drill доступен только для pivot-виджета'}, status=400)

    config = widget.config if isinstance(widget.config, dict) else {}
    row_by = config.get('row_by')
    col_by = config.get('col_by') or None
    metrics_cfg = config.get('metrics') or []
    if not metrics_cfg and config.get('agg'):
        metrics_cfg = [{
            'name':   config.get('metric_name', 'Значение'),
            'agg':    config.get('agg'),
            'field':  config.get('field'),
            'filter': config.get('metric_filter', ''),
            'format': config.get('format', 'int'),
        }]
    if not row_by or not metrics_cfg:
        return JsonResponse({'error': 'Виджет не сконфигурирован'}, status=400)

    row_key = request.GET.get('row_key', '')
    col_key = request.GET.get('col_key', '__all__')
    try:
        metric_idx = int(request.GET.get('metric_idx', 0))
    except (TypeError, ValueError):
        metric_idx = 0
    if not (0 <= metric_idx < len(metrics_cfg)):
        return JsonResponse({'error': 'Неверный metric_idx'}, status=400)
    try:
        limit  = max(1, min(int(request.GET.get('limit', 50)), 200))
        offset = max(0, int(request.GET.get('offset', 0)))
    except (TypeError, ValueError):
        limit, offset = 50, 0

    entity = widget.entity_type
    try:
        base_q = parse_cql(widget.filter_query, {'me': request.user.username},
                           entity_type=entity)
    except CQLError as e:
        return JsonResponse({'error': f'CQL: {e}'}, status=400)

    if entity == 'hawb':
        from .models import HouseWaybill
        qs = HouseWaybill.objects.filter(base_q).select_related('mawb', 'assigned_to').distinct()
    else:
        qs = Cargo.objects.filter(base_q).select_related('warehouse').distinct()

    # Row filter
    row_def, row_is_standalone = _resolve_group(row_by, entity)
    if not row_def:
        return JsonResponse({'error': f'Недопустимое поле строк: {row_by!r}'}, status=400)
    qs = qs.filter(_pivot_group_filter(row_def, row_is_standalone, row_key))

    # Col filter (пропускаем если col_by не задан или col_key == '__all__')
    col_def = None
    col_is_standalone = False
    if col_by and col_key != '__all__':
        col_def, col_is_standalone = _resolve_group(col_by, entity)
        if not col_def:
            return JsonResponse({'error': f'Недопустимое поле колонок: {col_by!r}'}, status=400)
        qs = qs.filter(_pivot_group_filter(col_def, col_is_standalone, col_key))
    elif col_by:
        col_def, col_is_standalone = _resolve_group(col_by, entity)

    # Metric sub-filter: для ratio используем фильтр числителя
    metric = metrics_cfg[metric_idx]
    metric_name = metric.get('name') or f'Метрика {metric_idx + 1}'
    if (metric.get('agg') or '').lower() == 'ratio':
        sub_filter = ((metric.get('numerator') or {}).get('filter') or '').strip()
    else:
        sub_filter = (metric.get('filter') or '').strip()
    if sub_filter:
        try:
            qs = qs.filter(parse_cql(sub_filter, {'me': request.user.username},
                                     entity_type=entity))
        except CQLError as e:
            return JsonResponse({'error': f'Фильтр метрики: {e}'}, status=400)

    # Columns — отдаём весь каталог (фронт может скрыть ненужные)
    catalog = get_column_catalog(entity)
    db_fields = set()
    for col in catalog:
        db_fields.update(col['db_fields'])
    if entity == 'cargo':
        db_fields.add('awb_number')
        row_link = '/cargo/{awb_number}/'
    else:
        db_fields.update({'id', 'hawb_number'})
        row_link = '/hawb/{id}/'

    total = qs.count()
    items = list(qs.order_by('-created_at')[offset:offset + limit].values(*db_fields))
    for it in items:
        for k, v in list(it.items()):
            if isinstance(v, _dt.datetime):
                it[k] = v.strftime('%d.%m.%Y %H:%M')
            elif isinstance(v, _dt.date):
                it[k] = v.strftime('%d.%m.%Y')

    def _lookup_label(gdef, key):
        if key == _PIVOT_MISSING:
            return '(не указано)'
        ch = gdef.get('choices')
        if ch is not None:
            for rk, lbl in ch.items():
                if str(rk) == str(key):
                    return str(lbl)
        if 'choices_fn' in gdef:
            for rk, lbl in gdef['choices_fn']().items():
                if str(rk) == str(key):
                    return str(lbl)
        return str(key)

    if row_is_standalone:
        row_label = 'Без партии' if str(row_key) == 'True' else 'Внутри партии'
    else:
        row_label = _lookup_label(row_def, row_key)

    col_label = None
    if col_by and col_key != '__all__':
        if col_is_standalone:
            col_label = 'Без партии' if str(col_key) == 'True' else 'Внутри партии'
        else:
            col_label = _lookup_label(col_def, col_key)

    return JsonResponse({
        'columns':       [serialize_column(c) for c in catalog],
        'items':         items,
        'total':         total,
        'has_more':      offset + len(items) < total,
        'row_label':     row_label,
        'col_label':     col_label,
        'metric_name':   metric_name,
        'row_link':      row_link,
        'row_by_label':  row_def.get('label', row_by),
        'col_by_label':  (col_def.get('label', col_by) if col_def else None),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# WORKFLOW / BUSINESS PROCESS
# ═══════════════════════════════════════════════════════════════════════════════

from .models import Workflow, WorkflowStep, WorkflowTransition, AutomationRule


@login_required
def workflow_list(request):
    """Страница со списком бизнес-процессов."""
    return render(request, 'cargo/workflow_list.html')


@login_required
def workflow_editor(request, workflow_id):
    """Визуальный редактор бизнес-процесса."""
    wf = get_object_or_404(Workflow, id=workflow_id)
    return render(request, 'cargo/workflow_editor.html', {'workflow': wf})


# ── Workflow CRUD API ───────────────────────────────────────────────────────

def _step_to_dict(s):
    return {
        'id': s.id, 'name': s.name, 'step_type': s.step_type,
        'stage': s.stage,
        'hawb_logistics_status': s.hawb_logistics_status,
        'hawb_customs_status': s.hawb_customs_status,
        'config': s.config,
        'pos_x': s.pos_x, 'pos_y': s.pos_y,
        'color': s.color, 'order': s.order,
        'sla_hours_override': float(s.sla_hours_override) if s.sla_hours_override is not None else None,
    }


def _transition_to_dict(t):
    return {
        'id': t.id, 'from_step': t.from_step_id, 'to_step': t.to_step_id,
        'label': t.label, 'condition': t.condition, 'order': t.order,
    }


def _automation_to_dict(a):
    return {
        'id': a.id, 'transition': a.transition_id,
        'name': a.name, 'action_type': a.action_type,
        'config': a.config, 'is_active': a.is_active, 'order': a.order,
    }


def _workflow_to_dict(wf, full=False):
    d = {
        'id': wf.id, 'name': wf.name, 'description': wf.description,
        'is_active': wf.is_active,
        'entity_type': wf.entity_type,
        'entity_type_display': wf.get_entity_type_display(),
        'trigger_conditions': wf.trigger_conditions,
        'auto_start': wf.auto_start,
        'created_by': wf.created_by.get_full_name() if wf.created_by else '',
        'updated_at': wf.updated_at.strftime('%d.%m.%Y %H:%M'),
        'steps_count': getattr(wf, '_steps_count', None) if getattr(wf, '_steps_count', None) is not None else wf.steps.count(),
    }
    if full:
        d['steps']       = [_step_to_dict(s) for s in wf.steps.all()]
        d['transitions'] = [_transition_to_dict(t) for t in wf.transitions.all()]
        d['automations'] = [_automation_to_dict(a) for a in wf.automations.all()]
    return d


@login_required
@require_http_methods(['GET', 'POST'])
def api_workflows(request):
    """GET — список, POST — создать."""
    if request.method == 'GET':
        wfs = (
            Workflow.objects
            .prefetch_related('steps', 'transitions', 'automations')
            .annotate(_steps_count=Count('steps'))
            .order_by('id')
        )
        items, page = _paginate(request, wfs)
        return JsonResponse({
            'workflows': [_workflow_to_dict(w) for w in items],
            'pagination': page,
        })

    data, err = _parse_json_body(request)
    if err: return err
    wf = Workflow.objects.create(
        name=data.get('name', 'Новый процесс'),
        description=data.get('description', ''),
        entity_type=data.get('entity_type', 'cargo'),
        trigger_conditions=data.get('trigger_conditions', {}),
        auto_start=data.get('auto_start', True),
        created_by=request.user,
    )
    return JsonResponse(_workflow_to_dict(wf, full=True), status=201)


@login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def api_workflow(request, workflow_id):
    try:
        wf = Workflow.objects.get(id=workflow_id)
    except Workflow.DoesNotExist:
        return JsonResponse({'error': 'Не найден'}, status=404)

    if request.method == 'GET':
        return JsonResponse(_workflow_to_dict(wf, full=True))

    if request.method == 'DELETE':
        wf.delete()
        return JsonResponse({'ok': True})

    data, err = _parse_json_body(request)
    if err: return err
    wf.name = data.get('name', wf.name)
    wf.description = data.get('description', wf.description)
    wf.is_active = data.get('is_active', wf.is_active)
    wf.entity_type = data.get('entity_type', wf.entity_type)
    wf.trigger_conditions = data.get('trigger_conditions', wf.trigger_conditions)
    wf.auto_start = data.get('auto_start', wf.auto_start)
    wf.save()
    return JsonResponse(_workflow_to_dict(wf, full=True))


# ── Steps API ──────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['POST'])
def api_workflow_steps(request, workflow_id):
    """Создать шаг."""
    try:
        wf = Workflow.objects.get(id=workflow_id)
    except Workflow.DoesNotExist:
        return JsonResponse({'error': 'Процесс не найден'}, status=404)

    data, err = _parse_json_body(request)
    if err: return err
    sla_override = data.get('sla_hours_override')
    if sla_override in ('', None):
        sla_override = None
    step = WorkflowStep.objects.create(
        workflow=wf,
        name=data.get('name', 'Новый шаг'),
        step_type=data.get('step_type', 'stage'),
        stage=data.get('stage', ''),
        hawb_logistics_status=data.get('hawb_logistics_status', ''),
        hawb_customs_status=data.get('hawb_customs_status', ''),
        config=data.get('config', {}),
        pos_x=data.get('pos_x', 100),
        pos_y=data.get('pos_y', 100),
        color=data.get('color', '#3b82f6'),
        order=data.get('order', 0),
        sla_hours_override=sla_override,
    )
    return JsonResponse(_step_to_dict(step), status=201)


@login_required
@require_http_methods(['PUT', 'DELETE'])
def api_workflow_step(request, workflow_id, step_id):
    try:
        step = WorkflowStep.objects.get(id=step_id, workflow_id=workflow_id)
    except WorkflowStep.DoesNotExist:
        return JsonResponse({'error': 'Шаг не найден'}, status=404)

    if request.method == 'DELETE':
        step.delete()
        return JsonResponse({'ok': True})

    data, err = _parse_json_body(request)
    if err: return err
    for field in ('name', 'step_type', 'stage', 'hawb_logistics_status', 'hawb_customs_status', 'config', 'pos_x', 'pos_y', 'color', 'order'):
        if field in data:
            setattr(step, field, data[field])
    if 'sla_hours_override' in data:
        v = data['sla_hours_override']
        step.sla_hours_override = None if v in ('', None) else v
    step.save()
    return JsonResponse(_step_to_dict(step))


# ── Transitions API ────────────────────────────────────────────────────────

@login_required
@require_http_methods(['POST'])
def api_workflow_transitions(request, workflow_id):
    """Создать переход."""
    try:
        wf = Workflow.objects.get(id=workflow_id)
    except Workflow.DoesNotExist:
        return JsonResponse({'error': 'Процесс не найден'}, status=404)

    data, err = _parse_json_body(request)
    if err: return err
    try:
        from_step = WorkflowStep.objects.get(id=data['from_step'], workflow=wf)
        to_step   = WorkflowStep.objects.get(id=data['to_step'],   workflow=wf)
    except (WorkflowStep.DoesNotExist, KeyError):
        return JsonResponse({'error': 'Шаг не найден'}, status=400)

    condition = data.get('condition', '')
    if condition:
        from cargo.cql_parser import parse_cql, CQLError as _CQLError
        try:
            parse_cql(condition, entity_type=wf.entity_type)
        except _CQLError as e:
            return JsonResponse({'error': f'CQL-условие: {e}'}, status=400)

    tr = WorkflowTransition.objects.create(
        workflow=wf, from_step=from_step, to_step=to_step,
        label=data.get('label', ''),
        condition=condition,
    )
    return JsonResponse(_transition_to_dict(tr), status=201)


@login_required
@require_http_methods(['PUT', 'DELETE'])
def api_workflow_transition(request, workflow_id, transition_id):
    try:
        tr = WorkflowTransition.objects.get(id=transition_id, workflow_id=workflow_id)
    except WorkflowTransition.DoesNotExist:
        return JsonResponse({'error': 'Переход не найден'}, status=404)

    if request.method == 'DELETE':
        tr.delete()
        return JsonResponse({'ok': True})

    data, err = _parse_json_body(request)
    if err: return err
    if 'condition' in data and data['condition']:
        from cargo.cql_parser import parse_cql, CQLError as _CQLError
        try:
            parse_cql(data['condition'], entity_type=tr.workflow.entity_type)
        except _CQLError as e:
            return JsonResponse({'error': f'CQL-условие: {e}'}, status=400)
    for field in ('label', 'condition', 'order'):
        if field in data:
            setattr(tr, field, data[field])
    tr.save()
    return JsonResponse(_transition_to_dict(tr))


# ── Automations API ────────────────────────────────────────────────────────

@login_required
@require_http_methods(['POST'])
def api_workflow_automations(request, workflow_id):
    """Создать правило автоматизации."""
    try:
        wf = Workflow.objects.get(id=workflow_id)
    except Workflow.DoesNotExist:
        return JsonResponse({'error': 'Процесс не найден'}, status=404)

    data, err = _parse_json_body(request)
    if err: return err
    rule = AutomationRule.objects.create(
        workflow=wf,
        transition_id=data.get('transition'),
        name=data.get('name', 'Новое правило'),
        action_type=data.get('action_type', 'assign_user'),
        config=data.get('config', {}),
        is_active=data.get('is_active', True),
    )
    return JsonResponse(_automation_to_dict(rule), status=201)


@login_required
@require_http_methods(['PUT', 'DELETE'])
def api_workflow_automation(request, workflow_id, automation_id):
    try:
        rule = AutomationRule.objects.get(id=automation_id, workflow_id=workflow_id)
    except AutomationRule.DoesNotExist:
        return JsonResponse({'error': 'Правило не найдено'}, status=404)

    if request.method == 'DELETE':
        rule.delete()
        return JsonResponse({'ok': True})

    data, err = _parse_json_body(request)
    if err: return err
    for field in ('name', 'action_type', 'config', 'is_active', 'transition', 'order'):
        if field in data:
            if field == 'transition':
                rule.transition_id = data[field]
            else:
                setattr(rule, field, data[field])
    rule.save()
    return JsonResponse(_automation_to_dict(rule))


# ── Batch save (positions of all steps) ────────────────────────────────────

@login_required
@require_http_methods(['POST'])
def api_workflow_save_layout(request, workflow_id):
    """Сохранить позиции всех шагов одним запросом."""
    try:
        wf = Workflow.objects.get(id=workflow_id)
    except Workflow.DoesNotExist:
        return JsonResponse({'error': 'Процесс не найден'}, status=404)

    items, err = _parse_json_body(request)
    if err: return err
    for item in items:
        WorkflowStep.objects.filter(id=item['id'], workflow=wf).update(
            pos_x=item.get('x', 0), pos_y=item.get('y', 0),
        )
    return JsonResponse({'ok': True})


# ─────────────────────────── WORKFLOW INSTANCES API ─────────────────────────

@login_required
@require_http_methods(['GET'])
def api_cargo_workflow_instances(request, awb_number):
    """Список инстансов воркфлоу для партии."""
    from . import workflow_runner
    cargo = get_object_or_404(Cargo, awb_number=awb_number)
    instances = workflow_runner.get_instances_for_entity('cargo', cargo.pk)
    return JsonResponse({'instances': [workflow_runner.serialize_instance(i) for i in instances]})


@login_required
@require_http_methods(['GET'])
def api_hawb_workflow_instances(request, hawb_id):
    """Список инстансов воркфлоу для накладной."""
    from .models import HouseWaybill
    from . import workflow_runner
    hawb = get_object_or_404(HouseWaybill, pk=hawb_id)
    instances = workflow_runner.get_instances_for_entity('hawb', hawb.pk)
    return JsonResponse({'instances': [workflow_runner.serialize_instance(i) for i in instances]})


@login_required
@require_http_methods(['POST'])
def api_workflow_start(request, workflow_id):
    """
    Ручной запуск воркфлоу на сущности.
    Body: {"entity_type": "hawb"|"cargo", "entity_id": <int>}
    """
    from .models import HouseWaybill
    from . import workflow_runner
    data, err = _parse_json_body(request)
    if err: return err
    entity_type = data.get('entity_type', '')
    entity_id = data.get('entity_id')

    if entity_type not in ('cargo', 'hawb') or not entity_id:
        return JsonResponse({'error': 'Укажите entity_type и entity_id'}, status=400)

    if entity_type == 'cargo':
        entity = get_object_or_404(Cargo, pk=entity_id)
    else:
        entity = get_object_or_404(HouseWaybill, pk=entity_id)

    instance, error = workflow_runner.start_workflow_manually(
        workflow_id, entity, entity_type, user=request.user
    )
    if error:
        return JsonResponse({'error': error}, status=400)
    return JsonResponse({'instance': workflow_runner.serialize_instance(instance)}, status=201)


@login_required
@require_http_methods(['POST'])
def api_workflow_instance_cancel(request, instance_id):
    """Отмена инстанса воркфлоу."""
    from . import workflow_runner
    instance, error = workflow_runner.cancel_instance(instance_id, user=request.user)
    if error:
        return JsonResponse({'error': error}, status=400)
    return JsonResponse({'instance': workflow_runner.serialize_instance(instance)})


# ============================================================================
# SLA POLICIES
# ============================================================================

from django.db import IntegrityError  # noqa: E402
from .models import SLAPolicy, HouseWaybill as _HouseWaybill  # noqa: E402


def _sla_policy_to_dict(p):
    return {
        'id': p.id,
        'name': p.name,
        'entity_type': p.entity_type,
        'entity_type_display': p.get_entity_type_display(),
        'status_field': p.status_field,
        'status_field_display': p.get_status_field_display(),
        'status_value': p.status_value,
        'hours': float(p.hours),
        'warning_threshold_pct': p.warning_threshold_pct,
        'is_active': p.is_active,
        'updated_at': p.updated_at.strftime('%d.%m.%Y %H:%M'),
    }


def _sla_status_choices():
    """Маппинг status_field → список [{value,label}] для выпадающих списков."""
    return {
        'stage': [{'value': v, 'label': l} for v, l in STAGE_CHOICES],
        'logistics_status': [{'value': v, 'label': l} for v, l in _HouseWaybill.LOGISTICS_STATUS_CHOICES],
        'customs_status':   [{'value': v, 'label': l} for v, l in _HouseWaybill.CUSTOMS_STATUS_CHOICES],
    }


@login_required
def sla_policies_page(request):
    """Страница администрирования SLA-нормативов."""
    return render(request, 'cargo/sla_admin.html', {
        'status_choices': json.dumps(_sla_status_choices(), ensure_ascii=False),
    })


@login_required
@require_http_methods(['GET', 'POST'])
def api_sla_policies(request):
    if request.method == 'GET':
        entity_type = request.GET.get('entity_type')
        qs = SLAPolicy.objects.all().order_by('id')
        if entity_type:
            qs = qs.filter(entity_type=entity_type)
        items, page = _paginate(request, qs)
        return JsonResponse({
            'policies': [_sla_policy_to_dict(p) for p in items],
            'pagination': page,
        })

    data, err = _parse_json_body(request)
    if err: return err
    try:
        policy = SLAPolicy.objects.create(
            name=data.get('name', ''),
            entity_type=data['entity_type'],
            status_field=data['status_field'],
            status_value=data['status_value'],
            hours=data['hours'],
            warning_threshold_pct=int(data.get('warning_threshold_pct', 75)),
            is_active=bool(data.get('is_active', True)),
        )
    except KeyError as e:
        return JsonResponse({'error': f'Не задано поле: {e}'}, status=400)
    except IntegrityError:
        return JsonResponse({'error': 'Политика для этого (сущность/поле/значение) уже существует'}, status=400)
    from .sla import invalidate_policy_cache
    invalidate_policy_cache()
    return JsonResponse(_sla_policy_to_dict(policy), status=201)


@login_required
@require_http_methods(['PUT', 'DELETE'])
def api_sla_policy(request, policy_id):
    try:
        policy = SLAPolicy.objects.get(id=policy_id)
    except SLAPolicy.DoesNotExist:
        return JsonResponse({'error': 'Не найдено'}, status=404)

    if request.method == 'DELETE':
        policy.delete()
        from .sla import invalidate_policy_cache
        invalidate_policy_cache()
        return JsonResponse({'ok': True})

    data, err = _parse_json_body(request)
    if err: return err
    for field in ('name', 'entity_type', 'status_field', 'status_value',
                  'hours', 'warning_threshold_pct', 'is_active'):
        if field in data:
            setattr(policy, field, data[field])
    try:
        policy.save()
    except IntegrityError:
        return JsonResponse({'error': 'Политика для этого (сущность/поле/значение) уже существует'}, status=400)
    from .sla import invalidate_policy_cache
    invalidate_policy_cache()
    return JsonResponse(_sla_policy_to_dict(policy))


# ─────────────────────────── ИМПОРТ GOOGLE SHEETS ───────────────────────────

@login_required
def sheets_imports_page(request):
    """Список ImportedSheetRow с фильтрами и сводкой по статусам."""
    from .models import ImportedSheetRow, SheetSource, SheetImportRun
    qs = ImportedSheetRow.objects.select_related(
        'source', 'matched_hawb', 'matched_cargo'
    ).order_by('source__kind', 'source__name', 'source_row_index')

    f_source = request.GET.get('source') or ''
    f_status = request.GET.get('status') or ''
    f_from   = (request.GET.get('from') or '').strip()
    f_to     = (request.GET.get('to') or '').strip()
    q        = (request.GET.get('q') or '').strip()
    if f_source:
        qs = qs.filter(source_id=f_source)
    if f_status:
        qs = qs.filter(match_status=f_status)
    if f_from:
        from datetime import datetime
        try:
            qs = qs.filter(arrival_date__gte=datetime.strptime(f_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if f_to:
        from datetime import datetime
        try:
            qs = qs.filter(arrival_date__lte=datetime.strptime(f_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    if q:
        qs = qs.filter(
            Q(hawb_number_norm__icontains=q) |
            Q(hawb_number_raw__icontains=q) |
            Q(inn_raw__icontains=q) |
            Q(declaration_number__icontains=q)
        )

    summary_map = {
        row['match_status']: row['n']
        for row in ImportedSheetRow.objects.values('match_status').annotate(n=Count('id'))
    }
    status_summary = [
        (code, label, summary_map.get(code, 0))
        for code, label in ImportedSheetRow.MATCH_STATUS_CHOICES
    ]

    sources = SheetSource.objects.order_by('kind', 'name')
    last_run = SheetImportRun.objects.order_by('-started_at').first()
    visible_qs = qs[:500]

    return render(request, 'cargo/imports/sheets_index.html', {
        'rows': visible_qs,
        'sources': sources,
        'f_source': f_source,
        'f_status': f_status,
        'f_from': f_from,
        'f_to': f_to,
        'q': q,
        'status_summary': status_summary,
        'total_count': ImportedSheetRow.objects.count(),
        'filtered_count': qs.count(),
        'last_run': last_run,
        'status_choices': ImportedSheetRow.MATCH_STATUS_CHOICES,
    })


@login_required
def sheets_row_detail(request, row_id: int):
    """Подробный вид одной строки с diff Sheets / БД."""
    from .models import ImportedSheetRow
    row = get_object_or_404(
        ImportedSheetRow.objects.select_related(
            'source', 'matched_hawb', 'matched_cargo', 'promoted_hawb'
        ),
        pk=row_id,
    )
    events = row.emitted_events.all().order_by('-occurred_at') if row.matched_hawb_id else []
    diff = row.diff_summary or {}
    candidates = diff.get('_candidates') or []
    reason = diff.get('_reason') or ''
    diff_fields = {k: v for k, v in diff.items() if not k.startswith('_')}
    return render(request, 'cargo/imports/sheets_row_detail.html', {
        'row': row,
        'events': events,
        'candidates': candidates,
        'reason': reason,
        'diff_fields': diff_fields,
    })


@login_required
@require_POST
def sheets_row_promote(request, row_id: int):
    """Создать HAWB из orphan-строки и связать с promoted_hawb."""
    from .models import ImportedSheetRow
    from .services.sheets.promote import promote_row
    row = get_object_or_404(ImportedSheetRow, pk=row_id)
    try:
        hawb = promote_row(row, user=request.user)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('sheets_row_detail', row_id=row.pk)
    messages.success(request, f'Создана HAWB {hawb.hawb_number} из строки #{row.source_row_index}.')
    return redirect('hawb_detail', hawb_id=hawb.pk)


@login_required
@require_POST
def sheets_bulk_promote(request):
    """Пакетный promote: создать HAWB по списку id из чекбоксов."""
    from django.db import transaction
    from .models import ImportedSheetRow
    from .services.sheets.promote import promote_row
    ids = request.POST.getlist('row_ids')
    if not ids:
        messages.warning(request, 'Не выбрана ни одна строка.')
        return redirect('sheets_imports')

    created = 0
    skipped = 0
    errors: list[str] = []
    qs = ImportedSheetRow.objects.filter(pk__in=ids, match_status='orphan', source__kind='general')
    with transaction.atomic():
        for row in qs.select_related('source'):
            try:
                promote_row(row, user=request.user)
                created += 1
            except ValueError as e:
                skipped += 1
                errors.append(f'#{row.source_row_index}: {e}')
    if created:
        messages.success(request, f'Создано HAWB: {created}.')
    if skipped:
        messages.warning(
            request,
            f'Пропущено: {skipped}. ' + ('Первые: ' + '; '.join(errors[:3]) if errors else '')
        )
    return redirect('sheets_imports')


@login_required
@require_POST
def sheets_row_ignore(request, row_id: int):
    """Пометить строку как ignored, чтобы не светилась в списках."""
    from .models import ImportedSheetRow
    row = get_object_or_404(ImportedSheetRow, pk=row_id)
    row.match_status = 'ignored'
    row.save(update_fields=['match_status'])
    messages.success(request, f'Строка #{row.source_row_index} помечена как игнорируемая.')
    return redirect('sheets_imports')


@login_required
@require_POST
def sheets_row_rematch(request, row_id: int):
    """Перематчить одну строку (после правки HAWB вручную или создания alias)."""
    from .models import ImportedSheetRow
    from .services.sheets.matcher import match_row
    from .services.sheets.events import emit_workflow_events
    row = get_object_or_404(ImportedSheetRow, pk=row_id)
    match_row(row)
    row.save()
    if row.source.kind == 'crm' and row.matched_hawb_id:
        emit_workflow_events(row)
    messages.success(request, f'Перематчено. Новый статус: {row.get_match_status_display()}.')
    return redirect('sheets_row_detail', row_id=row.pk)


@login_required
@require_POST
def sheets_run_now(request, source_id: int):
    """Синхронно запустить импорт одного источника."""
    from .models import SheetSource
    from .services.sheets.importer import SheetImporter
    source = get_object_or_404(SheetSource, pk=source_id)
    importer = SheetImporter(source, user=request.user)
    run = importer.run_once()
    if run.status == 'ok':
        messages.success(
            request,
            f'Импорт {source.name} OK: total={run.rows_total} '
            f'matched={run.rows_matched} orphan={run.rows_orphan} '
            f'conflict={run.rows_conflict}.'
        )
    else:
        messages.error(request, f'Импорт {source.name} упал: {run.error_message[:300]}')
    return redirect('sheets_runs')


@login_required
def sheets_runs_page(request):
    """История прогонов."""
    from .models import SheetImportRun
    runs = (SheetImportRun.objects
            .select_related('source', 'triggered_by')
            .order_by('-started_at')[:200])
    return render(request, 'cargo/imports/sheets_runs.html', {'runs': runs})


@login_required
def sheets_sources_page(request):
    """Список источников с кнопкой запуска. CRUD — через /admin/."""
    from .models import SheetSource
    sources = SheetSource.objects.order_by('kind', 'name')
    return render(request, 'cargo/imports/sheets_sources.html', {'sources': sources})
